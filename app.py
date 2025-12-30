import streamlit as st
import streamlit_authenticator as stauth
import streamlit.components.v1 as components
import yaml
from yaml.loader import SafeLoader
import json
import pandas as pd
from google import genai
from google.genai import types
import os
from dotenv import load_dotenv
import tempfile
import zipfile
import time
import re
from typing import List, Dict, Set , Tuple
from collections import defaultdict
from io import BytesIO
import logging
from datetime import datetime
import traceback
import matplotlib.ticker as mticker
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import seaborn as sns
import arabic_reshaper
from bidi.algorithm import get_display
from matplotlib.colors import LinearSegmentedColormap, BoundaryNorm
import plotly.express as px
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import math
from datetime import datetime, timedelta
import base64
from matplotlib import font_manager
import numpy as np
# ============================================================================
# Ø¨Ø®Ø´ 1: ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø§ÙˆÙ„ÛŒÙ‡
# ============================================================================

load_dotenv()

def load_css(file_name):
    try:
        with open(file_name, "r", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        st.error(f"ÙØ§ÛŒÙ„ '{file_name}' Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app_log.txt', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

st.set_page_config(
    page_title="AI Financial Analyzer",
    page_icon="ğŸ“Š",
    layout="wide",
    initial_sidebar_state="expanded"
)

load_css("style.css")


def load_api_keys_from_env():
    """
    Ø§ÛŒÙ† ØªØ§Ø¨Ø¹ ØªÙ„Ø§Ø´ Ù…ÛŒâ€ŒÚ©Ù†Ø¯ Ú©Ù„ÛŒØ¯Ù‡Ø§ Ø±Ø§ Ø§Ø² Ù…ØªØºÛŒØ±Ù‡Ø§ÛŒ Ù…Ø­ÛŒØ·ÛŒ Ø¨Ø®ÙˆØ§Ù†Ø¯.
    Ø¯Ùˆ Ø±ÙˆØ´ Ø±Ø§ Ù¾Ø´ØªÛŒØ¨Ø§Ù†ÛŒ Ù…ÛŒâ€ŒÚ©Ù†Ø¯:
    1. ÛŒÚ© Ù…ØªØºÛŒØ± Ø¨Ù‡ Ù†Ø§Ù… GOOGLE_API_KEYS Ú©Ù‡ Ú©Ù„ÛŒØ¯Ù‡Ø§ Ø¨Ø§ Ú©Ø§Ù…Ø§ Ø¬Ø¯Ø§ Ø´Ø¯Ù‡â€ŒØ§Ù†Ø¯.
    2. Ù…ØªØºÛŒØ±Ù‡Ø§ÛŒ Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡ Ù…Ø«Ù„ API_KEY_1, API_KEY_2 Ùˆ ...
    """
    keys = []
    
    # Ø±ÙˆØ´ Û±: Ø®ÙˆØ§Ù†Ø¯Ù† Ù‡Ù…Ù‡ Ú©Ù„ÛŒØ¯Ù‡Ø§ Ø§Ø² ÛŒÚ© Ù…ØªØºÛŒØ± Ø±Ø´ØªÙ‡â€ŒØ§ÛŒ (Ù…Ù†Ø§Ø³Ø¨ Ø¨Ø±Ø§ÛŒ Ú©Ù¾ÛŒ Ø±Ø§Ø­Øª Ø¯Ø± Ú¯ÛŒØªâ€ŒÙ‡Ø§Ø¨)
    # Ù…Ø«Ø§Ù„ Ù…Ù‚Ø¯Ø§Ø±: key1,key2,key3
    all_keys_string = os.getenv("GOOGLE_API_KEYS")
    if all_keys_string:
        # Ø¬Ø¯Ø§ Ú©Ø±Ø¯Ù† Ø¨Ø§ Ú©Ø§Ù…Ø§ Ùˆ Ø­Ø°Ù ÙØ§ØµÙ„Ù‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¶Ø§ÙÛŒ
        keys = [k.strip() for k in all_keys_string.split(',') if k.strip()]
        
    # Ø±ÙˆØ´ Û²: Ø§Ú¯Ø± Ø±ÙˆØ´ Ø§ÙˆÙ„ Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ØŒ Ø¯Ù†Ø¨Ø§Ù„ Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡ Ø¨Ú¯Ø±Ø¯
    if not keys:
        i = 1
        while True:
            key = os.getenv(f"API_KEY_{i}") # Ù…Ø«Ù„Ø§ API_KEY_1
            if key:
                keys.append(key)
                i += 1
            else:
                break
    
    return keys

# Ø¯Ø±ÛŒØ§ÙØª Ú©Ù„ÛŒØ¯Ù‡Ø§
loaded_keys = load_api_keys_from_env()

# Ø§Ú¯Ø± Ú©Ù„ÛŒØ¯ÛŒ Ø¯Ø± Ù…Ø­ÛŒØ· Ù†Ø¨ÙˆØ¯ØŒ ÛŒÚ© Ù„ÛŒØ³Øª Ø®Ø§Ù„ÛŒ ÛŒØ§ ÛŒÚ© Ù¾ÛŒØ§Ù… Ø®Ø·Ø§ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†
if loaded_keys:
    DEFAULT_API_KEYS = loaded_keys
else:
    # Ø§ÛŒÙ† Ù‚Ø³Ù…Øª ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² Ú©Ø±Ø´ Ú©Ø±Ø¯Ù† Ø¨Ø±Ù†Ø§Ù…Ù‡ Ø¯Ø± Ø­Ø§Ù„Øª Ø¨Ø¯ÙˆÙ† Ú©Ù„ÛŒØ¯ Ø§Ø³Øª
    DEFAULT_API_KEYS = [] 
    # ÛŒØ§ Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒØ¯ Ø®Ø·Ø§ Ø¯Ù‡ÛŒØ¯: raise ValueError("Ù‡ÛŒÚ† API Key ÛŒØ§ÙØª Ù†Ø´Ø¯!")

if 'api_keys' not in st.session_state:
    st.session_state.api_keys = DEFAULT_API_KEYS.copy()

# ============================================================================
# Ø¨Ø®Ø´ 2: Ø§Ø­Ø±Ø§Ø² Ù‡ÙˆÛŒØª
# ============================================================================

with open('config.yaml') as file:
    config = yaml.load(file, Loader=SafeLoader)

hashed_passwords = stauth.Hasher(["elnagh", "abc_fin_cba","123"]).generate()
config['credentials']['usernames']['admin']['password'] = hashed_passwords[0]
config['credentials']['usernames']['fin.analyst']['password'] = hashed_passwords[1]
config['credentials']['usernames']['h.khandani']['password'] = hashed_passwords[2]

authenticator = stauth.Authenticate(
    config['credentials'],
    config['cookie']['name'],
    config['cookie']['key'],
    config['cookie']['expiry_days'],
)

if 'authentication_status' not in st.session_state:
    st.session_state.authentication_status = None

if st.session_state.authentication_status is None:
    name, authentication_status, username = authenticator.login(location='main')
    st.session_state.authentication_status = authentication_status
    st.session_state.name = name
    st.session_state.username = username

if st.session_state.authentication_status == False:
    st.error('Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ ÛŒØ§ Ø±Ù…Ø² Ø¹Ø¨ÙˆØ± Ø§Ø´ØªØ¨Ø§Ù‡ Ø§Ø³Øª')
    st.stop()

if st.session_state.authentication_status == None:
    st.warning('Ù„Ø·ÙØ§Ù‹ Ù†Ø§Ù… Ú©Ø§Ø±Ø¨Ø±ÛŒ Ùˆ Ø±Ù…Ø² Ø¹Ø¨ÙˆØ± Ø®ÙˆØ¯ Ø±Ø§ ÙˆØ§Ø±Ø¯ Ú©Ù†ÛŒØ¯')
    st.stop()

# ============================================================================
# Ø¨Ø®Ø´ 3: Ø³Ø§ÛŒØ¯Ø¨Ø§Ø±
# ============================================================================

if st.session_state.authentication_status:
    
    with st.sidebar:
        st.header("âš™ï¸ ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø¨Ø±Ù†Ø§Ù…Ù‡")
        # st.divider()
        # st.subheader("ğŸ”‘ Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ API")
        
        key_input_method = st.radio(
            "Ø±ÙˆØ´ ÙˆØ±ÙˆØ¯ Ú©Ù„ÛŒØ¯Ù‡Ø§:",
            ["Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ Ù¾ÛŒØ´â€ŒÙØ±Ø¶", "Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ Ø³ÙØ§Ø±Ø´ÛŒ"],
            label_visibility="collapsed"
        )
        
        if key_input_method == "Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ Ù¾ÛŒØ´â€ŒÙØ±Ø¶":
            st.session_state.api_keys = DEFAULT_API_KEYS.copy()
            st.success(f"âœ… {len(st.session_state.api_keys)} Ú©Ù„ÛŒØ¯ Ù„ÙˆØ¯ Ø´Ø¯")
        else:
            custom_keys_text = st.text_area(
                "Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ API (Ø¬Ø¯Ø§ Ú©Ù†Ù†Ø¯Ù‡ enter):",
                height=120,
                placeholder="AIzaSy...\nAIzaSy..."
            )
            if custom_keys_text:
                st.session_state.api_keys = [key.strip() for key in custom_keys_text.split('\n') if key.strip()]
                st.success(f"âœ… {len(st.session_state.api_keys)} Ú©Ù„ÛŒØ¯ Ø³ÙØ§Ø±Ø´ÛŒ")
            else:
                st.session_state.api_keys = DEFAULT_API_KEYS.copy()

        # st.divider()

        # st.subheader(" Ù¾Ø§Ø±Ø§Ù…ØªØ±Ù‡Ø§ÛŒ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª API")
        st.markdown("""
        <style>
        /* Ø§Ø³ØªØ§ÛŒÙ„ Ø³ÙØ§Ø±Ø´ÛŒ Ø¨Ø±Ø§ÛŒ number input Ø¯Ø± Ø³Ø§ÛŒØ¯Ø¨Ø§Ø± */
        [data-testid="stSidebar"] .stNumberInput > div > div > input {
            text-align: center !important;
            font-size: 16px !important;
            font-weight: bold !important;
            background-color: #f0f2f6 !important;
            border: 2px solid #e0e0e0 !important;
            border-radius: 8px !important;
            padding: 8px !important;
        }
        
        [data-testid="stSidebar"] .stNumberInput > div > div > input:focus {
            border-color: #667eea !important;
            box-shadow: 0 0 0 2px rgba(102, 126, 234, 0.2) !important;
        }
        
        /* Ø¯Ú©Ù…Ù‡â€ŒÙ‡Ø§ÛŒ + Ùˆ - Ø¨Ø²Ø±Ú¯ØªØ± Ùˆ ÙˆØ§Ø¶Ø­â€ŒØªØ± */
        [data-testid="stSidebar"] .stNumberInput button {
            background-color: #667eea !important;
            color: white !important;
            border: none !important;
            border-radius: 6px !important;
            width: 35px !important;
            height: 35px !important;
            font-size: 18px !important;
            font-weight: bold !important;
            margin: 0 3px !important;
        }
        
        [data-testid="stSidebar"] .stNumberInput button:hover {
            background-color: #5568d3 !important;
            transform: scale(1.05);
        }
        
        /* Ú©Ø§Ù†ØªÛŒÙ†Ø± number input */
        [data-testid="stSidebar"] .stNumberInput > div {
            background-color: white !important;
            border-radius: 10px !important;
            padding: 5px !important;
            box-shadow: 0 2px 6px rgba(0,0,0,0.08) !important;
        }
        </style>
        """, unsafe_allow_html=True)
        


        # st.markdown('<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 12px; border-radius: 10px; text-align: center; margin-bottom: 15px;"><p style="color: white; margin: 0; font-size: 14px; font-weight: bold;">ğŸ›ï¸ Ù¾Ø§Ø±Ø§Ù…ØªØ±Ù‡Ø§ÛŒ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª API</p></div>', unsafe_allow_html=True)
        # Initialize session state for API limits if not exists
        if 'max_tokens_per_min' not in st.session_state:
            st.session_state.max_tokens_per_min = 125000
        if 'max_requests_per_min' not in st.session_state:
            st.session_state.max_requests_per_min = 2
        if 'max_requests_per_day' not in st.session_state:
            st.session_state.max_requests_per_day = 50
        

        with st.expander("âš¡ ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ù¾ÛŒØ´Ø±ÙØªÙ‡ Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ API", expanded=False):   
            # Tokens per minute
            # st.markdown('<div style="background: #e3f2fd; padding: 10px; border-radius: 8px; margin-bottom: 12px; border-right: 4px solid #2196f3;">', unsafe_allow_html=True)
            st.markdown('<p style="margin: 0 0 8px 0; font-size: 12px; font-weight: 600; color: #7b1fa2; text-align: right;">ğŸŸ£ Ø­Ø¯Ø§Ú©Ø«Ø± ØªÙˆÚ©Ù† Ø¯Ø± Ø¯Ù‚ÛŒÙ‚Ù‡ (Ù‡Ø± API)</p>', unsafe_allow_html=True)
            st.session_state.max_tokens_per_min = st.number_input(
                "max_tokens_label",
                min_value=1000,
                max_value=1000000,
                value=st.session_state.max_tokens_per_min,
                step=5000,
                label_visibility="collapsed",
                help="ØªØ¹Ø¯Ø§Ø¯ ØªÙˆÚ©Ù†â€ŒÙ‡Ø§ÛŒ Ù‚Ø§Ø¨Ù„ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ù‡Ø± Ø¯Ù‚ÛŒÙ‚Ù‡ Ø¨Ø±Ø§ÛŒ Ù‡Ø± API Key"
            )
            # st.markdown(f'<p style="margin: 5px 0 0 0; font-size: 11px; color: #666; text-align: center;">Ù…Ù‚Ø¯Ø§Ø± ÙØ¹Ù„ÛŒ: <strong>{st.session_state.max_tokens_per_min:,}</strong> ØªÙˆÚ©Ù†</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Requests per minute
            # st.markdown('<div style="background: #fff3e0; padding: 10px; border-radius: 8px; margin-bottom: 12px; border-right: 4px solid #ff9800;">', unsafe_allow_html=True)
            st.markdown('<p style="margin: 0 0 8px 0; font-size: 12px; font-weight: 600; color: #7b1fa2; text-align: right;">ğŸŸ£ Ø­Ø¯Ø§Ú©Ø«Ø± Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø¯Ø± Ø¯Ù‚ÛŒÙ‚Ù‡ (Ù‡Ø± API)</p>', unsafe_allow_html=True)
            st.session_state.max_requests_per_min = st.number_input(
                "max_requests_min_label",
                min_value=1,
                max_value=100,
                value=st.session_state.max_requests_per_min,
                step=1,
                label_visibility="collapsed",
                help="ØªØ¹Ø¯Ø§Ø¯ Ø¯Ø±Ø®ÙˆØ§Ø³Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø¬Ø§Ø² Ø¯Ø± Ù‡Ø± Ø¯Ù‚ÛŒÙ‚Ù‡ Ø¨Ø±Ø§ÛŒ Ù‡Ø± API Key"
            )
            # st.markdown(f'<p style="margin: 5px 0 0 0; font-size: 11px; color: #666; text-align: center;">Ù…Ù‚Ø¯Ø§Ø± ÙØ¹Ù„ÛŒ: <strong>{st.session_state.max_requests_per_min}</strong> Ø¯Ø±Ø®ÙˆØ§Ø³Øª</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Requests per day
            # st.markdown('<div style="background: #f3e5f5; padding: 10px; border-radius: 8px; margin-bottom: 12px; border-right: 4px solid #9c27b0;">', unsafe_allow_html=True)
            st.markdown('<p style="margin: 0 0 8px 0; font-size: 12px; font-weight: 600; color: #7b1fa2; text-align: right;">ğŸŸ£ Ø­Ø¯Ø§Ú©Ø«Ø± Ø¯Ø±Ø®ÙˆØ§Ø³Øª Ø¯Ø± Ø±ÙˆØ² (Ù‡Ø± API)</p>', unsafe_allow_html=True)
            st.session_state.max_requests_per_day = st.number_input(
                "max_requests_day_label",
                min_value=1,
                max_value=10000,
                value=st.session_state.max_requests_per_day,
                step=10,
                label_visibility="collapsed",
                help="ØªØ¹Ø¯Ø§Ø¯ Ú©Ù„ Ø¯Ø±Ø®ÙˆØ§Ø³Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø¬Ø§Ø² Ø¯Ø± Ù‡Ø± Ø±ÙˆØ² Ø¨Ø±Ø§ÛŒ Ù‡Ø± API Key"
            )
            # st.markdown(f'<p style="margin: 5px 0 0 0; font-size: 11px; color: #666; text-align: center;">Ù…Ù‚Ø¯Ø§Ø± ÙØ¹Ù„ÛŒ: <strong>{st.session_state.max_requests_per_day}</strong> Ø¯Ø±Ø®ÙˆØ§Ø³Øª</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # st.markdown("---")
        st.markdown("""
        <div style="background: linear-gradient(135deg, #e3f2fd 0%, #f3e5f5 100%); padding: 12px; border-radius: 8px; border: 1px solid #e0e0e0;">
            <p style="margin: 0; font-size: 12px; color: #424242; text-align: right; line-height: 1.6;">
            ğŸ’¡ <strong>Ù‡Ø´Ø¯Ø§Ø±:</strong> Ø§ÛŒÙ† Ù…Ù‚Ø§Ø¯ÛŒØ± Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ Gemini API ØªÙ†Ø¸ÛŒÙ… Ø´Ø¯Ù‡â€ŒØ§Ù†Ø¯. 
            ØªØºÛŒÛŒØ± Ø¢Ù†â€ŒÙ‡Ø§ Ù…ÛŒâ€ŒØªÙˆØ§Ù†Ø¯ Ø¨Ø± Ø³Ø±Ø¹Øª Ùˆ Ø¯Ù‚Øª Ù¾Ø±Ø¯Ø§Ø²Ø´ ØªØ£Ø«ÛŒØ± Ø¨Ú¯Ø°Ø§Ø±Ø¯.
            Ø¯Ø± ØµÙˆØ±ØªÛŒ Ú©Ù‡ Ø§Ø·Ù„Ø§Ø¹ Ø¯Ù‚ÛŒÙ‚ Ø§Ø² Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ù‡Ø§ Ùˆ ØªØºÛŒÛŒØ±Ø§Øª Ù…Ø¯Ù„ Ù†Ø¯Ø§Ø±ÛŒØ¯ ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø±Ø§ ØªØºÛŒÛŒØ± Ù†Ø¯Ù‡ÛŒØ¯.
             </p>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        

        
        #########
       
        
       


        st.info('''
        âš™ï¸ **Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡ÙˆØ´Ù…Ù†Ø¯**

        Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø¨ØµÙˆØ±Øª Ù‡Ù…Ø²Ù…Ø§Ù† Ùˆ Ù…ÙˆØ§Ø²ÛŒ Ø§Ù†Ø¬Ø§Ù… Ø®ÙˆÙ‡Ø¯ Ø´Ø¯
        
        ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ù‡Ù…Ø²Ù…Ø§Ù† Ø¨Ù‡ ØµÙˆØ±Øª Ø®ÙˆØ¯Ú©Ø§Ø± 
        Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…ÙˆØ§Ø±Ø¯ Ø²ÛŒØ± Ù…Ø­Ø§Ø³Ø¨Ù‡ Ù…ÛŒâ€ŒØ´ÙˆØ¯:

        âœ…  ØªØ¹Ø¯Ø§Ø¯ API Keys  
        âœ… Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ Ù…Ø¯Ù„  
        âœ… ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§  
        âœ… Ø­Ø§ÙØ¸Ù‡ Ø³ÛŒØ³ØªÙ…
        ''')

        st.markdown("<div style='margin-top: auto;'></div>", unsafe_allow_html=True)
        authenticator.logout('ğŸšª Ø®Ø±ÙˆØ¬ Ø§Ø² Ø³ÛŒØ³ØªÙ…', 'sidebar')


    class APILimitsManager:
        """Ù…Ø¯ÛŒØ±ÛŒØª Ù‡ÙˆØ´Ù…Ù†Ø¯ Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ API Ùˆ Ù…Ø­Ø§Ø³Ø¨Ù‡ ØªØ¹Ø¯Ø§Ø¯ workers Ø¨Ù‡ÛŒÙ†Ù‡"""
        
        # # Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ Gemini API
        # MAX_TOKENS_PER_MIN = 125_000      # tokens per minute per API
        # MAX_REQUESTS_PER_MIN = 2          # Maximum requests per minute per API
        # MAX_REQUESTS_PER_DAY = 50         # Maximum requests per day per API
        
        # ØªØ®Ù…ÛŒÙ†â€ŒÙ‡Ø§ Ø¨Ø±Ø§ÛŒ Ù‡Ø± ÙØ§ÛŒÙ„ PDF
        AVG_TOKENS_PER_FILE = 20_000      # ØªØ®Ù…ÛŒÙ† Ù…ØªÙˆØ³Ø· tokens Ø¨Ø±Ø§ÛŒ Ù‡Ø± ÙØ§ÛŒÙ„
        AVG_PROCESSING_TIME = 30          # ØªØ®Ù…ÛŒÙ† Ø²Ù…Ø§Ù† Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡Ø± ÙØ§ÛŒÙ„ (Ø«Ø§Ù†ÛŒÙ‡)
        
        def __init__(self, api_keys: list , max_tokens_per_min: int, max_requests_per_min: int, max_requests_per_day: int):
            """
            Args:
                api_keys: Ù„ÛŒØ³Øª API keys Ù…ÙˆØ¬ÙˆØ¯
            """
            self.num_api_keys = len(api_keys)
            self.api_usage = {key: {'requests_today': 0, 'last_reset': datetime.now()} 
                            for key in api_keys}

             # âœ… Ù…Ù‚Ø§Ø¯ÛŒØ± Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ Ø§Ú©Ù†ÙˆÙ† Ø§Ø² ÙˆØ±ÙˆØ¯ÛŒâ€ŒÙ‡Ø§ÛŒ ØªØ§Ø¨Ø¹ init Ú¯Ø±ÙØªÙ‡ Ù…ÛŒâ€ŒØ´ÙˆÙ†Ø¯
            self.MAX_TOKENS_PER_MIN = max_tokens_per_min
            self.MAX_REQUESTS_PER_MIN = max_requests_per_min
            self.MAX_REQUESTS_PER_DAY = max_requests_per_day      


        def calculate_optimal_workers(self, num_files: int, file_sizes: list = None) -> dict:
            """
            Ù…Ø­Ø§Ø³Ø¨Ù‡ ØªØ¹Ø¯Ø§Ø¯ Ø¨Ù‡ÛŒÙ†Ù‡ workers Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ API
            
            Args:
                num_files: ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ ÙˆØ±ÙˆØ¯ÛŒ
                file_sizes: Ø§Ù†Ø¯Ø§Ø²Ù‡ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ (Ø§Ø®ØªÛŒØ§Ø±ÛŒ) Ø¨Ø±Ø§ÛŒ ØªØ®Ù…ÛŒÙ† Ø¯Ù‚ÛŒÙ‚â€ŒØªØ±
            
            Returns:
                dict: Ø´Ø§Ù…Ù„ ØªØ¹Ø¯Ø§Ø¯ workersØŒ Ø²Ù…Ø§Ù† ØªØ®Ù…ÛŒÙ†ÛŒØŒ Ùˆ ØªÙˆØ¶ÛŒØ­Ø§Øª
            """
            
            # 1ï¸âƒ£ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Requests Per Minute
            # Ù‡Ø± API Ù…ÛŒâ€ŒØªÙˆØ§Ù†Ø¯ 2 request Ø¯Ø± Ø¯Ù‚ÛŒÙ‚Ù‡ Ø¯Ø§Ø´ØªÙ‡ Ø¨Ø§Ø´Ø¯
            max_workers_rpm = self.num_api_keys * self.MAX_REQUESTS_PER_MIN
            
            # 2ï¸âƒ£ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Tokens Per Minute
            # Ø¨Ø§ ÙØ±Ø¶ Ù‡Ø± ÙØ§ÛŒÙ„ 20K token
            max_files_per_min_tokens = (self.num_api_keys * self.MAX_TOKENS_PER_MIN) / self.AVG_TOKENS_PER_FILE
            max_workers_tokens = math.floor(max_files_per_min_tokens)
            
            # 3ï¸âƒ£ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Daily Requests
            # Ù‡Ø± API: 50 request Ø¯Ø± Ø±ÙˆØ²
            max_daily_files = self.num_api_keys * self.MAX_REQUESTS_PER_DAY
            
            # 4ï¸âƒ£ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø¹Ù…Ù„ÛŒ: Ø²Ù…Ø§Ù† Ù¾Ø±Ø¯Ø§Ø²Ø´
            # Ø§Ú¯Ø± Ù‡Ø± ÙØ§ÛŒÙ„ 30 Ø«Ø§Ù†ÛŒÙ‡ Ø·ÙˆÙ„ Ø¨Ú©Ø´Ø¯ Ùˆ Ù…Ø§ 60 Ø«Ø§Ù†ÛŒÙ‡ Ø¯Ø§Ø±ÛŒÙ…
            # Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒÙ… Ø­Ø¯Ø§Ú©Ø«Ø± 2 Ø¨Ø§Ø± Ø§Ø² Ù‡Ø± API Ø§Ø³ØªÙØ§Ø¯Ù‡ Ú©Ù†ÛŒÙ… (Ù…Ø·Ø§Ø¨Ù‚ Ø¨Ø§ RPM)
            processing_based_workers = max_workers_rpm
            
            # 5ï¸âƒ£ Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù…ØªØ±ÛŒÙ† Ù…Ù‚Ø¯Ø§Ø± (bottleneck)
            optimal_workers = min(
                max_workers_rpm,          # Ù…Ø­Ø¯ÙˆØ¯ÛŒØª RPM
                max_workers_tokens,       # Ù…Ø­Ø¯ÙˆØ¯ÛŒØª tokens
                num_files,                # ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§
                10                        # Ø­Ø¯Ø§Ú©Ø«Ø± Ù…Ù†Ø·Ù‚ÛŒ Ø¨Ø±Ø§ÛŒ Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² Ø§Ø´ØºØ§Ù„ Ù…Ù†Ø§Ø¨Ø¹
            )
            
            # Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² Ø§ÛŒÙ†Ú©Ù‡ Ø­Ø¯Ø§Ù‚Ù„ 1 worker Ø¯Ø§Ø±ÛŒÙ…
            optimal_workers = max(1, optimal_workers)
            
            # 6ï¸âƒ£ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø²Ù…Ø§Ù† ØªØ®Ù…ÛŒÙ†ÛŒ
            # Ø¨Ø§ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù…ÙˆØ§Ø²ÛŒ
            estimated_time_parallel = (num_files / optimal_workers) * self.AVG_PROCESSING_TIME
            # Ø¨Ø¯ÙˆÙ† Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù…ÙˆØ§Ø²ÛŒ
            estimated_time_sequential = num_files * self.AVG_PROCESSING_TIME
            
            # 7ï¸âƒ£ Ø¨Ø±Ø±Ø³ÛŒ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø±ÙˆØ²Ø§Ù†Ù‡
            daily_limit_ok = num_files <= max_daily_files
            
            # 8ï¸âƒ£ ØªØ¹ÛŒÛŒÙ† Ø§Ø³ØªØ±Ø§ØªÚ˜ÛŒ
            if num_files <= max_workers_rpm:
                strategy = "fast_parallel"
                message = f"âœ… Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø³Ø±ÛŒØ¹: Ù‡Ù…Ù‡ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø¨Ù‡ Ø·ÙˆØ± Ù‡Ù…Ø²Ù…Ø§Ù† ({optimal_workers} worker)"
            elif num_files <= max_daily_files:
                strategy = "batch_parallel"
                message = f"âš¡ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø³ØªÙ‡â€ŒØ§ÛŒ: {optimal_workers} worker Ø¨Ø§ Ú†Ù†Ø¯ batch"
            else:
                strategy = "limited"
                optimal_workers = min(optimal_workers, 3)
                message = f"âš ï¸ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø±ÙˆØ²Ø§Ù†Ù‡: ÙÙ‚Ø· {max_daily_files} ÙØ§ÛŒÙ„ Ø§Ù…Ú©Ø§Ù†â€ŒÙ¾Ø°ÛŒØ± Ø§Ø³Øª"
            
            return {
                'optimal_workers': optimal_workers,
                'strategy': strategy,
                'message': message,
                'estimated_time_minutes': estimated_time_parallel / 60,
                'speedup_factor': estimated_time_sequential / estimated_time_parallel,
                'limits': {
                    'max_rpm': max_workers_rpm,
                    'max_tokens': max_workers_tokens,
                    'max_daily': max_daily_files,
                    'files_count': num_files,
                    'daily_limit_ok': daily_limit_ok
                },
                'explanation': self._generate_explanation(
                    optimal_workers, 
                    num_files, 
                    max_workers_rpm,
                    estimated_time_parallel
                )
            }
        
        def _generate_explanation(self, workers: int, files: int, max_rpm: int, time_min: float) -> str:
            """ØªÙˆÙ„ÛŒØ¯ ØªÙˆØ¶ÛŒØ­Ø§Øª Ø¨Ø±Ø§ÛŒ Ú©Ø§Ø±Ø¨Ø±"""
            explanations = []
            
            explanations.append(f"ğŸ”§ **ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø´Ø¯Ù‡:**")
            explanations.append(f"  â€¢ ØªØ¹Ø¯Ø§Ø¯ API Keys: {self.num_api_keys}")
            explanations.append(f"  â€¢ ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§: {files}")
            explanations.append(f"  â€¢ Workers Ø¨Ù‡ÛŒÙ†Ù‡: {workers}")
            explanations.append(f"  â€¢ Ø²Ù…Ø§Ù† ØªØ®Ù…ÛŒÙ†ÛŒ: {time_min:.1f} Ø¯Ù‚ÛŒÙ‚Ù‡")
            
            explanations.append(f"\nğŸ“Š **Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ API:**")
            explanations.append(f"  â€¢ Ø­Ø¯Ø§Ú©Ø«Ø± Ù‡Ù…Ø²Ù…Ø§Ù†: {max_rpm} ÙØ§ÛŒÙ„ Ø¯Ø± Ø¯Ù‚ÛŒÙ‚Ù‡")
            explanations.append(f"  â€¢ Ù‡Ø± API: {self.MAX_REQUESTS_PER_MIN} request/min")
            explanations.append(f"  â€¢ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø±ÙˆØ²Ø§Ù†Ù‡: {self.MAX_REQUESTS_PER_DAY * self.num_api_keys} ÙØ§ÛŒÙ„")
            
            if files > max_rpm:
                batches = math.ceil(files / workers)
                explanations.append(f"\nâš¡ **Ø§Ø³ØªØ±Ø§ØªÚ˜ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´:**")
                explanations.append(f"  â€¢ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± {batches} Ø¯Ø³ØªÙ‡")
                explanations.append(f"  â€¢ Ù‡Ø± Ø¯Ø³ØªÙ‡: {workers} ÙØ§ÛŒÙ„ Ù‡Ù…Ø²Ù…Ø§Ù†")
            
            return "\n".join(explanations)

    
    # ============================================================================
    # ğŸ”„ ØªØ§Ø¨Ø¹ Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡: Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¨Ø§ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø®ÙˆØ¯Ú©Ø§Ø± workers
    # ============================================================================

    def process_files_concurrent_smart(uploaded_files):
        """
        âœ… Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡Ù…Ø²Ù…Ø§Ù† Ø¨Ø§ Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø®ÙˆØ¯Ú©Ø§Ø± ØªØ¹Ø¯Ø§Ø¯ workers Ø¨Ù‡ÛŒÙ†Ù‡
        
        ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ø¬Ø¯ÛŒØ¯:
        - Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø®ÙˆØ¯Ú©Ø§Ø± workers Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ø­Ø¯ÙˆØ¯ÛŒØªâ€ŒÙ‡Ø§ÛŒ API
        - Ù†Ù…Ø§ÛŒØ´ ØªÙˆØ¶ÛŒØ­Ø§Øª Ú©Ø§Ù…Ù„ Ø¨Ø±Ø§ÛŒ Ú©Ø§Ø±Ø¨Ø±
        - Ø¨Ù‡ÛŒÙ†Ù‡â€ŒØ³Ø§Ø²ÛŒ Ù…Ù†Ø§Ø¨Ø¹
        """
        
        # 1ï¸âƒ£ Ù…Ø­Ø§Ø³Ø¨Ù‡ ØªØ¹Ø¯Ø§Ø¯ workers Ø¨Ù‡ÛŒÙ†Ù‡
        limits_manager = APILimitsManager(   
            api_keys=st.session_state.api_keys,
            max_tokens_per_min=st.session_state.max_tokens_per_min,
            max_requests_per_min=st.session_state.max_requests_per_min,
            max_requests_per_day=st.session_state.max_requests_per_day
        )

        optimization = limits_manager.calculate_optimal_workers(
            num_files=len(uploaded_files)
        )
        
        optimal_workers = optimization['optimal_workers']
        
        # 2ï¸âƒ£ Ù†Ù…Ø§ÛŒØ´ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ø¨Ø±Ø§ÛŒ Ú©Ø§Ø±Ø¨Ø±
        st.markdown('<div class="modern-card">', unsafe_allow_html=True)
        
        # Ù†Ù…Ø§ÛŒØ´ ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø´Ø¯Ù‡
        # with st.expander("ğŸ¤– ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø®ÙˆØ¯Ú©Ø§Ø± (Ú©Ù„ÛŒÚ© Ø¨Ø±Ø§ÛŒ Ø¬Ø²Ø¦ÛŒØ§Øª)", expanded=False):
        #     st.markdown(optimization['explanation'])
        
        # Ù†Ù…Ø§ÛŒØ´ Ù¾ÛŒØ§Ù… Ø§ØµÙ„ÛŒ
        # st.info(optimization['message'])
        

        
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        
        # 3ï¸âƒ£ Ø¨Ø±Ø±Ø³ÛŒ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø±ÙˆØ²Ø§Ù†Ù‡
        if not optimization['limits']['daily_limit_ok']:
            st.error(
                f"âš ï¸ ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ ({len(uploaded_files)}) Ø¨ÛŒØ´ØªØ± Ø§Ø² Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø±ÙˆØ²Ø§Ù†Ù‡ "
                f"({optimization['limits']['max_daily']}) Ø§Ø³Øª. "
                f"Ù„Ø·ÙØ§Ù‹ ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø±Ø§ Ú©Ø§Ù‡Ø´ Ø¯Ù‡ÛŒØ¯ ÛŒØ§ API key Ù‡Ø§ÛŒ Ø¨ÛŒØ´ØªØ±ÛŒ Ø§Ø¶Ø§ÙÙ‡ Ú©Ù†ÛŒØ¯."
            )
            return None
        
        # 4ï¸âƒ£ Ø´Ø±ÙˆØ¹ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¨Ø§ workers Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø´Ø¯Ù‡
        analyzer = FinancialAnalyzer()
        total_files = len(uploaded_files)
        max_retry_attempts = 3
        
        st.markdown('<div class="modern-card"><h3>Ø¯Ø± Ø­Ø§Ù„ Ù¾Ø±Ø¯Ø§Ø²Ø´...</h3></div>', unsafe_allow_html=True)
        progress_bar = st.progress(0)
        status_placeholder = st.empty()
        status_container = st.container()
        
        # Metrics
        col1, col2, col3, col4 = st.columns(4)
        metric_success = col1.empty()
        metric_failed = col2.empty()
        metric_retrying = col3.empty()
        metric_total = col4.empty()
        
        results = [None] * total_files
        completed = 0
        failed_count = 0
        retry_count = 0
        start_time = time.time()
        files_to_retry = []
        
        # 5ï¸âƒ£ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø§ÙˆÙ„ÛŒÙ‡ Ø¨Ø§ optimal_workers
        with ThreadPoolExecutor(max_workers=optimal_workers) as executor:
            future_to_index = {
                executor.submit(process_single_file, analyzer, file, i, total_files, 1, max_retry_attempts): i
                for i, file in enumerate(uploaded_files)
            }
            
            for future in as_completed(future_to_index):
                index, filename, result, error, needs_retry = future.result()
                completed += 1
                
                if error:
                    if needs_retry:
                        files_to_retry.append((index, uploaded_files[index], filename, 2))
                        retry_count += 1
                        with status_container:
                            st.warning(f'ğŸ”„ **{filename}** Ù†ÛŒØ§Ø² Ø¨Ù‡ ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯ Ø¯Ø§Ø±Ø¯ ({completed}/{total_files})')
                    else:
                        results[index] = (filename, {"error": f"Ø®Ø·Ø§: {error}"})
                        failed_count += 1
                        with status_container:
                            st.error(f'âŒ **{filename}**: Ø®Ø·Ø§ÛŒ ØºÛŒØ±Ù‚Ø§Ø¨Ù„ Ø¨Ø§Ø²ÛŒØ§Ø¨ÛŒ')
                else:
                    results[index] = (filename, result)
                    with status_container:
                        st.success(f'âœ… **{filename}** ({completed}/{total_files})')
                
                progress_bar.progress(completed / total_files)
                
                metric_success.metric("âœ… Ù…ÙˆÙÙ‚", len([r for r in results if r and 'error' not in r[1]]))
                metric_failed.metric("âŒ Ù†Ø§Ù…ÙˆÙÙ‚", failed_count)
                metric_retrying.metric("ğŸ”„ Ø¯Ø± Ø§Ù†ØªØ¸Ø§Ø± ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯", len(files_to_retry))
                metric_total.metric("ğŸ“Š Ú©Ù„", total_files)
                
                elapsed = time.time() - start_time
                avg_time = elapsed / completed
                remaining = (total_files - completed + len(files_to_retry)) * avg_time
                status_placeholder.info(
                    f'ğŸ“Š Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø§ÙˆÙ„ÛŒÙ‡: {completed}/{total_files} | '
                    f'â±ï¸ Ø²Ù…Ø§Ù†: {elapsed:.1f}s | â³ ØªØ®Ù…ÛŒÙ†: {remaining:.1f}s'
                )
        
        # 6ï¸âƒ£ Ù…Ø±Ø­Ù„Ù‡ Retry (Ø§Ú¯Ø± Ù†ÛŒØ§Ø² Ø¨Ø§Ø´Ø¯)
        if files_to_retry:
            st.markdown("---")
            st.markdown("### ğŸ”„ ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯ Ø¨Ø±Ø§ÛŒ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ù†Ø§Ù…ÙˆÙÙ‚...")
            
            retry_attempt = 1
            # Ø¯Ø± retry Ø§Ø² Ù†ØµÙ workers Ø§Ø³ØªÙØ§Ø¯Ù‡ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
            retry_workers = max(1, optimal_workers // 2)
            
            while files_to_retry and retry_attempt <= max_retry_attempts:
                st.info(f"ğŸ”„ Ø¯ÙˆØ± {retry_attempt} Ø§Ø² ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯ ({len(files_to_retry)} ÙØ§ÛŒÙ„) Ø¨Ø§ {retry_workers} worker")
                
                delay = min(5 * retry_attempt, 15)
                with st.spinner(f'â³ ØµØ¨Ø± {delay} Ø«Ø§Ù†ÛŒÙ‡ Ù‚Ø¨Ù„ Ø§Ø² ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯...'):
                    time.sleep(delay)
                
                current_retry_list = files_to_retry.copy()
                files_to_retry = []
                
                with ThreadPoolExecutor(max_workers=retry_workers) as executor:
                    future_to_data = {
                        executor.submit(
                            process_single_file, 
                            analyzer, 
                            file_data, 
                            idx, 
                            total_files, 
                            attempt_num,
                            max_retry_attempts
                        ): (idx, fname, attempt_num)
                        for idx, file_data, fname, attempt_num in current_retry_list
                    }
                    
                    for future in as_completed(future_to_data):
                        idx, fname, attempt_num = future_to_data[future]
                        index, filename, result, error, needs_retry = future.result()
                        
                        if error:
                            if needs_retry and attempt_num < max_retry_attempts:
                                files_to_retry.append((index, uploaded_files[index], filename, attempt_num + 1))
                                with status_container:
                                    st.warning(f'ğŸ”„ **{filename}** - ØªÙ„Ø§Ø´ {attempt_num + 1}/{max_retry_attempts}')
                            else:
                                results[index] = (filename, {"error": f"Ø®Ø·Ø§ Ø¨Ø¹Ø¯ Ø§Ø² {attempt_num} ØªÙ„Ø§Ø´: {error}"})
                                failed_count += 1
                                with status_container:
                                    st.error(f'âŒ **{filename}**: Ù†Ø§Ù…ÙˆÙÙ‚ Ø¨Ø¹Ø¯ Ø§Ø² {attempt_num} ØªÙ„Ø§Ø´')
                        else:
                            results[index] = (filename, result)
                            with status_container:
                                st.success(f'âœ… **{filename}** Ù…ÙˆÙÙ‚ Ø¯Ø± ØªÙ„Ø§Ø´ {attempt_num}!')
                            failed_count = max(0, failed_count - 1)
                        
                        successful = len([r for r in results if r and 'error' not in r[1]])
                        metric_success.metric("âœ… Ù…ÙˆÙÙ‚", successful)
                        metric_failed.metric("âŒ Ù†Ø§Ù…ÙˆÙÙ‚", failed_count)
                        metric_retrying.metric("ğŸ”„ Ø¯Ø± Ø§Ù†ØªØ¸Ø§Ø± ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯", len(files_to_retry))
                
                retry_attempt += 1
        
        # 7ï¸âƒ£ Ú¯Ø²Ø§Ø±Ø´ Ù†Ù‡Ø§ÛŒÛŒ
        total_duration = time.time() - start_time
        successful = len([r for r in results if r and 'error' not in r[1]])
        
        st.markdown("---")
        
        if failed_count == 0:
            status_placeholder.success(
                f'ğŸ‰ Ù‡Ù…Ù‡ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø´Ø¯Ù†Ø¯! '
                f'{total_files} ÙØ§ÛŒÙ„ Ø¯Ø± {total_duration:.1f} Ø«Ø§Ù†ÛŒÙ‡ ({total_duration/60:.1f} Ø¯Ù‚ÛŒÙ‚Ù‡)'
            )
        else:
            status_placeholder.warning(
                f'âš ï¸ Ù¾Ø±Ø¯Ø§Ø²Ø´ ØªÚ©Ù…ÛŒÙ„ Ø´Ø¯: {successful}/{total_files} Ù…ÙˆÙÙ‚ØŒ {failed_count} Ù†Ø§Ù…ÙˆÙÙ‚ '
                f'Ø¯Ø± {total_duration:.1f} Ø«Ø§Ù†ÛŒÙ‡ ({total_duration/60:.1f} Ø¯Ù‚ÛŒÙ‚Ù‡)'
            )
        
        if retry_count > 0:
            st.info(f'â„¹ï¸ ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒÛŒ Ú©Ù‡ Ù†ÛŒØ§Ø² Ø¨Ù‡ ØªÙ„Ø§Ø´ Ù…Ø¬Ø¯Ø¯ Ø¯Ø§Ø´ØªÙ†Ø¯: {retry_count}')
        
        # Ù†Ù…Ø§ÛŒØ´ Ù…Ù‚Ø§ÛŒØ³Ù‡ Ø¨Ø§ Ø²Ù…Ø§Ù† ØªØ®Ù…ÛŒÙ†ÛŒ
        estimated_time = optimization['estimated_time_minutes'] * 60
        if abs(total_duration - estimated_time) < estimated_time * 0.2:  # Â±20%
            st.success(f"âœ… Ø²Ù…Ø§Ù† Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù…Ø·Ø§Ø¨Ù‚ ØªØ®Ù…ÛŒÙ† Ø¨ÙˆØ¯!")
        elif total_duration < estimated_time:
            st.success(f"ğŸš€ Ù¾Ø±Ø¯Ø§Ø²Ø´ {(estimated_time - total_duration):.0f} Ø«Ø§Ù†ÛŒÙ‡ Ø³Ø±ÛŒØ¹â€ŒØªØ± Ø§Ø² ØªØ®Ù…ÛŒÙ† Ø¨ÙˆØ¯!")
        
        return results
    


    # ========================================================================
    # Ø¨Ø®Ø´ 4: Ú©Ù„Ø§Ø³â€ŒÙ‡Ø§ Ùˆ ØªÙˆØ§Ø¨Ø¹ Ø§ØµÙ„ÛŒ (Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ±)
    # ========================================================================

    class APIKeyManager:
        def __init__(self, api_keys: List[str]):
            if not api_keys:
                raise ValueError("API keys list cannot be empty.")
            self.api_keys = api_keys
            self.current_index = 0
            self.failures = {key: 0 for key in api_keys}
            self.max_failures = 3
            self.lock = threading.Lock()  # âœ… Ø§Ø¶Ø§ÙÙ‡ Ø´Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ thread-safety
            
        def get_next_key(self) -> str:
            with self.lock:  # âœ… Ù…Ø­Ø§ÙØ¸Øª Ø§Ø² race condition
                attempts = 0
                while attempts < len(self.api_keys):
                    key = self.api_keys[self.current_index]
                    self.current_index = (self.current_index + 1) % len(self.api_keys)
                    if self.failures.get(key, 0) < self.max_failures:
                        return key
                    attempts += 1
                logger.warning("All API keys have failed, resetting failure counters")
                self.failures = {key: 0 for key in self.api_keys}
                return self.api_keys[0]
        
        def mark_failure(self, key: str):
            with self.lock:
                if key in self.failures:
                    self.failures[key] += 1
                    logger.warning(f"API key failure count for {key[:8]}...: {self.failures[key]}")
        
        def mark_success(self, key: str):
            with self.lock:
                if key in self.failures:
                    self.failures[key] = 0

    api_key_manager = APIKeyManager(st.session_state.api_keys)

    def get_client_with_retry():
        api_key = api_key_manager.get_next_key()
        return genai.Client(api_key=api_key), api_key

    # ========================================================================
    # Ø¨Ø®Ø´ 5: ØªÙˆØ§Ø¨Ø¹ Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§Ø±Ø³ÛŒ Ùˆ Ø§Ø¯ØºØ§Ù…
    # ========================================================================

# Ø§ÛŒÙ† Ú©Ø¯ Ø±Ø§ Ø¨Ù‡ Ø·ÙˆØ± Ú©Ø§Ù…Ù„ Ø¬Ø§ÛŒÚ¯Ø²ÛŒÙ† ØªØ§Ø¨Ø¹ setup_persian_font ÙØ¹Ù„ÛŒ Ø®ÙˆØ¯ Ú©Ù†ÛŒØ¯

    def setup_persian_font():
        """
        ØªÙ†Ø¸ÛŒÙ… ÙÙˆÙ†Øª Vazirmatn Ø¨Ø±Ø§ÛŒ Matplotlib Ø¨Ø§ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² FontManager
        """
        try:
            
            
            # âœ… Ù…Ø³ÛŒØ± ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª
            font_path = 'fonts/NotoNaskhArabic-Regular.ttf'
            
            if os.path.exists(font_path):
                # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† ÙÙˆÙ†Øª Ø¨Ù‡ Font Manager
                font_manager.fontManager.addfont(font_path)
                
                # âœ… ØªÙ†Ø¸ÛŒÙ… Ø¨Ù‡ Ø¹Ù†ÙˆØ§Ù† ÙÙˆÙ†Øª Ù¾ÛŒØ´â€ŒÙØ±Ø¶
                plt.rcParams['font.family'] = 'Noto Naskh Arabic'
                plt.rcParams['axes.unicode_minus'] = False
                
                logger.info("âœ… ÙÙˆÙ†Øª Noto Naskh Arabic Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ù„ÙˆØ¯ Ø´Ø¯")
                
                # Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† FontProperties Ø¨Ø±Ø§ÛŒ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ù…Ø³ØªÙ‚ÛŒÙ…
                return FontProperties(fname=font_path)
                
            else:
                logger.warning(f"âš ï¸ ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯: {font_path}")
                # Fallback Ø¨Ù‡ Tahoma
                plt.rcParams['font.family'] = ['Tahoma', 'Arial', 'DejaVu Sans']
                plt.rcParams['axes.unicode_minus'] = False
                return FontProperties(family='Tahoma')
                
        except Exception as e:
            logger.error(f"âŒ Ø®Ø·Ø§ Ø¯Ø± ØªÙ†Ø¸ÛŒÙ… ÙÙˆÙ†Øª: {e}")
            plt.rcParams['font.family'] = ['Tahoma', 'Arial']
            plt.rcParams['axes.unicode_minus'] = False
            return FontProperties()

    def process_persian_text(text):
        reshaped_text = arabic_reshaper.reshape(str(text))
        bidi_text = get_display(reshaped_text)
        return bidi_text

    def merge_excel_files(excel_files):
        if len(excel_files) < 2:
            return None
        
        merged_data = {}
        company_names = []
        
        try:
            for file_path in excel_files:
                df_summary = pd.read_excel(file_path, sheet_name="Ø¨Ø®Ø´1_Ø®Ù„Ø§ØµÙ‡")
                if not df_summary.empty and 'Ù†Ø§Ù…_Ø´Ø±Ú©Øª' in df_summary.columns:
                    company_names.append(df_summary['Ù†Ø§Ù…_Ø´Ø±Ú©Øª'].iloc[0])
            
            if len(set(company_names)) > 1:
                logger.warning("Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ Ù‡Ù…â€ŒÙ†Ø§Ù… Ù†ÛŒØ³ØªÙ†Ø¯")
                return None
            
            sheet_names = ["Ø¨Ø®Ø´1_Ø®Ù„Ø§ØµÙ‡", "Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ", "Ø¨Ø®Ø´3_Ú†Ú©_Ù„ÛŒØ³Øª", "Ú¯Ø²Ø§Ø±Ø´_Ù‚Ø§Ù†ÙˆÙ†ÛŒ"]
            
            for sheet in sheet_names:
                dfs = []
                for file_path in excel_files:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet)
                        dfs.append(df)
                    except:
                        continue
                if dfs:
                    merged_data[sheet] = pd.concat(dfs, ignore_index=True)
            
            return merged_data if merged_data else None
        except Exception as e:
            logger.error(f"Ø®Ø·Ø§ Ø¯Ø± Ø§Ø¯ØºØ§Ù…: {e}")
            return None

    # ========================================================================
    # Ø¨Ø®Ø´ 6: ØªÙˆØ§Ø¨Ø¹ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø± (7 Ù†Ù…ÙˆØ¯Ø§Ø±)
    # ========================================================================

    def plot_opinion_trend(df, font):
        """Ù†Ù…ÙˆØ¯Ø§Ø± Ø®Ø·ÛŒ Ø±ÙˆÙ†Ø¯ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± - Ø¨Ø§ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù…ÛŒØ§Ù†Ù‡ (Median) Ø¨Ø±Ø§ÛŒ ØªØ¬Ù…ÛŒØ¹ ØµØ­ÛŒØ­"""
        opinion_mp = {'Ù…Ù‚Ø¨ÙˆÙ„': 0, 'Ù…Ø´Ø±ÙˆØ·': 1, 'Ù…Ø±Ø¯ÙˆØ¯': 2, 'Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±': 3}
        
        df = df.copy()
        df['year'] = pd.to_numeric(df['year'], errors='coerce').astype('Int64')
        df = df.dropna(subset=['year'])
        df["opinion_mp"] = df['Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±'].map(opinion_mp)
        df = df.sort_values('year')
        
        # --- Ø§ØµÙ„Ø§Ø­ Ù…Ù†Ø·Ù‚ ØªØ¬Ù…ÛŒØ¹ ---
        # Ø¨Ù‡ Ø¬Ø§ÛŒ Ù…ÛŒØ§Ù†Ú¯ÛŒÙ†ØŒ Ø§Ø² Ù…ÛŒØ§Ù†Ù‡ (Median) Ø§Ø³ØªÙØ§Ø¯Ù‡ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ… Ùˆ Ø¢Ù† Ø±Ø§ Ú¯Ø±Ø¯ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
        # ØªØ§ Ø®Ø±ÙˆØ¬ÛŒ Ø­ØªÙ…Ø§ ÛŒÚ©ÛŒ Ø§Ø² Ø§Ø¹Ø¯Ø§Ø¯ ØµØ­ÛŒØ­ 0ØŒ 1ØŒ 2 ÛŒØ§ 3 Ø¨Ø§Ø´Ø¯.
        df_grouped = df.groupby('year')['opinion_mp'].median().round().astype(int).reset_index()
        
        fig, ax = plt.subplots(figsize=(12, 7), facecolor='#f8f9fa')
        ax.set_facecolor('#ffffff')

        # 1. Ø±Ø³Ù… Ù†Ù‚Ø§Ø· Ù¾Ø±Ø§Ú©Ù†Ø¯Ù‡ (Scatter) Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡
        # Ø§ÛŒÙ† Ù†Ù‚Ø§Ø· Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ù†Ø¯ Ú©Ù‡ Ø¯Ø± Ù‡Ø± Ø³Ø§Ù„ Ú†Ù‡ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ÛŒ Ø¯ÛŒÚ¯Ø±ÛŒ Ù‡Ù… ÙˆØ¬ÙˆØ¯ Ø¯Ø§Ø´ØªÙ‡ Ø§Ø³Øª (Ø¨Ø¯ÙˆÙ† Ø§ØªØµØ§Ù„ Ø®Ø·ÛŒ)
        # jitter (Ù„Ø±Ø²Ø´) Ø¬Ø²Ø¦ÛŒ Ø§Ø¶Ø§ÙÙ‡ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ… ØªØ§ Ù†Ù‚Ø§Ø· Ø±ÙˆÛŒ Ù‡Ù… Ù†ÛŒÙØªÙ†Ø¯
        jitter_y = df['opinion_mp'] + np.random.uniform(-0.1, 0.1, size=len(df))
        ax.scatter(df['year'], jitter_y, color='#3498db', alpha=0.15, s=50, zorder=1)

        # 2. Ø±Ø³Ù… Ø®Ø· Ø±ÙˆÙ†Ø¯ Ø§ØµÙ„ÛŒ (Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…ÛŒØ§Ù†Ù‡)
        ax.plot(df_grouped['year'], df_grouped['opinion_mp'], 
                marker='o', markersize=12, linestyle='-', 
                linewidth=3, color='#2980b9', zorder=2, label='Ø±ÙˆÙ†Ø¯ Ú©Ù„ÛŒ (Ù…ÛŒØ§Ù†Ù‡)')
        
        # ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ø¸Ø§Ù‡Ø±ÛŒ
        ax.set_xlabel(process_persian_text('Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ'), fontproperties=font, size=20, weight='bold', color='#34495e')
        ax.set_ylabel(process_persian_text('Ù†ÙˆØ¹ Ø§Ø¸Ù‡Ø§Ø± Ù†Ø¸Ø±'), fontproperties=font, size=20, weight='bold', color='#34495e')
        
        ax.set_yticks(list(opinion_mp.values()))
        y_labels = [process_persian_text(label) for label in opinion_mp.keys()]
        ax.set_yticklabels(y_labels, fontproperties=font, fontsize=20)
        
        if len(df_grouped) > 0:
            years = sorted(df_grouped['year'].unique())
            ax.set_xticks(years)
            ax.set_xticklabels([int(y) for y in years], fontproperties=font, fontsize=20)
        
        ax.set_ylim(-0.5, 3.5)
        ax.grid(True, linestyle='--', alpha=0.3, color='#bdc3c7')
        
        # Ø­Ø°Ù Ø­Ø§Ø´ÛŒÙ‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¶Ø§ÙÛŒ
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#bdc3c7')
        ax.spines['bottom'].set_color('#bdc3c7')
        
        if ax.get_legend():
            ax.get_legend().remove()
                
        return fig


    def plot_risk_trend(df, font):
        """Ù†Ù…ÙˆØ¯Ø§Ø± Ø®Ø·ÛŒ Ø±ÙˆÙ†Ø¯ Ø³Ø·Ø­ Ø±ÛŒØ³Ú© - Ø¨Ø§ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù…ÛŒØ§Ù†Ù‡ (Median)"""
        risk_level = ['Ù¾Ø§ÛŒÛŒÙ†', 'Ù…ØªÙˆØ³Ø·', 'Ø¨Ø§Ù„Ø§', 'Ø¨Ø­Ø±Ø§Ù†ÛŒ']
        risk_mp = {'Ù¾Ø§ÛŒÛŒÙ†': 0, 'Ù…ØªÙˆØ³Ø·': 1, 'Ø¨Ø§Ù„Ø§': 2, 'Ø¨Ø­Ø±Ø§Ù†ÛŒ': 3}
        
        df = df.copy()
        df['year'] = pd.to_numeric(df['year'], errors='coerce').astype('Int64')
        df = df.dropna(subset=['year'])
        df["risk_mp"] = df['Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_Ú©Ù„ÛŒ_Ø¨Ù†Ø§_Ø¨Ù‡_Ú¯Ø²Ø§Ø±Ø´'].map(risk_mp)
        df = df.sort_values('year')
        
        # --- Ø§ØµÙ„Ø§Ø­ Ù…Ù†Ø·Ù‚ ØªØ¬Ù…ÛŒØ¹ ---
        # Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù…ÛŒØ§Ù†Ù‡ Ùˆ Ú¯Ø±Ø¯ Ú©Ø±Ø¯Ù† Ø¨Ù‡ Ù†Ø²Ø¯ÛŒÚ©ØªØ±ÛŒÙ† Ø¹Ø¯Ø¯ ØµØ­ÛŒØ­
        df_grouped = df.groupby('year')['risk_mp'].median().round().astype(int).reset_index()

        fig, ax = plt.subplots(figsize=(12, 7), facecolor='#f8f9fa')
        ax.set_facecolor('#ffffff')
        
        # 1. Ø±Ø³Ù… ØªÙ…Ø§Ù… Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ Ø¨Ù‡ ØµÙˆØ±Øª Ú©Ù…â€ŒØ±Ù†Ú¯ (Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ ØªÙˆØ²ÛŒØ¹)
        jitter_y = df['risk_mp'] + np.random.uniform(-0.1, 0.1, size=len(df))
        ax.scatter(df['year'], jitter_y, color='#e74c3c', alpha=0.15, s=50, zorder=1)

        # 2. Ø±Ø³Ù… Ø®Ø· Ø±ÙˆÙ†Ø¯ Ø§ØµÙ„ÛŒ (Ù…ÛŒØ§Ù†Ù‡)
        ax.plot(df_grouped['year'], df_grouped['risk_mp'], 
                marker='o', markersize=12, linestyle='-', 
                linewidth=3, color='#c0392b', zorder=2)
        
        ax.set_xlabel(process_persian_text('Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ'), fontproperties=font, size=20, weight='bold', color='#34495e')
        ax.set_ylabel(process_persian_text('Ø³Ø·Ø­ Ø±ÛŒØ³Ú©'), fontproperties=font, size=20, weight='bold', color='#34495e')

        ax.set_yticks(list(risk_mp.values()))
        y_labels = [process_persian_text(label) for label in risk_mp.keys()]
        ax.set_yticklabels(y_labels, fontproperties=font, fontsize=20)
        
        if len(df_grouped) > 0:
            years = sorted(df_grouped['year'].unique())
            ax.set_xticks(years)
            ax.set_xticklabels([int(y) for y in years], fontproperties=font, fontsize=20)
        
        ax.set_ylim(-0.5, len(risk_level) - 0.5)
        ax.grid(True, linestyle='--', alpha=0.3, color='#bdc3c7')
        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#bdc3c7')
        ax.spines['bottom'].set_color('#bdc3c7')
        
        if ax.get_legend():
            ax.get_legend().remove()
            
        return fig

    def plot_checklist_heatmap(df, font):
        """Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ - Ø³Ø§ÛŒØ² Ø¨Ø²Ø±Ú¯ØªØ±"""
        status_mapping = {
            'Ù…ØµØ¯Ø§Ù‚ Ù†Ø¯Ø§Ø±Ø¯': 0,
            'Ø¨Ø±Ø±Ø³ÛŒ Ø´Ø¯Ù‡ - Ø±ÛŒØ³Ú© Ø®Ø§ØµÛŒ Ú¯Ø²Ø§Ø±Ø´ Ù†Ø´Ø¯Ù‡': 1,
            'Ù…Ø³Ø¦Ù„Ù‡ Ú©Ù„ÛŒØ¯ÛŒ Ù…Ù†Ø¬Ø± Ø¨Ù‡ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± Ù…Ø´Ø±ÙˆØ·': 2,
            'Ø±ÛŒØ³Ú© Ø¨Ø­Ø±Ø§Ù†ÛŒ': 3
        }
        
        df["status_mp"] = df['ÙˆØ¶Ø¹ÛŒØª'].map(status_mapping)
        # irrelevant_topics = [
        #     'Ø¢Ø®Ø±ÛŒÙ† ØµÙØ­Ù‡ Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ú©Ù‡ Ø´Ø§Ù…Ù„ Ø§Ù…Ø¶Ø§ Ø³Ø§Ø²Ù…Ø§Ù† Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…ÛŒØ´ÙˆØ¯',
        #     'ØµÙØ­Ù‡ Ø§Ù…Ø¶Ø§ Ù‡Ø§ÛŒ Ø³Ø§Ø²Ù…Ø§Ù† Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ'
        # ]
        # df = df[~df['Ù…ÙˆØ¶ÙˆØ¹'].isin(irrelevant_topics)]

        heatmap_index =[
                        "Ú©ÙØ§ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡",
                        "Ù†Ø³Ø¨Øªâ€ŒÙ‡Ø§ Ø¯Ø± Ú†Ø§Ø±Ú†ÙˆØ¨ Ø¨Ø§Ø²Ù„",
                        "Ø±ÛŒØ³Ú© Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ",
                        "Ù…Ø¯ÛŒØ±ÛŒØª Ø¯Ø§Ø±Ø§ÛŒÛŒ Ùˆ Ø¨Ø¯Ù‡ÛŒ (ALM)",
                        "Ø±ÛŒØ³Ú© Ù†Ø±Ø® Ø¨Ù‡Ø±Ù‡",
                        "ØªÙ…Ø±Ú©Ø² Ø±ÛŒØ³Ú© Ø§Ø¹ØªØ¨Ø§Ø±ÛŒ",
                        "Ø°Ø®ÛŒØ±Ù‡â€ŒÚ¯ÛŒØ±ÛŒ (Ú©Ù„ÛŒ)",
                        "ØµÙˆØ±Øª Ø¬Ø±ÛŒØ§Ù† ÙˆØ¬ÙˆÙ‡ Ù†Ù‚Ø¯",
                        "Ú©Ù†ØªØ±Ù„â€ŒÙ‡Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ Ùˆ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ Ø¯Ø§Ø®Ù„ÛŒ",
                        "Ø­Ø§Ú©Ù…ÛŒØª Ø´Ø±Ú©ØªÛŒ",
                        "Ø§ÙˆØ±Ø§Ù‚ Ø¨Ù‡Ø§Ø¯Ø§Ø± Ùˆ Ø³Ø±Ù…Ø§ÛŒÙ‡â€ŒÚ¯Ø°Ø§Ø±ÛŒâ€ŒÙ‡Ø§",
                        "ØªØ³Ø¹ÛŒØ± Ø§Ø±Ø² Ùˆ Ø¹Ù…Ù„ÛŒØ§Øª Ø®Ø§Ø±Ø¬ÛŒ",
                        "ØªØ¹Ù‡Ø¯Ø§Øª Ø§Ø±Ø²ÛŒ Ùˆ Ø§Ø®ØªÙ„Ø§Ù Ø¨Ø§ Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒ",
                        "Ø°Ø®ÛŒØ±Ù‡ Ù…Ø·Ø§Ù„Ø¨Ø§Øª Ù…Ø´Ú©ÙˆÚ©â€ŒØ§Ù„ÙˆØµÙˆÙ„",
                        "Ù…Ø·Ø§Ù„Ø¨Ø§Øª Ù…Ø´Ú©ÙˆÚ©â€ŒØ§Ù„ÙˆØµÙˆÙ„",
                        "ØªØ³Ù‡ÛŒÙ„Ø§Øª Ùˆ Ø§Ø¹ØªØ¨Ø§Ø±Ø§Øª",
                        "Ø³Ø±Ù…Ø§ÛŒÙ‡â€ŒÚ¯Ø°Ø§Ø±ÛŒ Ø¯Ø± Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ ÙˆØ§Ø¨Ø³ØªÙ‡",
                        "Ú©Ø§Ù‡Ø´ Ø§Ø±Ø²Ø´ Ø¯Ø§Ø±Ø§ÛŒÛŒâ€ŒÙ‡Ø§",
                        "Ø§ÙØ´Ø§ÛŒ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¹Ù…Ù„ÛŒØ§ØªÛŒ",
                        "Ù†Ø³Ø¨Øªâ€ŒÙ‡Ø§ÛŒ Ø¨Ø¯Ù‡ÛŒ Ùˆ Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ",
                        "Ù†Ø³Ø¨Øª Ú©ÙØ§ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡",
                        "Ù…Ø¹Ø§Ù…Ù„Ø§Øª Ø¨Ø§ Ø§Ø´Ø®Ø§Øµ ÙˆØ§Ø¨Ø³ØªÙ‡",
                        "ØªØ¯Ø§ÙˆÙ… ÙØ¹Ø§Ù„ÛŒØª",
                        "Ø§Ù†Ø·Ø¨Ø§Ù‚ Ø¨Ø§ Ù…Ù‚Ø±Ø±Ø§Øª Ø¶Ø¯Ù¾ÙˆÙ„Ø´ÙˆÛŒÛŒ (AML/CFT)"
]

        heatmap_data = df.pivot_table(index='Ù…ÙˆØ¶ÙˆØ¹', columns='year', values='status_mp', aggfunc='max').fillna(0).astype(int)
    # Ø§ÛŒÙ† Ú©Ø§Ø± ØªØ±ØªÛŒØ¨ Ø±Ø§ Ø­ÙØ¸ Ù…ÛŒâ€ŒÚ©Ù†Ø¯ Ùˆ Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ù†Ø§Ù…ÙˆØ¬ÙˆØ¯ Ø±Ø§ Ø­Ø°Ù Ù…ÛŒâ€ŒÚ©Ù†Ø¯
        filtered_index = [topic for topic in heatmap_index if topic in heatmap_data.index]

        # Reindex Ø¨Ø§ Ù„ÛŒØ³Øª ÙÛŒÙ„ØªØ± Ø´Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² ØªØ±ØªÛŒØ¨ ØµØ­ÛŒØ­
        heatmap_data = heatmap_data.reindex(filtered_index)

        custom_colors = ["#E0E7FF", "#8EA4E9", "#A970DE", "#F46E52"]
        custom_cmap = LinearSegmentedColormap.from_list("modern_risk", custom_colors, N=4)
        norm = BoundaryNorm(boundaries=[-0.5, 0.5, 1.5, 2.5, 3.5], ncolors=4)
        
        # âœ… Ø³Ø§ÛŒØ² Ø¨Ø²Ø±Ú¯ Ø¨Ø±Ø§ÛŒ heatmap: 16x10
        fig, ax = plt.subplots(figsize=(12, 7.5), facecolor='#f8f9fa')
        
        sns.heatmap(
            heatmap_data, annot=False, cmap=custom_cmap, norm=norm,
            linewidths=2.5, linecolor='white', fmt='d', cbar=True, ax=ax,
            annot_kws={'fontsize': 11, 'weight': 'bold'},
            cbar_kws={'shrink': 0.8}
        )
        
        # ax.set_title(process_persian_text('Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ú©Ù„ÛŒØ¯ÛŒ Ø¯Ø± Ø·ÙˆÙ„ Ø²Ù…Ø§Ù†'),
        #             fontproperties=font, size=20, weight='bold', pad=25, color='#2c3e50')
        ax.set_xlabel(process_persian_text('Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ'), fontproperties=font, size=14, weight='bold', color='#34495e')
        ax.set_ylabel(process_persian_text('Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ'), fontproperties=font, size=14, weight='bold', color='#34495e')
        
        y_labels = [process_persian_text(label) for label in heatmap_data.index]
        x_labels = [process_persian_text(str(int(label))) for label in heatmap_data.columns]
        ax.set_yticklabels(y_labels, fontproperties=font, rotation=0, fontsize=11)
        ax.set_xticklabels(x_labels, fontproperties=font, rotation=0, fontsize=12)
        
        cbar = ax.collections[0].colorbar
        cbar.set_ticks(list(status_mapping.values()))
        cbar_labels = [process_persian_text(label) for label in status_mapping.keys()]
        cbar.set_ticklabels(cbar_labels, fontproperties=font, fontsize=11)
        cbar.ax.set_title(process_persian_text('Ø³Ø·Ø­ Ø±ÛŒØ³Ú©'), fontproperties=font, fontsize=13, pad=15)
        cbar.outline.set_linewidth(2)
        cbar.outline.set_edgecolor('#bdc3c7')
        
        # for spine in ax.spines.values():
        #     spine.set_visible(False)
        
        plt.tight_layout()
        return fig
    

    def plot_risk_stacked_bar(df, font):
        """Ù†Ù…ÙˆØ¯Ø§Ø± Ø³ØªÙˆÙ†ÛŒ Ø§Ù†Ø¨Ø§Ø´ØªÙ‡ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ - ÙÙ‚Ø· Ù…ÙˆØ§Ø±Ø¯ Ø¯Ø§Ø±Ø§ÛŒ Ù…ÙˆØ¶ÙˆØ¹ÛŒØª"""
        

        df = df[df.get('Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯', True) == True].copy()
        
        if df.empty:
            # Ø§Ú¯Ø± Ø¯Ø§Ø¯Ù‡â€ŒØ§ÛŒ Ù†Ø¨ÙˆØ¯ØŒ ÛŒÚ© figure Ø®Ø§Ù„ÛŒ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†
            fig, ax = plt.subplots(figsize=(12, 10), facecolor='#f8f9fa')
            ax.text(0.5, 0.5, process_persian_text('Ø¯Ø§Ø¯Ù‡â€ŒØ§ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯'), 
                    ha='center', va='center', fontsize=16, fontproperties=font)
            ax.axis('off')
            return fig
        
        risk_over_time = pd.crosstab(df['year'], df['Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ_Ø±ÛŒØ³Ú©'])
        
        fig, ax = plt.subplots(figsize=(12, 10), facecolor='#f8f9fa')
        ax.set_facecolor('#ffffff')
        
        colors = ['#3498db', '#e74c3c', '#2ecc71', '#f39c12', '#9b59b6', '#1abc9c', '#e67e22']
        risk_over_time.plot(kind='bar', stacked=True, color=colors, ax=ax, width=0.8, edgecolor='white', linewidth=2)
        
        for c in ax.containers:
            labels = [int(v.get_height()) if v.get_height() > 0 else '' for v in c]
            ax.bar_label(c, labels=[f'{v}' if v else '' for v in labels], 
                        label_type='center', color='white', weight='bold', fontsize=20, fontproperties=font)
        
        # ax.set_title(process_persian_text('Ø±ÙˆÙ†Ø¯ ØªØ¹Ø¯Ø§Ø¯ Ùˆ ØªØ±Ú©ÛŒØ¨ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ Ø¯Ø± Ø·ÙˆÙ„ Ø²Ù…Ø§Ù†'), 
        #             fontproperties=font, size=18, weight='bold', pad=20, color='#2c3e50')
        ax.set_xlabel(process_persian_text('Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ'), fontproperties=font, size=20, weight='bold', color='#34495e')
        ax.set_ylabel(process_persian_text('ØªØ¹Ø¯Ø§Ø¯ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡ Ø´Ø¯Ù‡'), fontproperties=font, size=25, weight='bold', color='#34495e')
        ax.tick_params(axis='x', rotation=0, labelsize=20)
        ax.tick_params(axis='y', labelsize=20)
        
        ax.grid(True, axis='y', linestyle='--', alpha=0.3, color='#bdc3c7')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#bdc3c7')
        ax.spines['bottom'].set_color('#bdc3c7')
        
        legend = ax.get_legend()
        # legend.set_title(process_persian_text('Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ø±ÛŒØ³Ú©'), prop=font)
        title_font = font.copy()
        title_font.set_size(17) 
        legend.set_title(process_persian_text('Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ø±ÛŒØ³Ú©'), prop=title_font)
        legend.set_frame_on(True)
        legend.get_frame().set_facecolor('#ffffff')
        legend.get_frame().set_edgecolor('#bdc3c7')
        legend.get_frame().set_linewidth(1.5)
        for text in legend.get_texts():
            text.set_text(process_persian_text(text.get_text()))
            text.set_fontproperties(font)
            text.set_fontsize(17)
        # 6. Ø¬Ø§Ø¨Ø¬Ø§ÛŒÛŒ legend Ø¨Ù‡ Ø¨ÛŒØ±ÙˆÙ† Ø§Ø² Ù†Ù…ÙˆØ¯Ø§Ø±
        # bbox_to_anchor=(1.02, 1) Ø¨Ù‡ Ù…Ø¹Ù†ÛŒ:
        # 1.02: Ú©Ù…ÛŒ Ø¨Ù‡ Ø³Ù…Øª Ø±Ø§Ø³Øª Ø¨ÛŒØ±ÙˆÙ† Ø§Ø² Ù…Ø­ÙˆØ± x Ù†Ù…ÙˆØ¯Ø§Ø± (102%)
        # 1: Ø¯Ø± Ø¨Ø§Ù„Ø§ÛŒ Ù…Ø­ÙˆØ± y Ù†Ù…ÙˆØ¯Ø§Ø± (100%)
        # loc='upper left' Ø¨Ù‡ matplotlib Ù…ÛŒâ€ŒÚ¯ÙˆÛŒØ¯ Ú©Ù‡ Ù†Ù‚Ø·Ù‡ Ø¨Ø§Ù„Ø§-Ú†Ù¾ Ú©Ø§Ø¯Ø± legend Ø±Ø§ Ø¯Ø± Ø¢Ù† Ù…Ø®ØªØµØ§Øª Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡.
#         legend.set_bbox_to_anchor((1.02, 1))
# rect=[0, 0, 0.85, 1]
        # plt.tight_layout()
        plt.subplots_adjust(right=0.82)
        return fig


    def plot_risk_sunburst(df):
        """ğŸ“Š Ù†Ù…ÙˆØ¯Ø§Ø± Sunburst Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ - Ù†Ø³Ø®Ù‡ Ù†Ù‡Ø§ÛŒÛŒ Ø¨Ø§ Ø·Ø±Ø§Ø­ÛŒ Ù…Ø´Ø§Ø¨Ù‡ ØªØ®Ù„ÙØ§Øª Ùˆ Ù…Ø±Ú©Ø² Ø¨Ø§ Ø¹Ù†ÙˆØ§Ù† 'Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡'"""

        sunburst_data = df.groupby(['Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ_Ø±ÛŒØ³Ú©', 'Ø²ÛŒØ±Ø´Ø§Ø®Ù‡_Ø±ÛŒØ³Ú©']).size().reset_index(name='ØªØ¹Ø¯Ø§Ø¯')

        fig = px.sunburst(
            sunburst_data,
            path=[px.Constant("Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡"), 'Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ_Ø±ÛŒØ³Ú©', 'Ø²ÛŒØ±Ø´Ø§Ø®Ù‡_Ø±ÛŒØ³Ú©'],
            values='ØªØ¹Ø¯Ø§Ø¯',
            color='Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ_Ø±ÛŒØ³Ú©',
            color_discrete_sequence=px.colors.qualitative.Set3,
            hover_data={'ØªØ¹Ø¯Ø§Ø¯': True}
        )

        fig.update_layout(
            width=900,
            height=700,
            # title={
            #     'text': 'ØªØ­Ù„ÛŒÙ„ Ø³Ù„Ø³Ù„Ù‡â€ŒÙ…Ø±Ø§ØªØ¨ÛŒ ØªÙˆØ²ÛŒØ¹ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§',
            #     'y': 0.98,
            #     'x': 0.5,
            #     'xanchor': 'center',
            #     'yanchor': 'top',
            #     'font': {'size': 22, 'family': 'Tahoma', 'color': '#2c3e50'}
            # },
            font_family="Noto Naskh Arabic",
            font_size=13,
            margin=dict(t=80, l=40, r=40, b=40),
            paper_bgcolor='#f8f9fa',
            plot_bgcolor='#ffffff',
            hovermode='closest',
            hoverlabel=dict(
                bgcolor="white",
                font_size=20,
                font_family="Noto Naskh Arabic",
                align="right"
            ),
            # âœ… Ø§ÙØ²ÙˆØ¯Ù† Ù…ØªÙ† Ù…Ø±Ú©Ø²

        )

        # ØªÙ†Ø¸ÛŒÙ…Ø§Øª ØªØ±Ø³ÛŒÙ… Ø¬Ø²Ø¦ÛŒØ§Øª Ù†Ù…ÙˆØ¯Ø§Ø±
        fig.update_traces(
            textinfo="label",
            insidetextorientation='horizontal',  # Ø§ÙÙ‚ÛŒ Ø¨Ø±Ø§ÛŒ Ø®ÙˆØ§Ù†Ø§ÛŒÛŒ Ø¨Ù‡ØªØ±
            marker=dict(
                line=dict(color='white', width=2.5)
            ),
            textfont=dict(
                size=20,
                family="Noto Naskh Arabic",
                color="black"
            ),
            hovertemplate='<b style="font-family:Noto Naskh Arabic">%{label}</b><br>' +
                        '<span style="font-family:Noto Naskh Arabic">ØªØ¹Ø¯Ø§Ø¯: %{value}</span><br>' +
                        '<span style="font-family:Noto Naskh Arabic">Ù†Ø³Ø¨Øª: %{percentRoot:.1%}</span>' +
                        '<extra></extra>',
            hoverlabel=dict(
                bgcolor="rgba(255,255,255,0.95)",
                bordercolor="#2c3e50",
                font=dict(family="Noto Naskh Arabic", size=13, color="black")
            )
        )

        return fig



    def plot_violations_stacked_bar(df, font):
        """Ù†Ù…ÙˆØ¯Ø§Ø± Ø³ØªÙˆÙ†ÛŒ Ø§Ù†Ø¨Ø§Ø´ØªÙ‡ ØªØ®Ù„ÙØ§Øª - ÙÙ‚Ø· Ù…ÙˆØ§Ø±Ø¯ Ø¯Ø§Ø±Ø§ÛŒ Ù…ÙˆØ¶ÙˆØ¹ÛŒØª"""
        
        # âœ… ÙÛŒÙ„ØªØ± Ú©Ø±Ø¯Ù† ÙÙ‚Ø· Ù…ÙˆØ§Ø±Ø¯ Ø¯Ø§Ø±Ø§ÛŒ Ù…ÙˆØ¶ÙˆØ¹ÛŒØª
        df = df[df.get('Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯', True) == True].copy()
        
        if df.empty:
            # Ø§Ú¯Ø± Ø¯Ø§Ø¯Ù‡â€ŒØ§ÛŒ Ù†Ø¨ÙˆØ¯ØŒ ÛŒÚ© figure Ø®Ø§Ù„ÛŒ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†
            fig, ax = plt.subplots(figsize=(12,10), facecolor='#f8f9fa')
            ax.text(0.5, 0.5, process_persian_text('Ø¯Ø§Ø¯Ù‡â€ŒØ§ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯'), 
                    ha='center', va='center', fontsize=16, fontproperties=font)
            ax.axis('off')
            return fig
        
        violation_counts = pd.crosstab(df['year'], df['Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ'])
        
        fig, ax = plt.subplots(figsize=(12,10), facecolor='#f8f9fa')
        ax.set_facecolor('#ffffff')
        
        colors = ['#e74c3c', '#3498db', '#2ecc71', '#f39c12', '#9b59b6', '#1abc9c']
        violation_counts.plot(kind='bar', stacked=True, color=colors, ax=ax, width=0.8, edgecolor='white', linewidth=2)
        
        for c in ax.containers:
            labels = [int(v.get_height()) if v.get_height() > 0 else '' for v in c]
            ax.bar_label(c, labels=[f'{v}' if v else '' for v in labels], 
                        label_type='center', color='white', weight='bold', fontsize=11, fontproperties=font)
        
        # ax.set_title(process_persian_text('Ø±ÙˆÙ†Ø¯ ØªØ¹Ø¯Ø§Ø¯ Ùˆ ØªØ±Ú©ÛŒØ¨ ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ø¯Ø± Ø·ÙˆÙ„ Ø²Ù…Ø§Ù†'), 
        #             fontproperties=font, size=18, weight='bold', pad=20, color='#2c3e50')
        ax.set_xlabel(process_persian_text('Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ'), fontproperties=font, size=20, weight='bold', color='#34495e')
        ax.set_ylabel(process_persian_text('ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ§Ø±Ø¯ Ø¹Ø¯Ù… Ø±Ø¹Ø§ÛŒØª'), fontproperties=font, size=25, weight='bold', color='#34495e')
        ax.tick_params(axis='x', rotation=0, labelsize=20)
        ax.tick_params(axis='y', labelsize=20)
        
        ax.grid(True, axis='y', linestyle='--', alpha=0.3, color='#bdc3c7')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#bdc3c7')
        ax.spines['bottom'].set_color('#bdc3c7')
        
        legend = ax.get_legend()
        # legend.set_title(process_persian_text('Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ ØªØ®Ù„Ù'), prop=font)
        title_font = font.copy()
        title_font.set_size(16) 
        legend.set_title(process_persian_text('Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ø±ÛŒØ³Ú©'), prop=title_font)
        legend.set_frame_on(True)
        legend.set_loc('upper right')
        legend.get_frame().set_facecolor('#ffffff')
        legend.get_frame().set_edgecolor('#bdc3c7')
        legend.get_frame().set_linewidth(1.5)
        for text in legend.get_texts():
            text.set_text(process_persian_text(text.get_text()))
            text.set_fontproperties(font)
            text.set_fontsize(15)

        plt.subplots_adjust(right=0.82)

        return fig
    

    
    def plot_violations_sunburst(df):
        """Ù†Ù…ÙˆØ¯Ø§Ø± Sunburst ØªØ®Ù„ÙØ§Øª - Ù†Ø³Ø®Ù‡ Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡ Ø¨Ø§ Ù…ØªÙ† ÙˆØ§Ø¶Ø­"""
        
        sunburst_data = df.groupby(['Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ', 'Ø²ÛŒØ±Ø´Ø§Ø®Ù‡']).size().reset_index(name='ØªØ¹Ø¯Ø§Ø¯')

        fig = px.sunburst(
            sunburst_data,
            path=[px.Constant("ØªÙ…Ø§Ù… ØªØ®Ù„ÙØ§Øª"), 'Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ', 'Ø²ÛŒØ±Ø´Ø§Ø®Ù‡'],
            values='ØªØ¹Ø¯Ø§Ø¯',
            color='Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ',
            # title='ØªØ­Ù„ÛŒÙ„ Ø³Ø§Ø®ØªØ§Ø± Ùˆ ØªÙˆØ²ÛŒØ¹ ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ',
            color_discrete_sequence=px.colors.qualitative.Set3,
            hover_data={'ØªØ¹Ø¯Ø§Ø¯': True}
        )
        
        fig.update_layout(
            width=900,
            height=700,
            # title={
            #     'text': 'ØªØ­Ù„ÛŒÙ„ Ø³Ø§Ø®ØªØ§Ø± Ùˆ ØªÙˆØ²ÛŒØ¹ ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ',
            #     'y': 0.98,
            #     'x': 0.5,
            #     'xanchor': 'center',
            #     'yanchor': 'top',
            #     'font': {'size': 22, 'family': 'Tahoma', 'color': '#2c3e50'}
            # },
            font_family="Noto Naskh Arabic",
            font_size=13,
            margin=dict(t=80, l=40, r=40, b=40),
            paper_bgcolor='#f8f9fa',
            plot_bgcolor='#ffffff',
            # âœ… ØªÙ†Ø¸ÛŒÙ…Ø§Øª Ù…Ù‡Ù… Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ ØµØ­ÛŒØ­ hover
            hovermode='closest',
            hoverlabel=dict(
                bgcolor="white",
                font_size=20,
                font_family="Noto Naskh Arabic",
                align="right"
            )
        )
        
        fig.update_traces(
            textinfo="label",
            insidetextorientation='horizontal',  # âœ… Ø§ÙÙ‚ÛŒ Ø¨Ø±Ø§ÛŒ Ø®ÙˆØ§Ù†Ø§ÛŒÛŒ Ø¨Ù‡ØªØ±
            marker=dict(
                line=dict(color='white', width=2.5)
            ),
            textfont=dict(
                size=20,  # âœ… ÙÙˆÙ†Øª Ù…ØªÙˆØ³Ø· Ø¨Ø±Ø§ÛŒ Ø®ÙˆØ§Ù†Ø§ÛŒÛŒ Ø¨Ù‡ØªØ±
                family="Noto Naskh Arabic",
                color="black"
            ),
            # âœ… Hover template Ø³Ø§Ø¯Ù‡ Ùˆ ÙˆØ§Ø¶Ø­ Ø¨Ø§ ÙØ§Ø±Ø³ÛŒ
            hovertemplate='<b style="font-family:Noto Naskh Arabic">%{label}</b><br>' +
                        '<span style="font-family:Noto Naskh Arabic">ØªØ¹Ø¯Ø§Ø¯: %{value}</span><br>' +
                        '<span style="font-family:Noto Naskh Arabic">Ù†Ø³Ø¨Øª: %{percentRoot:.1%}</span>' +
                        '<extra></extra>',
            hoverlabel=dict(
                bgcolor="rgba(255,255,255,0.95)",
                bordercolor="#2c3e50",
                font=dict(family="Noto Naskh Arabic", size=13, color="black")
            )
        )
        
        return fig
    # ========================================================================
    # Ø¨Ø®Ø´ Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡: Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡Ù…Ø²Ù…Ø§Ù† ÙØ§ÛŒÙ„â€ŒÙ‡Ø§
    # ========================================================================

    class FinancialAnalyzer:
        """Ú©Ù„Ø§Ø³ ØªØ­Ù„ÛŒÙ„Ú¯Ø± Ù…Ø§Ù„ÛŒ Ø¨Ø§ Ù¾Ø´ØªÛŒØ¨Ø§Ù†ÛŒ Ø§Ø² Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡Ù…Ø²Ù…Ø§Ù†"""
        
        def __init__(self):
            # Schema Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ±
            self.response_schema = {
                "type": "object",
                "properties": {
                    "ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ": {
                        "type": "object",
                        "description": "Ø³Ø§Ø®ØªØ§Ø± Ø§ØµÙ„ÛŒ Ú©Ù‡ ØªØ­Ù„ÛŒÙ„ Ú©Ø§Ù…Ù„ Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ø±Ø§ Ø¯Ø± Ø®ÙˆØ¯ Ø¬Ø§ÛŒ Ù…ÛŒâ€ŒØ¯Ù‡Ø¯.",
                        "properties": {
                            "Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ": {
                                "type": "object",
                                "description": "Ø´Ø§Ù…Ù„ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ø§ÙˆÙ„ÛŒÙ‡ Ú¯Ø²Ø§Ø±Ø´ Ùˆ Ù†ØªÛŒØ¬Ù‡â€ŒÚ¯ÛŒØ±ÛŒâ€ŒÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ Ø¯Ø± ÛŒÚ© Ù†Ú¯Ø§Ù‡.",
                                "properties": {
                                    "Ù†Ø§Ù…_Ø´Ø±Ú©Øª": {"type": "string", "description": "Ù†Ø§Ù… Ú©Ø§Ù…Ù„ Ø´Ø±Ú©Øª Ø§Ø² Ø±ÙˆÛŒ Ø¬Ù„Ø¯ Ú¯Ø²Ø§Ø±Ø´."},
                                    "Ù†Ø§Ù…_Ø­Ø³Ø§Ø¨Ø±Ø³": {"type": "string", "description": "Ù†Ø§Ù… Ù…ÙˆØ³Ø³Ù‡ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ."},
                                    "Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ": {"type": "string", "description": "Ø¯ÙˆØ±Ù‡ Ù…Ø§Ù„ÛŒ Ù…ÙˆØ±Ø¯ Ø±Ø³ÛŒØ¯Ú¯ÛŒØŒ Ù…Ø«Ù„Ø§: 'Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ Ù…Ù†ØªÙ‡ÛŒ Ø¨Ù‡ Û²Û¹ Ø§Ø³ÙÙ†Ø¯ Û±Û³Û¹Û¸'."},
                                    "Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±": {
                                        "type": "string",
                                        "description": "ÛŒÚ©ÛŒ Ø§Ø² Ù…ÙˆØ§Ø±Ø¯: Ù…Ù‚Ø¨ÙˆÙ„ØŒ Ù…Ø´Ø±ÙˆØ·ØŒ Ù…Ø±Ø¯ÙˆØ¯ØŒ Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±.",
                                        "enum": ["Ù…Ù‚Ø¨ÙˆÙ„", "Ù…Ø´Ø±ÙˆØ·", "Ù…Ø±Ø¯ÙˆØ¯", "Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±"]
                                    },
                                    "Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_Ú©Ù„ÛŒ_Ø¨Ù†Ø§_Ø¨Ù‡_Ú¯Ø²Ø§Ø±Ø´": {
                                        "type": "string",
                                        "description": "Ø³Ø·Ø­ Ø±ÛŒØ³Ú© Ú©Ù„ÛŒ Ø§Ø³ØªÙ†Ø¨Ø§Ø· Ø´Ø¯Ù‡ Ø§Ø² Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ø¨Ù†Ø§ Ø¨Ù‡ Ù…ØªÙ† Ú¯Ø²Ø§Ø±Ø´ Ùˆ Ø´ÙˆØ§Ù‡Ø¯ Ùˆ Ø¢Ù…Ø§Ø±Ù‡ Ù‡Ø§ÛŒ Ø¨ÛŒØ§Ù† Ø´Ø¯Ù‡ Ø§Ø² Ø¯ÛŒØ¯Ú¯Ø§Ù‡ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ",
                                        "enum": ["Ù¾Ø§ÛŒÛŒÙ†", "Ù…ØªÙˆØ³Ø·", "Ø¨Ø§Ù„Ø§", "Ø¨Ø­Ø±Ø§Ù†ÛŒ"]
                                    },
                                    "Ø¬Ø²ÛŒÛŒØ§Øª_Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_ØªØ¹ÛŒÛŒÙ†_Ø´Ø¯Ù‡": {
                                        "type": "string",
                                        "description": "Ø¬Ø²ÛŒÛŒØ§Øª Ùˆ Ø¯Ù„ÛŒÙ„ Ø³Ø·Ø­ Ø±ÛŒØ³Ú© Ú©Ù„ÛŒ Ø§Ø³ØªÙ†Ø¨Ø§Ø· Ø´Ø¯Ù‡ Ø§Ø² Ú¯Ø²Ø§Ø±Ø´."
                                    },
                                    "Ù†Ú©Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ_Ùˆ_Ù†ØªÛŒØ¬Ù‡_Ú¯ÛŒØ±ÛŒ": {
                                        "type": "array",
                                        "description": "Ø¢Ø±Ø§ÛŒÙ‡â€ŒØ§ÛŒ Ø§Ø² Û³ Ø±Ø´ØªÙ‡ Ø´Ø§Ù…Ù„ Ù…Ù‡Ù…â€ŒØªØ±ÛŒÙ† ÛŒØ§ÙØªÙ‡â€ŒÙ‡Ø§ Ùˆ Ù†ØªÛŒØ¬Ù‡â€ŒÚ¯ÛŒØ±ÛŒâ€ŒÙ‡Ø§.",
                                        "items": {"type": "string"}
                                    }
                                },
                                "required": [
                                    "Ù†Ø§Ù…_Ø´Ø±Ú©Øª", "Ù†Ø§Ù…_Ø­Ø³Ø§Ø¨Ø±Ø³", "Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ", "Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±",
                                    "Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_Ú©Ù„ÛŒ_Ø¨Ù†Ø§_Ø¨Ù‡_Ú¯Ø²Ø§Ø±Ø´", "Ø¬Ø²ÛŒÛŒØ§Øª_Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_ØªØ¹ÛŒÛŒÙ†_Ø´Ø¯Ù‡",
                                    "Ù†Ú©Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ_Ùˆ_Ù†ØªÛŒØ¬Ù‡_Ú¯ÛŒØ±ÛŒ"
                                ]
                            },
                            "Ø¨Ø®Ø´Û²_ØªØ¬Ø²ÛŒÙ‡_ØªØ­Ù„ÛŒÙ„_Ú¯Ø²Ø§Ø±Ø´": {
                                "type": "object",
                                "description": "ØªØ¬Ø²ÛŒÙ‡ Ùˆ ØªØ­Ù„ÛŒÙ„ Ø³Ø§Ø®ØªØ§Ø±ÛŒØ§ÙØªÙ‡ Ù…ØªÙ† Ú¯Ø²Ø§Ø±Ø´ØŒ Ø¨Ù†Ø¯ Ø¨Ù‡ Ø¨Ù†Ø¯.",
                                "properties": {
                                    "Ø¨Ù†Ø¯_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±": {
                                        "type": "object",
                                        "properties": {
                                            "Ù†ÙˆØ¹": {"type": "string"},
                                            "Ø®Ù„Ø§ØµÙ‡_Ø¯Ù„Ø§ÛŒÙ„": {"type": "string"}
                                        },
                                        "required": ["Ù†ÙˆØ¹", "Ø®Ù„Ø§ØµÙ‡_Ø¯Ù„Ø§ÛŒÙ„"]
                                    },
                                    "Ø¨Ù†Ø¯_Ù…Ø¨Ø§Ù†ÛŒ_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±": {
                                        "type": "object",
                                        "properties": {
                                            "Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯": {"type": "boolean"},
                                            "Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡": {
                                                "type": "array",
                                                "items": {
                                                    "type": "object",
                                                    "properties": {
                                                        "Ø´Ù…Ø§Ø±Ù‡_Ù…ÙˆØ±Ø¯": {"type": "integer"},
                                                        "Ø¹Ù†ÙˆØ§Ù†": {"type": "string"},
                                                        "Ø´Ø±Ø­": {"type": "string"},
                                                        "Ù†ÙˆØ¹_Ø¯Ù„ÛŒÙ„": {
                                                            "type": "string",
                                                            "enum": ["Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø¯Ø± Ø±Ø³ÛŒØ¯Ú¯ÛŒ", "Ø§Ù†Ø­Ø±Ø§Ù Ø§Ø² Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§ÛŒ Ø­Ø³Ø§Ø¨Ø¯Ø§Ø±ÛŒ", "Ø³Ø§ÛŒØ±"]
                                                        }
                                                    },
                                                    "required": ["Ø´Ù…Ø§Ø±Ù‡_Ù…ÙˆØ±Ø¯", "Ø¹Ù†ÙˆØ§Ù†", "Ø´Ø±Ø­", "Ù†ÙˆØ¹_Ø¯Ù„ÛŒÙ„"]
                                                }
                                            }
                                        },
                                        "required": ["Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯"]
                                    },
                                    "Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ": {
                                        "type": "object",
                                        "properties": {
                                            "Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯": {"type": "boolean"},
                                            "Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡": {
                                                "type": "array",
                                                "items": {
                                                    "type": "object",
                                                    "properties": {
                                                        "Ø§Ø±Ø¬Ø§Ø¹": {
                                                            "type": "object",
                                                            "properties": {
                                                                "Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯": {
                                                                    "type": "string",
                                                                    "description": "Ø´Ù…Ø§Ø±Ù‡ Ø¨Ù†Ø¯ Ù…Ø±Ø¨ÙˆØ·Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ .Ø¨ÛŒÙ† Ø¨Ù†Ø¯ Ù‡Ø§ , Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡ Ù…Ø§Ù†Ù†Ø¯ Û²,Û¶"
                                                                    },
                                                                "Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡": {
                                                                    "type": "string",
                                                                "description": "Ø´Ù…Ø§Ø±Ù‡ ØµÙØ­Ù‡ Ù…Ø±Ø¨ÙˆØ·Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ.Ú†Ù†Ø§Ù†Ú†Ù‡ Ø§ÛŒÙ† Ù…ÙˆØ±Ø¯ Ø¯Ø± Ú†Ù†Ø¯ Ø¨Ù†Ø¯ Ø¨Ù‡ Ø§Ù† Ø§Ø´Ø§Ø±Ù‡ Ø´Ø¯Ù‡ ØµÙØ­Ø§Øª Ù…Ù†Ø·Ø¨Ù‚ Ø¨Ø§ Ø¨Ù†Ø¯ Ø±Ø§ Ø¨Ù‡ ØªØ±ØªÛŒØ¨ Ø¨Ù†Ø¯ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù† Ø¨ÛŒÙ† ØµÙØ­Ø§Øª , Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡ Ù…Ø§Ù†Ù†Ø¯ Û±,Ûµ"
                                                                    }
                                                            },
                                                            "required": ["Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯", "Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡"]
                                                        },
                                                        "Ø¹Ù†ÙˆØ§Ù†": {"type": "string"},
                                                        "Ø´Ø±Ø­": {"type": "string"},
                                                        "Ø±ÛŒØ³Ú©_Ø¨Ø±Ø¬Ø³ØªÙ‡_Ø´Ø¯Ù‡": {"type": "string"},
                                                        "Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ_Ø±ÛŒØ³Ú©": {
                                                            "type": "string",
                                                            "enum": ["Ø±ÛŒØ³Ú© Ø§Ø¹ØªØ¨Ø§Ø±ÛŒ", "Ø±ÛŒØ³Ú© Ø¨Ø§Ø²Ø§Ø±", "Ø±ÛŒØ³Ú© Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ", "Ø±ÛŒØ³Ú© Ø¹Ù…Ù„ÛŒØ§ØªÛŒ", "Ø±ÛŒØ³Ú© Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ùˆ ØªØ·Ø¨ÛŒÙ‚", "Ø±ÛŒØ³Ú© Ø§Ø³ØªØ±Ø§ØªÚ˜ÛŒÚ©", "Ø±ÛŒØ³Ú© Ø´Ù‡Ø±Øª"]
                                                        },
                                                        "Ø²ÛŒØ±Ø´Ø§Ø®Ù‡_Ø±ÛŒØ³Ú©": {
                                                            "type": "string",
                                                            "description": """ Ø²ÛŒØ±Ø´Ø§Ø®Ù‡ Ø¯Ù‚ÛŒÙ‚ Ø±ÛŒØ³Ú© Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡ Ù…Ø±ØªØ¨Ø· Ø¨Ø§ Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡ .(Ø¯Ù‚ÛŒÙ‚Ø§ Ø§Ø² Ø§ÛŒÙ† Ù…ØµØ§Ø¯ÛŒÙ‚ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø´ÙˆØ¯. Ù…ØµØ§Ø¯Ù‚ Ù…Ù†Ø·Ø¨Ù‚ Ø¨Ø§ Ø§Ù„Ú¯ÙˆÛŒ Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ:[Ø²ÛŒØ±Ø´Ø§Ø®Ù‡ Ù‡Ø§] Ù…Ø§Ù†Ù†Ø¯ 
                                                                                                                                                                                            
                                                                            Û±. Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ø§Ø¹ØªØ¨Ø§Ø±ÛŒ:
                                                                            [Ø±ÛŒØ³Ú© Ù†Ú©ÙˆÙ„ ,
                                                                            Ø±ÛŒØ³Ú© ØªÙ…Ø±Ú©Ø² ,
                                                                            Ø±ÛŒØ³Ú© Ø·Ø±Ù Ù‚Ø±Ø§Ø±Ø¯Ø§Ø¯ ,
                                                                            Ø±ÛŒØ³Ú© Ú©Ø´ÙˆØ± ,
                                                                            Ø±ÛŒØ³Ú© ØªØ¶Ø¹ÛŒÙ ÙˆØ«Ø§ÛŒÙ‚ ,
                                                                            Ø±ÛŒØ³Ú© ÙˆØµÙˆÙ„ Ù…Ø·Ø§Ù„Ø¨Ø§Øª ]
                                                                            ,

                                                                            Û². Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ø¨Ø§Ø²Ø§Ø±:
                                                                            [Ø±ÛŒØ³Ú© Ù†Ø±Ø® Ø§Ø±Ø² ,
                                                                            Ø±ÛŒØ³Ú© Ù†Ø±Ø® Ø³ÙˆØ¯ ,
                                                                            Ø±ÛŒØ³Ú© Ù‚ÛŒÙ…Øª Ú©Ø§Ù„Ø§ ,
                                                                            Ø±ÛŒØ³Ú© Ù‚ÛŒÙ…Øª Ø³Ù‡Ø§Ù… ,
                                                                            Ø±ÛŒØ³Ú© Ù†ÙˆØ³Ø§Ù† Ø§Ø±Ø²Ø´ Ø³Ø±Ù…Ø§ÛŒÙ‡â€ŒÚ¯Ø°Ø§Ø±ÛŒâ€ŒÙ‡Ø§ ]
                                                                            ,

                                                                            Û³. Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ:
                                                                            [Ø±ÛŒØ³Ú© ØªØ§Ù…ÛŒÙ† Ù…Ø§Ù„ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ù†Ù‚Ø¯Ø´ÙˆÙ†Ø¯Ú¯ÛŒ Ø¨Ø§Ø²Ø§Ø± ,
                                                                            Ø±ÛŒØ³Ú© ØªØ³ÙˆÛŒÙ‡ Ø¨Ø§ Ù†Ù‡Ø§Ø¯Ù‡Ø§ÛŒ Ø­Ø§Ú©Ù…ÛŒØªÛŒ ]
                                                                            ,

                                                                            Û´. Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ø¹Ù…Ù„ÛŒØ§ØªÛŒ:
                                                                            [Ø±ÛŒØ³Ú© ÙØ±Ø¢ÛŒÙ†Ø¯Ù‡Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© ÙÙ†Ø§ÙˆØ±ÛŒ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ùˆ Ø§Ù…Ù†ÛŒØª Ø³Ø§ÛŒØ¨Ø±ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ù…Ù†Ø§Ø¨Ø¹ Ø§Ù†Ø³Ø§Ù†ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© ØªÙ‚Ù„Ø¨ ,
                                                                            Ø±ÛŒØ³Ú© Ø±ÙˆÛŒØ¯Ø§Ø¯Ù‡Ø§ÛŒ Ø®Ø§Ø±Ø¬ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ù…Ø¯Ù„ ,
                                                                            Ø±ÛŒØ³Ú© Ø¹Ø¯Ù… Ú©ÙØ§ÛŒØª Ù¾ÙˆØ´Ø´ Ø¨ÛŒÙ…Ù‡â€ŒØ§ÛŒ ]
                                                                            ,

                                                                            Ûµ. Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ùˆ ØªØ·Ø¨ÛŒÙ‚:
                                                                            [Ø±ÛŒØ³Ú© Ø¯Ø¹Ø§ÙˆÛŒ Ø­Ù‚ÙˆÙ‚ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ø¹Ø¯Ù… Ø±Ø¹Ø§ÛŒØª Ù…Ù‚Ø±Ø±Ø§Øª,
                                                                            Ø±ÛŒØ³Ú© Ù‚Ø±Ø§Ø±Ø¯Ø§Ø¯Ù‡Ø§ ,
                                                                            Ø±ÛŒØ³Ú© Ù…Ø§Ù„ÛŒØ§ØªÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ù¾ÙˆÙ„Ø´ÙˆÛŒÛŒ Ùˆ ØªØ§Ù…ÛŒÙ† Ù…Ø§Ù„ÛŒ ØªØ±ÙˆØ±ÛŒØ³Ù… ,
                                                                            Ø±ÛŒØ³Ú© Ø­Ø§Ú©Ù…ÛŒØª Ø´Ø±Ú©ØªÛŒ ]
                                                                            ,
                                                                            Û¶. Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ø§Ø³ØªØ±Ø§ØªÚ˜ÛŒÚ©:
                                                                            [Ø±ÛŒØ³Ú© Ø±Ù‚Ø§Ø¨Øª ,
                                                                            Ø±ÛŒØ³Ú© ØªØºÛŒÛŒØ±Ø§Øª ØªÚ©Ù†ÙˆÙ„ÙˆÚ˜ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© ØªØµÙ…ÛŒÙ…Ø§Øª Ù…Ø¯ÛŒØ±ÛŒØªÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ù¾Ø±ÙˆÚ˜Ù‡â€ŒÙ‡Ø§ Ùˆ Ø³Ø±Ù…Ø§ÛŒÙ‡â€ŒÚ¯Ø°Ø§Ø±ÛŒâ€ŒÙ‡Ø§ÛŒ Ú©Ù„Ø§Ù†,
                                                                            Ø±ÛŒØ³Ú© Ø§Ø¯ØºØ§Ù… Ùˆ ØªÙ…Ù„ÛŒÚ© ,
                                                                            Ø±ÛŒØ³Ú© ØªØºÛŒÛŒØ±Ø§Øª Ú©Ù„Ø§Ù† Ø§Ù‚ØªØµØ§Ø¯ÛŒ ]

                                                                            Û·. Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú© Ø´Ù‡Ø±Øª:
                                                                            [Ø±ÛŒØ³Ú© Ø±Ø¶Ø§ÛŒØª Ù…Ø´ØªØ±ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© ÙˆØ¬Ù‡Ù‡ Ø¹Ù…ÙˆÙ…ÛŒ ,
                                                                            Ø±ÛŒØ³Ú© Ø±ÙˆØ§Ø¨Ø· Ø¨Ø§ Ø°ÛŒÙ†ÙØ¹Ø§Ù† ],
                                                                            8.Ø³Ø§ÛŒØ±
                                                            """,

                                                        }
                                                    },
                                                    "required": ["Ø§Ø±Ø¬Ø§Ø¹", "Ø¹Ù†ÙˆØ§Ù†", "Ø´Ø±Ø­", "Ø±ÛŒØ³Ú©_Ø¨Ø±Ø¬Ø³ØªÙ‡_Ø´Ø¯Ù‡", "Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ_Ø±ÛŒØ³Ú©", "Ø²ÛŒØ±Ø´Ø§Ø®Ù‡_Ø±ÛŒØ³Ú©"]
                                                }
                                            }
                                        },
                                        "required": ["Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯"]
                                    },
                                    "Ú¯Ø²Ø§Ø±Ø´_Ø±Ø¹Ø§ÛŒØª_Ø§Ù„Ø²Ø§Ù…Ø§Øª_Ù‚Ø§Ù†ÙˆÙ†ÛŒ": {
                                        "type": "object",
                                        "properties": {
                                            "Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯": {"type": "boolean"},
                                            "ØªØ®Ù„ÙØ§Øª": {
                                                "type": "array",
                                                "items": {
                                                    "type": "object",
                                                    "properties": {
                                                        "Ø§Ø±Ø¬Ø§Ø¹": {
                                                            "type": "object",
                                                            "properties": {
                                                               "Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯": {
                                                                "type": "string",
                                                                "description": "Ø´Ù…Ø§Ø±Ù‡ Ø¨Ù†Ø¯ Ù…Ø±Ø¨ÙˆØ·Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ .Ø¨ÛŒÙ† Ø¨Ù†Ø¯ Ù‡Ø§ , Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡ Ù…Ø§Ù†Ù†Ø¯ Û²,Û¶"
                                                            },
                                                            "Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡": {
                                                                "type": "string",
                                                                "description": "Ø´Ù…Ø§Ø±Ù‡ ØµÙØ­Ù‡ Ù…Ø±Ø¨ÙˆØ·Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ.Ú†Ù†Ø§Ù†Ú†Ù‡ Ø§ÛŒÙ† Ù…ÙˆØ±Ø¯ Ø¯Ø± Ú†Ù†Ø¯ Ø¨Ù†Ø¯ Ø¨Ù‡ Ø§Ù† Ø§Ø´Ø§Ø±Ù‡ Ø´Ø¯Ù‡ ØµÙØ­Ø§Øª Ù…Ù†Ø·Ø¨Ù‚ Ø¨Ø§ Ø¨Ù†Ø¯ Ø±Ø§ Ø¨Ù‡ ØªØ±ØªÛŒØ¨ Ø¨Ù†Ø¯ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù† Ø¨ÛŒÙ† ØµÙØ­Ø§Øª , Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡ Ù…Ø§Ù†Ù†Ø¯ Û±,Ûµ"
                                                            }
                                                            },
                                                            "required": ["Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯", "Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡"]
                                                        },
                                                        "Ø¹Ù†ÙˆØ§Ù†_ØªØ®Ù„Ù": {"type": "string"},
                                                        "Ø´Ø±Ø­": {"type": "string"},
                                                        "Ù…Ø¨Ø§Ù†ÛŒ_Ù‚Ø§Ù†ÙˆÙ†ÛŒ_Ùˆ_Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§": {
                                                            "type": "array",
                                                            "items": {
                                                                "type": "string",
                                                                "enum": [
                                                                    "Ù‚Ø§Ù†ÙˆÙ† Ù¾ÙˆÙ„ÛŒ Ùˆ Ø¨Ø§Ù†Ú©ÛŒ Ú©Ø´ÙˆØ±",
                                                                    "Ù‚Ø§Ù†ÙˆÙ† Ø¹Ù…Ù„ÛŒØ§Øª Ø¨Ø§Ù†Ú©ÛŒ Ø¨Ø¯ÙˆÙ† Ø±Ø¨Ø§â€ŒÙ‹",
                                                                    "Ø¢ÛŒÛŒÙ† Ù†Ø§Ù…Ù‡ Ù‡Ø§ Ùˆ Ø¯Ø³ØªÙˆØ±Ø§Ù„Ø¹Ù…Ù„â€ŒÙ‡Ø§ÛŒ Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒ (Ù…Ù‡Ù…ØªØ±ÛŒÙ† Ø¨Ø®Ø´)",
                                                                    "Ø§Ø³Ø§Ø³Ù†Ø§Ù…Ù‡ Ø¨Ø§Ù†Ú©",
                                                                    "Ù‚Ø§Ù†ÙˆÙ† ØªØ¬Ø§Ø±Øª (Ø¯Ø± Ù…ÙˆØ§Ø±Ø¯ Ù…Ø±ØªØ¨Ø·)",
                                                                    "Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§ÛŒ Ø­Ø³Ø§Ø¨Ø¯Ø§Ø±ÛŒ",
                                                                    "Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§ÛŒ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"
                                                                ]
                                                            }
                                                        },
                                                        "Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ": {
                                                            "type": "string",
                                                            "description": "Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ ØªØ®Ù„Ù Ø¨Ø± Ø§Ø³Ø§Ø³ Ù…Ù†Ø´Ø£ Ù‚Ø§Ù†ÙˆÙ†.Ø§Ú¯Ø± ÛŒÚ© ØªØ®Ù„Ù Ø¨Ù‡ Ú†Ù†Ø¯ Ù‚Ø§Ù†ÙˆÙ† Ù…Ø±Ø¨ÙˆØ· Ø§Ø³ØªØŒ Ø§ÙˆÙ„ÙˆÛŒØª Ø¨Ø§ Ø¯Ø³ØªÙ‡â€ŒØ§ÛŒ Ø§Ø³Øª Ú©Ù‡ Ø®Ø§Øµâ€ŒØªØ± Ùˆ Ù…Ù‡Ù…â€ŒØªØ± Ø§Ø³Øª. ",
                                                            "enum": [
                                                            "Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù†Ù‡Ø§Ø¯ Ù†Ø§Ø¸Ø± Ø¨Ø§Ù†Ú©ÛŒ (Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒ)",
                                                            "Ø§Ù„Ø²Ø§Ù…Ø§Øª Ø¨Ø§Ø²Ø§Ø± Ø³Ø±Ù…Ø§ÛŒÙ‡ (Ø³Ø§Ø²Ù…Ø§Ù† Ø¨ÙˆØ±Ø³)",
                                                            "Ù‚ÙˆØ§Ù†ÛŒÙ† Ø­Ø§Ú©Ù…ÛŒØªÛŒ Ùˆ Ø´Ø±Ú©ØªÛŒ",
                                                            "Ù‚ÙˆØ§Ù†ÛŒÙ† Ù…Ø§Ù„ÛŒØ§ØªÛŒØŒ Ø¨ÛŒÙ…Ù‡ Ùˆ Ø¨ÙˆØ¯Ø¬Ù‡",
                                                            "Ù‚ÙˆØ§Ù†ÛŒÙ† Ù…Ø¨Ø§Ø±Ø²Ù‡ Ø¨Ø§ Ø¬Ø±Ø§Ø¦Ù… Ù…Ø§Ù„ÛŒ",
                                                            "Ø³Ø§ÛŒØ± Ù‚ÙˆØ§Ù†ÛŒÙ† Ùˆ Ù…Ù‚Ø±Ø±Ø§Øª"
                                                            ]
                                                        },
                                                        "Ø²ÛŒØ±Ø´Ø§Ø®Ù‡": {
                                                            "type": "string",
                                                            "description": """ Ø²ÛŒØ±Ø´Ø§Ø®Ù‡ Ø¯Ù‚ÛŒÙ‚ ØªØ®Ù„Ù Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡ Ù…Ø±ØªØ¨Ø· Ø¨Ø§ Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡ .(Ø¯Ù‚ÛŒÙ‚Ø§ Ø§Ø² Ø§ÛŒÙ† Ù…ØµØ§Ø¯ÛŒÙ‚ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø´ÙˆØ¯. Ù…ØµØ§Ø¯Ù‚ Ù…Ù†Ø·Ø¨Ù‚ Ø¨Ø§ Ø§Ù„Ú¯ÙˆÛŒ Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ:[Ø²ÛŒØ±Ø´Ø§Ø®Ù‡ Ù‡Ø§] Ù…Ø§Ù†Ù†Ø¯ 
                                                                :Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù†Ù‡Ø§Ø¯ Ù†Ø§Ø¸Ø± Ø¨Ø§Ù†Ú©ÛŒ (Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒ)
                                                                [ Ú©ÙØ§ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡ Ùˆ Ù…Ø¯ÛŒØ±ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡,
                                                                ØªØ³Ù‡ÛŒÙ„Ø§Øª Ùˆ ØªØ¹Ù‡Ø¯Ø§Øª Ú©Ù„Ø§Ù†,
                                                                Ù…Ø¹Ø§Ù…Ù„Ø§Øª Ø¨Ø§ Ø§Ø´Ø®Ø§Øµ ÙˆØ§Ø¨Ø³ØªÙ‡,
                                                                Ø·Ø¨Ù‚Ù‡â€ŒØ¨Ù†Ø¯ÛŒ Ø¯Ø§Ø±Ø§ÛŒÛŒâ€ŒÙ‡Ø§ Ùˆ Ø°Ø®ÛŒØ±Ù‡â€ŒÚ¯ÛŒØ±ÛŒ,
                                                                Ù…Ø¯ÛŒØ±ÛŒØª Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ Ùˆ Ø³Ù¾Ø±Ø¯Ù‡ Ù‚Ø§Ù†ÙˆÙ†ÛŒ,
                                                                Ù†Ø±Ø® Ø³ÙˆØ¯ Ùˆ Ú©Ø§Ø±Ù…Ø²Ø¯ Ø®Ø¯Ù…Ø§Øª,
                                                                Ø§Ù„Ø²Ø§Ù…Ø§Øª Ø§Ø±Ø²ÛŒ,
                                                                ÙˆØ§Ú¯Ø°Ø§Ø±ÛŒ Ø§Ù…ÙˆØ§Ù„ Ùˆ Ø¯Ø§Ø±Ø§ÛŒÛŒâ€ŒÙ‡Ø§ÛŒ Ù…Ø§Ø²Ø§Ø¯,
                                                                Ø§Ù„Ø²Ø§Ù…Ø§Øª ØµÙ†Ø¯ÙˆÙ‚ Ø¶Ù…Ø§Ù†Øª Ø³Ù¾Ø±Ø¯Ù‡â€ŒÙ‡Ø§Ù…],

                                                                Ø§Ù„Ø²Ø§Ù…Ø§Øª Ø¨Ø§Ø²Ø§Ø± Ø³Ø±Ù…Ø§ÛŒÙ‡ (Ø³Ø§Ø²Ù…Ø§Ù† Ø¨ÙˆØ±Ø³):
                                                                [Ø§ÙØ´Ø§ÛŒ Ø§Ø·Ù„Ø§Ø¹Ø§Øª,
                                                                Ø­Ø§Ú©Ù…ÛŒØª Ø´Ø±Ú©ØªÛŒ (Ø¯Ø³ØªÙˆØ±Ø§Ù„Ø¹Ù…Ù„ Ø¨ÙˆØ±Ø³),
                                                                ØªÚ©Ø§Ù„ÛŒÙ Ù…Ø¬Ø§Ù…Ø¹ Ø¹Ù…ÙˆÙ…ÛŒ (Ù…Ø±Ø¨ÙˆØ· Ø¨Ù‡ Ø¨ÙˆØ±Ø³),
                                                                Ù…Ø¹Ø§Ù…Ù„Ø§Øª Ø¨Ø§ Ø§Ø´Ø®Ø§Øµ ÙˆØ§Ø¨Ø³ØªÙ‡ (Ø¶ÙˆØ§Ø¨Ø· Ø¨ÙˆØ±Ø³)],

                                                                Ù‚ÙˆØ§Ù†ÛŒÙ† Ø­Ø§Ú©Ù…ÛŒØªÛŒ Ùˆ Ø´Ø±Ú©ØªÛŒ:
                                                                [Ù†Ù‚Ø¶ Ù…ÙØ§Ø¯ Ø§Ø³Ø§Ø³Ù†Ø§Ù…Ù‡,
                                                                Ù†Ù‚Ø¶ Ù‚Ø§Ù†ÙˆÙ† ØªØ¬Ø§Ø±Øª,
                                                                ØªÚ©Ø§Ù„ÛŒÙ Ø«Ø¨Øª Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§,]
                                                                    
                                                                Ù‚ÙˆØ§Ù†ÛŒÙ† Ù…Ø§Ù„ÛŒØ§ØªÛŒØŒ Ø¨ÛŒÙ…Ù‡ Ùˆ Ø¨ÙˆØ¯Ø¬Ù‡:
                                                                [Ù…Ø§Ù„ÛŒØ§Øª Ø¹Ù…Ù„Ú©Ø±Ø¯ Ùˆ ØªÚ©Ù„ÛŒÙÛŒ,
                                                                Ù…Ø§Ù„ÛŒØ§Øª Ø¨Ø± Ø§Ø±Ø²Ø´ Ø§ÙØ²ÙˆØ¯Ù‡ (VAT),
                                                                Ø¨ÛŒÙ…Ù‡ ØªØ§Ù…ÛŒÙ† Ø§Ø¬ØªÙ…Ø§Ø¹ÛŒ,
                                                                Ù‚ÙˆØ§Ù†ÛŒÙ† Ø¨ÙˆØ¯Ø¬Ù‡ Ø³Ù†ÙˆØ§ØªÛŒ Ùˆ ØªÙˆØ³Ø¹Ù‡],
                                                                Ù‚ÙˆØ§Ù†ÛŒÙ† Ù…Ø¨Ø§Ø±Ø²Ù‡ Ø¨Ø§ Ø¬Ø±Ø§Ø¦Ù… Ù…Ø§Ù„ÛŒ:
                                                                [Ù…Ø¨Ø§Ø±Ø²Ù‡ Ø¨Ø§ Ù¾ÙˆÙ„Ø´ÙˆÛŒÛŒ (AML),
                                                                Ù…Ø¨Ø§Ø±Ø²Ù‡ Ø¨Ø§ ØªØ§Ù…ÛŒÙ† Ù…Ø§Ù„ÛŒ ØªØ±ÙˆØ±ÛŒØ³Ù… (CFT)],
                                                                Ø³Ø§ÛŒØ± Ù‚ÙˆØ§Ù†ÛŒÙ† Ùˆ Ù…Ù‚Ø±Ø±Ø§Øª:
                                                                [Ù‚Ø§Ù†ÙˆÙ† Ú©Ø§Ø±,
                                                                Ù‚ÙˆØ§Ù†ÛŒÙ† Ø´Ù‡Ø±Ø¯Ø§Ø±ÛŒ Ùˆ Ù…Ø­ÛŒØ· Ø²ÛŒØ³Øª,
                                                                Ø­Ù‚ÙˆÙ‚ Ù…Ø§Ù„Ú©ÛŒØª ÙÚ©Ø±ÛŒ,
                                                                Ø³Ø§ÛŒØ±
                                                                ]
                                                            """,
                                                            
                                                        },

                                                        },
                                                        "required": [
                                                        "Ø§Ø±Ø¬Ø§Ø¹",
                                                        "Ø¹Ù†ÙˆØ§Ù†_ØªØ®Ù„Ù",
                                                        "Ø´Ø±Ø­",
                                                        "Ù…Ø¨Ø§Ù†ÛŒ_Ù‚Ø§Ù†ÙˆÙ†ÛŒ_Ùˆ_Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§",
                                                        "Ø¯Ø³ØªÙ‡_Ø§ØµÙ„ÛŒ",
                                                        "Ø²ÛŒØ±Ø´Ø§Ø®Ù‡"
                                                        ]

                                                }
                                            }
                                        },
                                        "required": ["Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯"]
                                    }
                                }
                            },
                            "Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ": {
                                "type": "array",
                                "description":" Ø¨Ø§ÛŒØ¯ ØªÙ…Ø§Ù… Ù…ÙˆØ¶ÙˆØ¹ Ù‡Ø§ÛŒ Ø²ÛŒØ± Ø±Ø§ Ú†Ú© Ú©Ù†ÛŒØ¯ Ùˆ Ø¯Ø± Ø®Ø±ÙˆØ¬ÛŒ Ø¨ÛŒØ§ÙˆØ±ÛŒØ¯ Ù‡Ù… Ù…ÙˆØ§Ø±Ø¯ÛŒ Ú©Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø¢Ù…Ø¯Ù‡ Ù‡Ù… Ù…ÙˆØ§Ø±Ø¯ÛŒ Ú©Ù‡ Ù†ÛŒØ§Ù…Ø¯Ù‡",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "Ù…ÙˆØ¶ÙˆØ¹": {
                                            "type": "string",
                                            "enum": [
                     # ğŸ¯ Ø§ÙˆÙ„ÙˆÛŒØª Ø®ÛŒÙ„ÛŒ Ø¨Ø§Ù„Ø§
                                                "Ú©ÙØ§ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡",
                                                "Ù†Ø³Ø¨Øªâ€ŒÙ‡Ø§ Ø¯Ø± Ú†Ø§Ø±Ú†ÙˆØ¨ Ø¨Ø§Ø²Ù„",
                                                "Ø±ÛŒØ³Ú© Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ",
                                                "Ù…Ø¯ÛŒØ±ÛŒØª Ø¯Ø§Ø±Ø§ÛŒÛŒ Ùˆ Ø¨Ø¯Ù‡ÛŒ (ALM)",
                                                "Ø±ÛŒØ³Ú© Ù†Ø±Ø® Ø¨Ù‡Ø±Ù‡",
                                                "ØªÙ…Ø±Ú©Ø² Ø±ÛŒØ³Ú© Ø§Ø¹ØªØ¨Ø§Ø±ÛŒ",
                                                "Ø°Ø®ÛŒØ±Ù‡â€ŒÚ¯ÛŒØ±ÛŒ (Ú©Ù„ÛŒ)",

                                                # ğŸ”¥ Ø§ÙˆÙ„ÙˆÛŒØª Ø¨Ø§Ù„Ø§
                                                "ØµÙˆØ±Øª Ø¬Ø±ÛŒØ§Ù† ÙˆØ¬ÙˆÙ‡ Ù†Ù‚Ø¯",
                                                "Ú©Ù†ØªØ±Ù„â€ŒÙ‡Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ Ùˆ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ Ø¯Ø§Ø®Ù„ÛŒ",
                                                "Ø­Ø§Ú©Ù…ÛŒØª Ø´Ø±Ú©ØªÛŒ",
                                                "Ø§ÙˆØ±Ø§Ù‚ Ø¨Ù‡Ø§Ø¯Ø§Ø± Ùˆ Ø³Ø±Ù…Ø§ÛŒÙ‡â€ŒÚ¯Ø°Ø§Ø±ÛŒâ€ŒÙ‡Ø§",
                                                "ØªØ³Ø¹ÛŒØ± Ø§Ø±Ø² Ùˆ Ø¹Ù…Ù„ÛŒØ§Øª Ø®Ø§Ø±Ø¬ÛŒ",
                                                "ØªØ¹Ù‡Ø¯Ø§Øª Ø§Ø±Ø²ÛŒ Ùˆ Ø§Ø®ØªÙ„Ø§Ù Ø¨Ø§ Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒ",
                                                "Ø°Ø®ÛŒØ±Ù‡ Ù…Ø·Ø§Ù„Ø¨Ø§Øª Ù…Ø´Ú©ÙˆÚ©â€ŒØ§Ù„ÙˆØµÙˆÙ„",
                                                "Ù…Ø·Ø§Ù„Ø¨Ø§Øª Ù…Ø´Ú©ÙˆÚ©â€ŒØ§Ù„ÙˆØµÙˆÙ„",
                                                "ØªØ³Ù‡ÛŒÙ„Ø§Øª Ùˆ Ø§Ø¹ØªØ¨Ø§Ø±Ø§Øª",
                                                "Ø³Ø±Ù…Ø§ÛŒÙ‡â€ŒÚ¯Ø°Ø§Ø±ÛŒ Ø¯Ø± Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ ÙˆØ§Ø¨Ø³ØªÙ‡",
                                                "Ú©Ø§Ù‡Ø´ Ø§Ø±Ø²Ø´ Ø¯Ø§Ø±Ø§ÛŒÛŒâ€ŒÙ‡Ø§",
                                                "Ø§ÙØ´Ø§ÛŒ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¹Ù…Ù„ÛŒØ§ØªÛŒ",
                                                "Ù†Ø³Ø¨Øªâ€ŒÙ‡Ø§ÛŒ Ø¨Ø¯Ù‡ÛŒ Ùˆ Ù†Ù‚Ø¯ÛŒÙ†Ú¯ÛŒ",
                                                "Ù†Ø³Ø¨Øª Ú©ÙØ§ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡",
                                                "Ù…Ø¹Ø§Ù…Ù„Ø§Øª Ø¨Ø§ Ø§Ø´Ø®Ø§Øµ ÙˆØ§Ø¨Ø³ØªÙ‡",
                                                "ØªØ¯Ø§ÙˆÙ… ÙØ¹Ø§Ù„ÛŒØª",
                                                "Ø§Ù†Ø·Ø¨Ø§Ù‚ Ø¨Ø§ Ù…Ù‚Ø±Ø±Ø§Øª Ø¶Ø¯Ù¾ÙˆÙ„Ø´ÙˆÛŒÛŒ (AML/CFT)",

                                                # # âš™ï¸ Ø§ÙˆÙ„ÙˆÛŒØª Ù…ØªÙˆØ³Ø·
                                                "Ø°Ø®ÛŒØ±Ù‡ Ù…Ø²Ø§ÛŒØ§ÛŒ Ù¾Ø§ÛŒØ§Ù† Ø®Ø¯Ù…Øª Ú©Ø§Ø±Ú©Ù†Ø§Ù†",
                                                "Ø±ÛŒØ³Ú© Ø´Ù‡Ø±Øª",
                                                "Ø°Ø®ÛŒØ±Ù‡ Ù…Ø§Ù„ÛŒØ§Øª Ø¨Ø± Ø¯Ø±Ø¢Ù…Ø¯",
                                                "Ø­Ù‚ÙˆÙ‚ ØµØ§Ø­Ø¨Ø§Ù† Ø³Ù‡Ø§Ù…",
                                                "Ø³ÛŒØ³ØªÙ…â€ŒÙ‡Ø§ÛŒ Ø§Ø·Ù„Ø§Ø¹Ø§ØªÛŒ Ùˆ ÙÙ†Ø§ÙˆØ±ÛŒ",
                                                "Ø§Ù†Ø·Ø¨Ø§Ù‚ Ø¨Ø§ Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§ÛŒ Ø¨ÛŒÙ†â€ŒØ§Ù„Ù…Ù„Ù„ÛŒ",
                                                "Ù¾ÙˆØ´Ø´ Ø¨ÛŒÙ…Ù‡â€ŒØ§ÛŒ Ø¯Ø§Ø±Ø§ÛŒÛŒâ€ŒÙ‡Ø§",
                                                "Ø¯Ø¹Ø§ÙˆÛŒ Ùˆ Ø¬Ø±Ø§Ø¦Ù… Ø­Ù‚ÙˆÙ‚ÛŒ",
                                                "Ú©ÛŒÙÛŒØª Ø§ÙØ´Ø§ÛŒ Ø§Ø·Ù„Ø§Ø¹Ø§Øª",
                                                "Ø±ÙˆÛŒØ¯Ø§Ø¯Ù‡Ø§ÛŒ Ø¨Ø¹Ø¯ Ø§Ø² ØªØ§Ø±ÛŒØ® ØªØ±Ø§Ø²Ù†Ø§Ù…Ù‡",
                                                "ØªØºÛŒÛŒØ± Ø±ÙˆÛŒÙ‡â€ŒÙ‡Ø§ÛŒ Ø­Ø³Ø§Ø¨Ø¯Ø§Ø±ÛŒ",
                                                "Ø¨Ø¯Ù‡ÛŒâ€ŒÙ‡Ø§ÛŒ Ø§Ø­ØªÙ…Ø§Ù„ÛŒ",
                                                "Ù†Ø³Ø¨Øªâ€ŒÙ‡Ø§ÛŒ Ø³ÙˆØ¯Ø¢ÙˆØ±ÛŒ",
                                                "Ù…Ø§Ù„ÛŒØ§Øª Ùˆ Ø¬Ø±Ø§Ø¦Ù… Ù…Ø§Ù„ÛŒØ§ØªÛŒ",
                                                "Ø³ÙˆØ¯ Ø³Ù‡Ø§Ù… Ø¯ÙˆÙ„Øª",
                                                "Ø¹Ø¯Ù… Ø¯Ø±ÛŒØ§ÙØª ØªØ£ÛŒÛŒØ¯ÛŒÙ‡â€ŒÙ‡Ø§ÛŒ Ø­Ø³Ø§Ø¨Ø¯Ø§Ø±ÛŒ",
                                                "Ø°Ø®ÛŒØ±Ù‡ Ø¯Ø¹Ø§ÙˆÛŒ Ø­Ù‚ÙˆÙ‚ÛŒ",
                                                "ØªÙ‡Ø§ØªØ± (Barter)",
                                                "ØµÙØ­Ù‡ Ø§Ù…Ø¶Ø§ Ù‡Ø§ÛŒ Ø³Ø§Ø²Ù…Ø§Ù† Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"
                                            ]
                                        },
                                        "Ø¯Ø±_Ú¯Ø²Ø§Ø±Ø´_Ø¢Ù…Ø¯Ù‡": {"type": "boolean"},
                                        "ÙˆØ¶Ø¹ÛŒØª": {
                                            "type": "string",
                                            "enum": [
                                                "Ù…ØµØ¯Ø§Ù‚ Ù†Ø¯Ø§Ø±Ø¯", "Ø¨Ø±Ø±Ø³ÛŒ Ø´Ø¯Ù‡ - Ø±ÛŒØ³Ú© Ø®Ø§ØµÛŒ Ú¯Ø²Ø§Ø±Ø´ Ù†Ø´Ø¯Ù‡",
                                                "Ù…Ø³Ø¦Ù„Ù‡ Ú©Ù„ÛŒØ¯ÛŒ Ù…Ù†Ø¬Ø± Ø¨Ù‡ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± Ù…Ø´Ø±ÙˆØ·", "Ø±ÛŒØ³Ú© Ø¨Ø­Ø±Ø§Ù†ÛŒ"
                                            ]
                                        },
                                        "Ø¬Ø²Ø¦ÛŒØ§Øª": {"type": "string"},
                                        "Ø§Ø±Ø¬Ø§Ø¹": {
                                            "type": "object",
                                            "properties": {
                                                "Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯": {
                                                "type": "string",
                                                "description": "Ø´Ù…Ø§Ø±Ù‡ Ø¨Ù†Ø¯ Ù…Ø±Ø¨ÙˆØ·Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ .Ø¨ÛŒÙ† Ø¨Ù†Ø¯ Ù‡Ø§ , Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡ Ù…Ø§Ù†Ù†Ø¯ Û²,Û¶"
                                            },
                                            "Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡": {
                                                "type": "string",
                                                "description": "Ø´Ù…Ø§Ø±Ù‡ ØµÙØ­Ù‡ Ù…Ø±Ø¨ÙˆØ·Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ù…Ø³ØªÙ‚Ù„ Ùˆ Ø¨Ø§Ø²Ø±Ø³ Ù‚Ø§Ù†ÙˆÙ†ÛŒ.Ú†Ù†Ø§Ù†Ú†Ù‡ Ø§ÛŒÙ† Ù…ÙˆØ±Ø¯ Ø¯Ø± Ú†Ù†Ø¯ Ø¨Ù†Ø¯ Ø¨Ù‡ Ø§Ù† Ø§Ø´Ø§Ø±Ù‡ Ø´Ø¯Ù‡ ØµÙØ­Ø§Øª Ù…Ù†Ø·Ø¨Ù‚ Ø¨Ø§ Ø¨Ù†Ø¯ Ø±Ø§ Ø¨Ù‡ ØªØ±ØªÛŒØ¨ Ø¨Ù†Ø¯ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù† Ø¨ÛŒÙ† ØµÙØ­Ø§Øª , Ù‚Ø±Ø§Ø± Ø¨Ø¯Ù‡ Ù…Ø§Ù†Ù†Ø¯ Û±,Ûµ"
                                            }
                                            }
                                        }
                                    },
                                    "required": ["Ù…ÙˆØ¶ÙˆØ¹", "Ø¯Ø±_Ú¯Ø²Ø§Ø±Ø´_Ø¢Ù…Ø¯Ù‡", "ÙˆØ¶Ø¹ÛŒØª", "Ø¬Ø²Ø¦ÛŒØ§Øª", "Ø§Ø±Ø¬Ø§Ø¹"]
                                }
                            }
                        }
                    }
                },
                "required": ["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
            }
        
        def extract_table_from_page(self, file_content: bytes, filename: str, max_retries: int = 5) -> Dict:

            prompt =  """Ù„Ø·ÙØ§Ù‹ Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ Ø±Ø§ ØªØ­Ù„ÛŒÙ„ Ú©Ù†ÛŒØ¯. Ù†Ú©ØªÙ‡ Ø¨Ø³ÛŒØ§Ø± Ù…Ù‡Ù… Ø¨Ø±Ø§ÛŒ Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ:
            - Ø¨Ø§ÛŒØ¯ ØªÙ…Ø§Ù…  Ù…ÙˆØ¶ÙˆØ¹ Ù‡Ø§ Ø±Ø§ Ú†Ú© Ú©Ù†ÛŒØ¯ Ùˆ Ø¯Ø± Ø®Ø±ÙˆØ¬ÛŒ Ø¨ÛŒØ§ÙˆØ±ÛŒØ¯
            - Ø¨Ø±Ø§ÛŒ Ù‡Ø± Ù…ÙˆØ¶ÙˆØ¹ØŒ ÙÛŒÙ„Ø¯ "Ø¯Ø±_Ú¯Ø²Ø§Ø±Ø´_Ø¢Ù…Ø¯Ù‡" Ø±Ø§ Ù…Ø´Ø®Øµ Ú©Ù†ÛŒØ¯ (true ÛŒØ§ false)
            - Ù‡Ù…Ù‡ Ù…ÙˆØ¶ÙˆØ¹ Ù‡Ø§ Ø¨Ø§ÛŒØ¯ Ø¯Ø± Ø¢Ø±Ø§ÛŒÙ‡ Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ Ø¨Ø§Ø´Ù†Ø¯"""
            
            for attempt in range(max_retries):
                try:
                    client, current_api_key = get_client_with_retry()
                    logger.info(f"Processing {filename} - Attempt {attempt + 1}")
                    response = client.models.generate_content(
                        model="gemini-2.5-flash",
                        contents=[types.Part.from_bytes(data=file_content, mime_type="application/pdf"), prompt],
                        config={
                            'system_instruction': "Ø´Ù…Ø§ ØªØ­Ù„ÛŒÙ„Ú¯Ø± Ù…Ø§Ù„ÛŒ Ù‡Ø³ØªÛŒØ¯.",
                            "response_mime_type": "application/json",
                            "response_schema": self.response_schema,
                            "temperature": 0.5
                        }
                    )
                    if not response or not response.text:
                        raise ValueError("API response was empty")
                    data = json.loads(response.text)
                    api_key_manager.mark_success(current_api_key)
                    logger.info(f"Successfully processed {filename}")
                    return data
                except Exception as e:
                    logger.error(f"Error: {str(e)}")
                    api_key_manager.mark_failure(current_api_key)
                    if attempt < max_retries - 1:
                        time.sleep(min(15, (2 ** attempt)))
                    else:
                        raise
            raise Exception(f"Failed after {max_retries} attempts")

    # ========================================================================
    # âœ… ØªØ§Ø¨Ø¹ Ø¬Ø¯ÛŒØ¯: Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡Ù…Ø²Ù…Ø§Ù† Ø¨Ø§ ThreadPoolExecutor
    # ========================================================================

    def process_single_file(analyzer, file_data, index, total, attempt=1, max_attempts=3):
        """
        Ù¾Ø±Ø¯Ø§Ø²Ø´ ÛŒÚ© ÙØ§ÛŒÙ„ Ø¨Ø§ Ù‚Ø§Ø¨Ù„ÛŒØª retry Ø®ÙˆØ¯Ú©Ø§Ø±
        
        Args:
            analyzer: Ù†Ù…ÙˆÙ†Ù‡ FinancialAnalyzer
            file_data: Ø¯Ø§Ø¯Ù‡ ÙØ§ÛŒÙ„ (dict ÛŒØ§ file object)
            index: Ø´Ù…Ø§Ø±Ù‡ ÙØ§ÛŒÙ„
            total: ØªØ¹Ø¯Ø§Ø¯ Ú©Ù„ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§
            attempt: ØªÙ„Ø§Ø´ ÙØ¹Ù„ÛŒ (1ØŒ 2ØŒ 3)
            max_attempts: Ø­Ø¯Ø§Ú©Ø«Ø± ØªØ¹Ø¯Ø§Ø¯ ØªÙ„Ø§Ø´
        
        Returns:
            tuple: (index, filename, result, error, needs_retry)
        """
        filename = file_data['name'] if isinstance(file_data, dict) else file_data.name
        file_content = file_data['content'] if isinstance(file_data, dict) else file_data.getvalue()
        
        try:
            logger.info(f"ğŸ”„ Processing {filename} - Attempt {attempt}/{max_attempts}")
            result = analyzer.extract_table_from_page(file_content, filename)
            return (index, filename, result, None, False)  # âœ… Ø§Ø¶Ø§ÙÙ‡ Ø´Ø¯Ù† False
        except Exception as e:
            error_msg = str(e)
            logger.error(f"âŒ Failed to process {filename} (Attempt {attempt}): {error_msg}")
            
            # Ø¨Ø±Ø±Ø³ÛŒ Ø§ÛŒÙ†Ú©Ù‡ Ø¢ÛŒØ§ Ù†ÛŒØ§Ø² Ø¨Ù‡ retry Ø¯Ø§Ø±Ø¯ ÛŒØ§ Ù†Ù‡
            needs_retry = attempt < max_attempts and is_retryable_error(error_msg)
            
            return (index, filename, None, error_msg, needs_retry)  # âœ… Ø§Ø¶Ø§ÙÙ‡ Ø´Ø¯Ù† needs_retry
        

    def is_retryable_error(error_msg: str) -> bool:
        """
        ØªØ´Ø®ÛŒØµ Ø§ÛŒÙ†Ú©Ù‡ Ø®Ø·Ø§ Ù‚Ø§Ø¨Ù„ retry Ø§Ø³Øª ÛŒØ§ Ù†Ù‡
        
        Ø®Ø·Ø§Ù‡Ø§ÛŒ Ù‚Ø§Ø¨Ù„ retry:
        - Timeout
        - Rate limit (429)
        - Server error (500, 503)
        - Network errors
        - API overloaded
        
        Ø®Ø·Ø§Ù‡Ø§ÛŒ ØºÛŒØ±Ù‚Ø§Ø¨Ù„ retry:
        - Invalid file format
        - Authentication error (403)
        - File too large
        """
        retryable_patterns = [
            'timeout',
            'timed out',
            'rate limit',
            '429',
            '500',
            '503',
            'server error',
            'network',
            'connection',
            'overloaded',
            'temporarily unavailable',
            'try again later',
            'unavailable'
        ]
        
        error_lower = error_msg.lower()
        return any(pattern in error_lower for pattern in retryable_patterns)
    
 
    # ========================================================================
    # âœ… ØªØ§Ø¨Ø¹ Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡: create_processing_section
    # ========================================================================

    def create_processing_section(uploaded_files):
        """
        Ø¨Ø®Ø´ Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø¨Ø§ Ù…Ø¯ÛŒØ±ÛŒØª Ø­Ø§Ù„Øª Ú©Ø§Ù…Ù„ (Ø¢Ù…Ø§Ø¯Ù‡ØŒ Ø¯Ø± Ø­Ø§Ù„ Ù¾Ø±Ø¯Ø§Ø²Ø´ØŒ Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯Ù‡)
        Ùˆ Ù‚Ø§Ø¨Ù„ÛŒØª ØªØ­Ù„ÛŒÙ„ Ù…Ø¬Ø¯Ø¯ - Ø¨Ø§ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² Ù†Ø§Ù… ÛŒÚ©Ø³Ø§Ù† session_state
        """
        # Ù…Ù‚Ø¯Ø§Ø±Ø¯Ù‡ÛŒ Ø§ÙˆÙ„ÛŒÙ‡ Ø­Ø§Ù„Øªâ€ŒÙ‡Ø§ Ø¯Ø± ØµÙˆØ±Øª Ø¹Ø¯Ù… ÙˆØ¬ÙˆØ¯
        if 'processing_active' not in st.session_state:
            st.session_state.processing_active = False
        if 'results' not in st.session_state:  # âœ… Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² 'results' Ø¨Ù‡ Ø¬Ø§ÛŒ 'processing_results'
            st.session_state.results = None

        if not uploaded_files:
            st.session_state.results = None  # Ù¾Ø§Ú© Ú©Ø±Ø¯Ù† Ù†ØªØ§ÛŒØ¬ Ø§Ú¯Ø± ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø­Ø°Ù Ø´ÙˆÙ†Ø¯
            return

        # =========================================================================
        # Ø¨Ø®Ø´ 1: Ù†Ù…Ø§ÛŒØ´ Ú©Ø§Ø±Øªâ€ŒÙ‡Ø§ÛŒ Ø¢Ù…Ø§Ø±ÛŒ (Ø§ÛŒÙ† Ø¨Ø®Ø´ Ù‡Ù…ÛŒØ´Ù‡ Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ù…ÛŒâ€ŒØ´ÙˆØ¯)
        # =========================================================================
        with st.container():
            st.subheader("ğŸš€ ÙˆØ¶Ø¹ÛŒØª Ù¾Ø±Ø¯Ø§Ø²Ø´")
            st.divider()

            col1, col2, col3 = st.columns(3)
            total_size_mb = sum(
                len(f['content']) if isinstance(f, dict) else f.size
                for f in uploaded_files
            ) / (1024 * 1024)

            with col1:
                st.markdown(f'''
                    <div class="metric-modern">
                        <p>ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§</p>
                        <div class="metric-value-box">
                            <div class="metric-value">{len(uploaded_files)}</div>
                        </div>
                    </div>
                ''', unsafe_allow_html=True)
                
            with col2:
                st.markdown(f'''
                    <div class="metric-modern">
                        <p>Ø­Ø¬Ù… Ú©Ù„</p>
                        <div class="metric-value-box">
                            <div class="metric-value">{total_size_mb:.1f} MB</div>
                        </div>
                    </div>
                ''', unsafe_allow_html=True)

            with col3:
                # ØªØ¹ÛŒÛŒÙ† ÙˆØ¶Ø¹ÛŒØª Ø¨Ø± Ø§Ø³Ø§Ø³ Ø­Ø§Ù„Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø®ØªÙ„Ù
                if st.session_state.results:
                    status_text = "ØªÚ©Ù…ÛŒÙ„ Ø´Ø¯ âœ…"
                    status_class = "metric-status-done"
                elif st.session_state.processing_active:
                    status_text = "Ø¯Ø± Ø­Ø§Ù„ Ù¾Ø±Ø¯Ø§Ø²Ø´ ğŸ”„"
                    status_class = "metric-status-processing"
                else:
                    status_text = "Ø¢Ù…Ø§Ø¯Ù‡ ğŸŸ¢"
                    status_class = "metric-status-ready"
                
                st.markdown(f'''
                    <div class="metric-modern {status_class}">
                        <p>ÙˆØ¶Ø¹ÛŒØª</p>
                        <div class="metric-value-box">
                            <div class="metric-value">{status_text}</div>
                        </div>
                    </div>
                ''', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # =========================================================================
            # Ø¨Ø®Ø´ 2: Ø¯Ú©Ù…Ù‡â€ŒÙ‡Ø§ Ùˆ Ù…Ù†Ø·Ù‚ Ù¾Ø±Ø¯Ø§Ø²Ø´ (Ø¨Ø± Ø§Ø³Ø§Ø³ Ø­Ø§Ù„Øª ÙØ¹Ù„ÛŒ)
            # =========================================================================

            # Ø­Ø§Ù„Øª 1: Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯Ù‡ Ùˆ Ù†ØªØ§ÛŒØ¬ Ù…ÙˆØ¬ÙˆØ¯ Ø§Ø³Øª
            if st.session_state.results:
                # Ù…Ø­Ø§Ø³Ø¨Ù‡ ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ù…ÙˆÙÙ‚
                successful_count = len([r for r in st.session_state.results if r and 'error' not in r[1]])
                failed_count = len(st.session_state.results) - successful_count
                
                # Ù†Ù…Ø§ÛŒØ´ Ù¾ÛŒØ§Ù… Ù…ÙˆÙÙ‚ÛŒØª Ø¨Ø§ Ø¬Ø²Ø¦ÛŒØ§Øª
                if failed_count == 0:
                    st.success(f"ğŸ‰ ØªØ­Ù„ÛŒÙ„ Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª ØªÚ©Ù…ÛŒÙ„ Ø´Ø¯! {successful_count} ÙØ§ÛŒÙ„ ØªØ­Ù„ÛŒÙ„ Ø´Ø¯. Ù†ØªØ§ÛŒØ¬ Ø¯Ø± ØªØ¨â€ŒÙ‡Ø§ÛŒ Ø¯ÛŒÚ¯Ø± Ù‚Ø§Ø¨Ù„ Ù…Ø´Ø§Ù‡Ø¯Ù‡ Ø§Ø³Øª.")
                else:
                    st.warning(f"âš ï¸ ØªØ­Ù„ÛŒÙ„ ØªÚ©Ù…ÛŒÙ„ Ø´Ø¯: {successful_count} ÙØ§ÛŒÙ„ Ù…ÙˆÙÙ‚ØŒ {failed_count} ÙØ§ÛŒÙ„ Ù†Ø§Ù…ÙˆÙÙ‚. Ù†ØªØ§ÛŒØ¬ Ø¯Ø± ØªØ¨â€ŒÙ‡Ø§ÛŒ Ø¯ÛŒÚ¯Ø± Ù‚Ø§Ø¨Ù„ Ù…Ø´Ø§Ù‡Ø¯Ù‡ Ø§Ø³Øª.")
                
                # Ø¯Ú©Ù…Ù‡ ØªØ­Ù„ÛŒÙ„ Ù…Ø¬Ø¯Ø¯
                col_btn1, col_btn2 = st.columns([1, 3])
                with col_btn1:
                    if st.button("ğŸ”„ ØªØ­Ù„ÛŒÙ„ Ù…Ø¬Ø¯Ø¯", type="primary", use_container_width=True):
                        # Ù¾Ø§Ú© Ú©Ø±Ø¯Ù† Ù†ØªØ§ÛŒØ¬ Ù‚Ø¨Ù„ÛŒ Ùˆ Ø±ÛŒØ³Øª Ú©Ø±Ø¯Ù† Ø­Ø§Ù„Øª Ø¨Ø±Ø§ÛŒ Ø´Ø±ÙˆØ¹ Ù…Ø¬Ø¯Ø¯
                        st.session_state.results = None
                        st.session_state.processing_active = False
                        st.rerun()

            # Ø­Ø§Ù„Øª 2: Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ø­Ø§Ù„ Ø§Ù†Ø¬Ø§Ù… Ø§Ø³Øª
            elif st.session_state.processing_active:
                st.warning("âš ï¸ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ø­Ø§Ù„ Ø§Ø¬Ø±Ø§ Ø§Ø³Øª. Ù„Ø·ÙØ§Ù‹ Ù…Ù†ØªØ¸Ø± Ø¨Ù…Ø§Ù†ÛŒØ¯...")
                info_html = """
                    <div style="
                        background-color: #e6f3ff; 
                        border-radius: 10px; 
                        padding: 1rem 1.5rem; 
                        margin-bottom: 1rem;
                        direction: rtl; 
                        border-right: 6px solid #1c83e1;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    ">
                        <h4 style="margin-top: 0; margin-bottom: 0.75rem; font-weight: bold;">ğŸ’¡ Ù†Ú©ØªÙ‡ Ù…Ù‡Ù…:</h4>
                        <ul style="padding-right: 25px; margin-bottom: 0;">
                            <li style="margin-bottom: 8px;">Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¯Ø± Ù¾Ø³â€ŒØ²Ù…ÛŒÙ†Ù‡ Ø¯Ø± Ø­Ø§Ù„ Ø§Ù†Ø¬Ø§Ù… Ø§Ø³Øª.</li>
                            <li style="margin-bottom: 8px;">Ù¾Ø³ Ø§Ø² Ø§ØªÙ…Ø§Ù…ØŒ Ù†ØªØ§ÛŒØ¬ Ø¨Ù‡ ØµÙˆØ±Øª Ø®ÙˆØ¯Ú©Ø§Ø± Ø¯Ø± ØªØ¨â€ŒÙ‡Ø§ÛŒ Ø¯ÛŒÚ¯Ø± Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ù…ÛŒâ€ŒØ´ÙˆØ¯.</li>
                            <li style="margin-bottom: 8px;">Ø¯Ø± ØµÙˆØ±Øª Ø¹Ø¯Ù… Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„ Ù…ÛŒØªÙˆØ§Ù†ÛŒØ¯ Ø§Ø² Ø¯Ú©Ù…Ù‡ ØªØ­Ù„ÛŒÙ„ Ù…Ø¬Ø¯Ø¯ Ø§Ø³ØªÙØ§Ø¯Ù‡ Ú©Ù†ÛŒØ¯.</li>
                            <li>Ú†Ù†Ø§Ù†Ú†Ù‡ Ø¯Ø± Ù†ØªØ§ÛŒØ¬ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø®Ø·Ø§ÛŒ 429 Ù…Ø´Ø§Ù‡Ø¯Ù‡ Ú©Ø±Ø¯ÛŒØ¯ Ø¨Ø§ÛŒØ¯ Ú©Ù„ÛŒØ¯Ù‡Ø§ÛŒ Ø®ÙˆØ¯ Ø±Ø§ ØªØºÛŒÛŒØ± Ø¯Ù‡ÛŒØ¯.</li>
                        </ul>
                    </div>
                    """
                st.markdown(info_html, unsafe_allow_html=True)
                # Ø¯Ø± Ø§ÛŒÙ† Ø­Ø§Ù„ØªØŒ ØªØ§Ø¨Ø¹ Ø§ØµÙ„ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ±Ø§Ø®ÙˆØ§Ù†ÛŒ Ù…ÛŒâ€ŒØ´ÙˆØ¯
                try:
                    results = process_files_concurrent_smart(uploaded_files)
                    st.session_state.results = results
                    st.session_state.processing_active = False
                    st.rerun()  # Ø¨Ø§Ø²Ø®ÙˆØ§Ù†ÛŒ ØµÙØ­Ù‡ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ Ø­Ø§Ù„Øª "Ø§Ù†Ø¬Ø§Ù… Ø´Ø¯Ù‡"
                except Exception as e:
                    st.session_state.processing_active = False
                    st.session_state.results = None
                    st.error(f"âŒ Ø®Ø·Ø§ÛŒ ØºÛŒØ±Ù…Ù†ØªØ¸Ø±Ù‡ Ø¯Ø± Ø­ÛŒÙ† Ù¾Ø±Ø¯Ø§Ø²Ø´: {str(e)}")
                    logger.error(f"Critical processing error: {traceback.format_exc()}")
                    st.rerun()

            # Ø­Ø§Ù„Øª 3: Ø¢Ù…Ø§Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ø´Ø±ÙˆØ¹ Ù¾Ø±Ø¯Ø§Ø²Ø´
            else:
                if st.button("ğŸš€ Ø´Ø±ÙˆØ¹ ØªØ­Ù„ÛŒÙ„", type="primary", use_container_width=True):
                    st.session_state.processing_active = True
                    st.session_state.results = None  # Ø§Ø·Ù…ÛŒÙ†Ø§Ù† Ø§Ø² Ù¾Ø§Ú© Ø¨ÙˆØ¯Ù† Ù†ØªØ§ÛŒØ¬ Ù‚Ø¨Ù„ÛŒ
                    st.rerun()


    def get_risk_class(risk_level):
        risk_classes = {'Ù¾Ø§ÛŒÛŒÙ†': 'risk-low', 'Ù…ØªÙˆØ³Ø·': 'risk-medium', 'Ø¨Ø§Ù„Ø§': 'risk-high', 'Ø¨Ø­Ø±Ø§Ù†ÛŒ': 'risk-critical'}
        return risk_classes.get(risk_level, '')

    def flatten_reference_data(df):
        if 'Ø§Ø±Ø¬Ø§Ø¹' in df.columns:
            df['Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯'] = df['Ø§Ø±Ø¬Ø§Ø¹'].apply(lambda x: x.get('Ø´Ù…Ø§Ø±Ù‡_Ø¨Ù†Ø¯', '') if isinstance(x, dict) else '')
            df['Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡'] = df['Ø§Ø±Ø¬Ø§Ø¹'].apply(lambda x: x.get('Ø´Ù…Ø§Ø±Ù‡_ØµÙØ­Ù‡', '') if isinstance(x, dict) else '')
            df = df.drop('Ø§Ø±Ø¬Ø§Ø¹', axis=1)
        return df

    def flatten_array_fields(df):
        for col in df.columns:
            df[col] = df[col].apply(lambda x: ", ".join(map(str, x)) if isinstance(x, list) else x)
        return df

    # ========================================================================
    # Ø¨Ø®Ø´ 9: ØªÙˆØ§Ø¨Ø¹ UI
    # ========================================================================

    def create_header():
        st.markdown('<div class="modern-header"><h1>ğŸ“Š ØªØ­Ù„ÛŒÙ„Ú¯Ø± Ù‡ÙˆØ´Ù…Ù†Ø¯ ØµÙˆØ±Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ</h1></div>', unsafe_allow_html=True)

    def create_file_upload_section():
        st.markdown('<div class="modern-card"> <h2>ğŸ“ Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§</h2></div>', unsafe_allow_html=True)
        st.divider()
        upload_method = st.radio("Ø±ÙˆØ´ Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ:", ["ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡", "Ù¾ÙˆØ´Ù‡ ZIP"], horizontal=True)
        uploaded_files = []
        if upload_method == "ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡":
            uploaded_files = st.file_uploader("ÙØ§ÛŒÙ„ Ù‡Ø§ÛŒ PDF Ø®ÙˆØ¯ Ø±Ø§ Ø§ÛŒÙ†Ø¬Ø§ Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ú©Ù†ÛŒØ¯", type=['pdf'], accept_multiple_files=True)
        else:
            zip_file = st.file_uploader("ÙØ§ÛŒÙ„ ZIP Ø´Ø§Ù…Ù„ PDF Ù‡Ø§ Ø±Ø§ Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯", type=['zip'])
            if zip_file:
                try:
                    with zipfile.ZipFile(zip_file, 'r') as z:
                        uploaded_files = [{'name': os.path.basename(f.filename), 'content': z.read(f.filename)} for f in z.infolist() if f.filename.lower().endswith('.pdf')]
                    if uploaded_files:
                        st.success(f'âœ… {len(uploaded_files)} ÙØ§ÛŒÙ„ PDF Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø´Ø¯')
                except Exception as e:
                    st.error(f'âŒ Ø®Ø·Ø§: {e}')
        return uploaded_files

    def process_files(uploaded_files):
        analyzer = FinancialAnalyzer()
        results = []
        st.markdown('<div class="modern-card"><h3>Ø¯Ø± Ø­Ø§Ù„ Ù¾Ø±Ø¯Ø§Ø²Ø´...</h3></div>', unsafe_allow_html=True)
        progress_bar = st.progress(0)
        status_placeholder = st.empty()
        total_files = len(uploaded_files)
        start_time = time.time()
        for i, file in enumerate(uploaded_files):
            filename = file['name'] if isinstance(file, dict) else file.name
            file_content = file['content'] if isinstance(file, dict) else file.getvalue()
            status_placeholder.info(f'ğŸ“„ Ø¯Ø± Ø­Ø§Ù„ ØªØ­Ù„ÛŒÙ„ ÙØ§ÛŒÙ„ {i+1} Ø§Ø² {total_files}: {filename}')
            try:
                result = analyzer.extract_table_from_page(file_content, filename)
                results.append((filename, result))
                status_placeholder.success(f'âœ… **{filename}** Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª ØªØ­Ù„ÛŒÙ„ Ø´Ø¯.')
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Failed to process {filename}: {error_msg}\n{traceback.format_exc()}")
                results.append((filename, {"error": f"Ø®Ø·Ø§ Ø¯Ø± ØªØ­Ù„ÛŒÙ„: {error_msg}"}))
                status_placeholder.error(f'âŒ Ø®Ø·Ø§ Ø¯Ø± ØªØ­Ù„ÛŒÙ„ **{filename}**: {error_msg[:100]}...')
            progress_bar.progress((i + 1) / total_files)
        total_duration = time.time() - start_time
        status_placeholder.success(f'ğŸ‰ ØªØ­Ù„ÛŒÙ„ ØªÚ©Ù…ÛŒÙ„ Ø´Ø¯! {total_files} ÙØ§ÛŒÙ„ Ø¯Ø± {total_duration/60:.1f} Ø¯Ù‚ÛŒÙ‚Ù‡ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø´Ø¯.')
        return results
    



    @st.cache_data
    def convert_to_excel(results):

        def style_excel_file(file_path):
            wb = load_workbook(file_path)

            # ğŸ¨ ØªØ¹Ø±ÛŒÙ Ø§Ø³ØªØ§ÛŒÙ„ Ø¹Ù…ÙˆÙ…ÛŒ ÙÙˆÙ†Øª Ùˆ ÙˆØ³Ø· Ú†ÛŒÙ†
            base_style = NamedStyle(name="base_style")
            base_style.font = Font(name="Calibri", size=12)
            base_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # ğŸ¨ Ø§Ø³ØªØ§ÛŒÙ„ Ù‡Ø¯Ø±
            header_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
            header_font = Font(name="Calibri", size=12, bold=True)

            # ğŸ¨ Ø±Ù†Ú¯ Ø±Ø¯ÛŒÙâ€ŒÙ‡Ø§ÛŒ ÛŒÚ©ÛŒ Ø¯Ø± Ù…ÛŒØ§Ù†
            row_fill_alt = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")


            for ws in wb.worksheets:

                # Ø³Øª Ú©Ø±Ø¯Ù† Ø§Ø³ØªØ§ÛŒÙ„ Ú©Ù„ Ø´ÛŒØª
                for row in ws.iter_rows():
                    for cell in row:
                        cell.style = base_style

                # ğŸ“Œ Ø§Ø³ØªØ§ÛŒÙ„ Ø¨Ø±Ø§ÛŒ Ù‡Ø¯Ø±Ù‡Ø§ (Ø±Ø¯ÛŒÙ Ø§ÙˆÙ„)
                # Ø§Ø³ØªØ§ÛŒÙ„ Ù‡Ø¯Ø± (Ø±Ø¯ÛŒÙ 1)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # âœ… Ø§ÙØ²Ø§ÛŒØ´ Ø§Ø±ØªÙØ§Ø¹ Ù‡Ø¯Ø±
                ws.row_dimensions[1].height = 35  
     


                # ØªÙ†Ø¸ÛŒÙ… Ø¹Ø±Ø¶ Ø³ØªÙˆÙ†â€ŒÙ‡Ø§ Ø¨Ø§ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø¨Ø±Ø§ÛŒ Ø³ØªÙˆÙ†â€ŒÙ‡Ø§ÛŒ Ù…ØªÙ†ÛŒ Ø·ÙˆÙ„Ø§Ù†ÛŒ
                long_text_columns = [ "Ù†Ú©Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ_Ùˆ_Ù†ØªÛŒØ¬Ù‡_Ú¯ÛŒØ±ÛŒ", "Ø¬Ø²Ø¦ÛŒØ§Øª", "Ø´Ø±Ø­", "Ø®Ù„Ø§ØµÙ‡_Ø¯Ù„Ø§ÛŒÙ„", "Ø¬Ø²ÛŒÛŒØ§Øª_Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_ØªØ¹ÛŒÛŒÙ†_Ø´Ø¯Ù‡"]

                for col in ws.columns:
                    max_length = 0
                    column = get_column_letter(col[0].column)
                    header = str(col[0].value)

                    # Ù…Ø­Ø§Ø³Ø¨Ù‡ Ø·ÙˆÙ„ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§
                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    # âœ… Ø´Ø±Ø·: Ø§Ú¯Ø± Ø³ØªÙˆÙ† Ø¬Ø²Ùˆ Ø³ØªÙˆÙ†â€ŒÙ‡Ø§ÛŒ ØªÙˆØ¶ÛŒØ­ÛŒ Ø¨ÙˆØ¯ â†’ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª Ø§Ø¹Ù…Ø§Ù„ Ú©Ù†
                    if any(x in header for x in long_text_columns):
                        ws.column_dimensions[column].width = min(max(max_length, 20), 40)
                    else:
                        # Ø³ØªÙˆÙ†â€ŒÙ‡Ø§ Ú©Ù…ÛŒ Ø¹Ø±ÛŒØ¶â€ŒØªØ± Ø§Ø² Ø­Ø§Ù„Øª Ù‚Ø¨Ù„ÛŒ
                        ws.column_dimensions[column].width = max(max_length + 3, 15)

                # ğŸŸ¦ Ø§Ø¹Ù…Ø§Ù„ Ø±Ù†Ú¯ ÛŒÚ©ÛŒ Ø¯Ø± Ù…ÛŒØ§Ù† Ø¨Ø±Ø§ÛŒ Ø±Ø¯ÛŒÙâ€ŒÙ‡Ø§
                for r in range(2, ws.max_row+1):
                    if r % 2 == 0:
                        for cell in ws[r]:
                            cell.fill = row_fill_alt
                           # âœ… ØªØ¹Ø±ÛŒÙ Border
                thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
                # âœ… Ø§Ø¹Ù…Ø§Ù„ Border Ø±ÙˆÛŒ Ú©Ù„ Ø³Ù„ÙˆÙ„â€ŒÙ‡Ø§ÛŒ Ø´ÛŒØª
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border


            wb.save(file_path)

 # =======================
    # âœ… Ø¨Ø®Ø´ Ø³Ø§Ø®Øª ÙØ§ÛŒÙ„ Ø§Ú©Ø³Ù„
    # =======================

        temp_dir = tempfile.mkdtemp()
        excel_files = []
        
        for filename, data in results:
            try:
                if "error" in data:
                    continue
                
                report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                
                # âœ… Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ
                try:
                    company_name = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]["Ù†Ø§Ù…_Ø´Ø±Ú©Øª"]
                    financial_year = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]["Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ"]
                    
                    year_match = re.search(r'(\d{4})', financial_year)
                    year = year_match.group(1) if year_match else "Unknown"
                    
                    clean_company_name = re.sub(r'[\\/:"*?<>|]+', "", company_name).strip()
                    if not clean_company_name:
                        clean_company_name = f"Company_{len(excel_files) + 1}"
                    
                    excel_filename = f"{clean_company_name}_{year}.xlsx"
                    
                except:
                    year = "Unknown"
                    excel_filename = f"Company_{len(excel_files) + 1}.xlsx"
                
                output_file = os.path.join(temp_dir, excel_filename)
                
                with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                    
                    # ========================================
                    # Ø¨Ø®Ø´ 1: Ø®Ù„Ø§ØµÙ‡ Ùˆ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ
                    # ========================================
                    try:
                        part1 = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                        df1 = pd.DataFrame.from_dict({
                            k: [v] if not isinstance(v, list) else [", ".join(v)] 
                            for k, v in part1.items()
                        })
                        # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† year
                        df1['year'] = year
                        df1.to_excel(writer, sheet_name="Ø¨Ø®Ø´1_Ø®Ù„Ø§ØµÙ‡", index=False)
                    except Exception as e:
                        logger.warning(f"Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¨Ø®Ø´ 1: {str(e)}")
                    
                    # ========================================
                    # Ø¨Ø®Ø´ 2: ØªØ¬Ø²ÛŒÙ‡ ØªØ­Ù„ÛŒÙ„ Ú¯Ø²Ø§Ø±Ø´
                    # ========================================
                    try:
                        part2 = report.get("Ø¨Ø®Ø´Û²_ØªØ¬Ø²ÛŒÙ‡_ØªØ­Ù„ÛŒÙ„_Ú¯Ø²Ø§Ø±Ø´", {})
                        
                        # Ø´ÛŒØª: Ø¨Ù†Ø¯ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±
                        if "Ø¨Ù†Ø¯_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±" in part2:
                            df_opinion = pd.DataFrame([part2["Ø¨Ù†Ø¯_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±"]])
                            # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† year
                            df_opinion['year'] = year
                            df_opinion.to_excel(writer, sheet_name="Ø¨Ù†Ø¯_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±", index=False)
                        
                        # Ø´ÛŒØª: Ø¨Ù†Ø¯ Ù…Ø¨Ø§Ù†ÛŒ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±
                        if "Ø¨Ù†Ø¯_Ù…Ø¨Ø§Ù†ÛŒ_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±" in part2:
                            basis_data = part2["Ø¨Ù†Ø¯_Ù…Ø¨Ø§Ù†ÛŒ_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±"]
                            if basis_data.get("Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯", False) and "Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡" in basis_data:
                                df_basis = pd.DataFrame(basis_data["Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡"])
                                df_basis = flatten_reference_data(df_basis)
                                df_basis = flatten_array_fields(df_basis)
                                # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† year
                                df_basis['year'] = year
                            else:
                                df_basis = pd.DataFrame([{"Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯": False, "year": year}])
                            df_basis.to_excel(writer, sheet_name="Ø¨Ù†Ø¯_Ù…Ø¨Ø§Ù†ÛŒ_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±", index=False)
                        
                        # Ø´ÛŒØª: Ø¨Ù†Ø¯ ØªØ§Ú©ÛŒØ¯ Ø¨Ø± Ù…Ø·Ø§Ù„Ø¨ Ø®Ø§Øµ
                        if "Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ" in part2:
                            emphasis_data = part2["Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ"]
                            if emphasis_data.get("Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯", False) and "Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡" in emphasis_data:
                                df_emphasis = pd.DataFrame(emphasis_data["Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡"])
                                df_emphasis = flatten_reference_data(df_emphasis)
                                df_emphasis = flatten_array_fields(df_emphasis)
                                # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† year
                                df_emphasis['year'] = year
                                # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† flag Ù…ÙˆØ¶ÙˆØ¹ÛŒØª
                                df_emphasis['Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯'] = True
                            else:
                                df_emphasis = pd.DataFrame([{"Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯": False, "year": year}])
                            df_emphasis.to_excel(writer, sheet_name="Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ", index=False)
                        
                        # Ø´ÛŒØª: Ú¯Ø²Ø§Ø±Ø´ Ø±Ø¹Ø§ÛŒØª Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ
                        if "Ú¯Ø²Ø§Ø±Ø´_Ø±Ø¹Ø§ÛŒØª_Ø§Ù„Ø²Ø§Ù…Ø§Øª_Ù‚Ø§Ù†ÙˆÙ†ÛŒ" in part2:
                            legal_data = part2["Ú¯Ø²Ø§Ø±Ø´_Ø±Ø¹Ø§ÛŒØª_Ø§Ù„Ø²Ø§Ù…Ø§Øª_Ù‚Ø§Ù†ÙˆÙ†ÛŒ"]
                            if legal_data.get("Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯", False) and "ØªØ®Ù„ÙØ§Øª" in legal_data:
                                violations = legal_data["ØªØ®Ù„ÙØ§Øª"]
                                processed_violations = []
                                
                                for violation in violations:
                                    processed_violation = violation.copy()
                                    # ØªØ¨Ø¯ÛŒÙ„ Ù„ÛŒØ³Øª Ø¨Ù‡ Ø±Ø´ØªÙ‡
                                    if "Ù…Ø¨Ø§Ù†ÛŒ_Ù‚Ø§Ù†ÙˆÙ†ÛŒ_Ùˆ_Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§" in processed_violation:
                                        processed_violation["Ù…Ø¨Ø§Ù†ÛŒ_Ù‚Ø§Ù†ÙˆÙ†ÛŒ_Ùˆ_Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§"] = ", ".join(
                                            processed_violation["Ù…Ø¨Ø§Ù†ÛŒ_Ù‚Ø§Ù†ÙˆÙ†ÛŒ_Ùˆ_Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§"]
                                        )
                                    processed_violations.append(processed_violation)
                                
                                df_legal = pd.DataFrame(processed_violations)
                                df_legal = flatten_reference_data(df_legal)
                                df_legal = flatten_array_fields(df_legal)
                                # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† year
                                df_legal['year'] = year
                                # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† flag Ù…ÙˆØ¶ÙˆØ¹ÛŒØª
                                df_legal['Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯'] = True
                            else:
                                df_legal = pd.DataFrame([{"Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯": False, "year": year}])
                            df_legal.to_excel(writer, sheet_name="Ú¯Ø²Ø§Ø±Ø´_Ù‚Ø§Ù†ÙˆÙ†ÛŒ", index=False)
                    
                    except Exception as e:
                        logger.warning(f"Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¨Ø®Ø´ 2: {str(e)}")
                    
                    # ========================================
                    # Ø¨Ø®Ø´ 3: Ú†Ú© Ù„ÛŒØ³Øª Ù…ÙˆØ¶ÙˆØ¹ÛŒ
                    # ========================================
                    try:
                        if "Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ" in report:
                            part3 = report["Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ"]
                            df3 = pd.DataFrame(part3)
                            df3 = flatten_reference_data(df3)
                            df3 = flatten_array_fields(df3)
                            # âœ… Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† year
                            df3['year'] = year
                            df3.to_excel(writer, sheet_name="Ø¨Ø®Ø´3_Ú†Ú©_Ù„ÛŒØ³Øª", index=False)
                            
                    except Exception as e:
                        logger.warning(f"Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¨Ø®Ø´ 3: {str(e)}")
                           
                           
                 # âœ… Ø§Ø¹Ù…Ø§Ù„ Ø§Ø³ØªØ§ÛŒÙ„ Ø¨Ø¹Ø¯ Ø§Ø² Ø°Ø®ÛŒØ±Ù‡ ÙØ§ÛŒÙ„
                style_excel_file(output_file)
                excel_files.append(output_file)
                logger.info(f"Successfully created Excel file: {excel_filename}")
                
            except Exception as e:
                logger.error(f"Ø®Ø·Ø§ Ø¯Ø± Ø§ÛŒØ¬Ø§Ø¯ Excel Ø¨Ø±Ø§ÛŒ {filename}: {str(e)}")
                logger.error(f"Traceback: {traceback.format_exc()}")
        
        return excel_files
    def create_results_section(results):
        if not results:
            return
        st.subheader("ğŸ“ Ù†ØªØ§ÛŒØ¬ ØªÙØµÛŒÙ„ÛŒ ØªØ­Ù„ÛŒÙ„")
        st.divider()
        for filename, result in results:
            with st.container():
                if 'error' in result:
                    st.markdown(f'<div class="new-result-card error-card"><div class="card-header"><h5>{filename} <span class="status-icon error">âœ—</span></h5></div><p class="error-message">Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„: {result["error"]}</p></div>', unsafe_allow_html=True)
                    continue
                try:
                    analysis = result['ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ']['Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ']
                    company_name = analysis.get('Ù†Ø§Ù…_Ø´Ø±Ú©Øª', 'N/A')
                    auditor_name = analysis.get('Ù†Ø§Ù…_Ø­Ø³Ø§Ø¨Ø±Ø³', 'N/A')
                    opinion_type = analysis.get('Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±', 'N/A')
                    risk_level = analysis.get('Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_Ú©Ù„ÛŒ_Ø¨Ù†Ø§_Ø¨Ù‡_Ú¯Ø²Ø§Ø±Ø´', 'N/A')
                    financial_year = analysis.get('Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ', 'N/A')
                    risk_class = get_risk_class(risk_level)
                    st.markdown(f'<div class="new-result-card"><div class="card-header"><h5>{filename} <span class="status-icon success">âœ“</span></h5></div><div class="new-card-grid"><div class="new-info-box"><div class="new-info-label">ğŸ¢ Ø´Ø±Ú©Øª</div><div class="new-info-value">{company_name}</div></div><div class="new-info-box"><div class="new-info-label">ğŸ“… Ø¯ÙˆØ±Ù‡ Ù…Ø§Ù„ÛŒ</div><div class="new-info-value">{financial_year}</div></div><div class="new-info-box"><div class="new-info-label">ğŸ‘¨â€ğŸ’¼ Ø­Ø³Ø§Ø¨Ø±Ø³</div><div class="new-info-value">{auditor_name}</div></div><div class="new-info-box"><div class="new-info-label">ğŸ“‹ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±</div><div class="new-info-value">{opinion_type}</div></div><div class="new-info-box new-risk-box {risk_class}"><div class="new-info-label">âš ï¸ Ø³Ø·Ø­ Ø±ÛŒØ³Ú©</div><div class="new-info-value">{risk_level}</div></div></div></div>', unsafe_allow_html=True)
                except Exception as e:
                    st.warning(f"âš ï¸ Ø®Ø·Ø§ÛŒÛŒ Ø¯Ø± Ù†Ù…Ø§ÛŒØ´ Ù†ØªØ§ÛŒØ¬ ÙØ§ÛŒÙ„ {filename} Ø±Ø® Ø¯Ø§Ø¯: {e}")
        st.markdown("---")
        st.subheader("ğŸ“¥ Ø¯Ø§Ù†Ù„ÙˆØ¯ Ø®Ø±ÙˆØ¬ÛŒâ€ŒÙ‡Ø§")
        if 'show_download_options' not in st.session_state:
            st.session_state.show_download_options = False
        if 'show_individual_files' not in st.session_state:
            st.session_state.show_individual_files = False
        if st.button("ğŸ“‚ Ù†Ù…Ø§ÛŒØ´ Ú¯Ø²ÛŒÙ†Ù‡â€ŒÙ‡Ø§ÛŒ Ø¯Ø§Ù†Ù„ÙˆØ¯", type="primary", key="show_downloads_main"):
            st.session_state.show_download_options = not st.session_state.show_download_options
            st.session_state.show_individual_files = False
        if st.session_state.show_download_options:
            col1, col2 = st.columns(2)
            with col1:
                if st.button("ğŸ”  Ø¯Ø§Ù†Ù„ÙˆØ¯ ÙØ§ÛŒÙ„ Ù‡Ø§ÛŒ Ø§Ú©Ø³Ù„", use_container_width=True, key="toggle_individual_files"):
                    st.session_state.show_individual_files = not st.session_state.show_individual_files
            with col2:
                try:
                    excel_files_zip = convert_to_excel(results)
                    if excel_files_zip:
                        zip_buffer = BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                            for f_path in excel_files_zip:
                                zf.write(f_path, os.path.basename(f_path))
                        zip_buffer.seek(0)
                        st.download_button(label="ğŸ“¦ Ø¯Ø§Ù†Ù„ÙˆØ¯ Ù‡Ù…Ù‡ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ (ZIP)", data=zip_buffer.getvalue(), file_name=f"Financial_Analysis_All_{datetime.now().strftime('%Y%m%d')}.zip", mime="application/zip", use_container_width=True, key="download_zip_final")
                except Exception as e:
                    st.error(f"Ø®Ø·Ø§ Ø¯Ø± Ø¢Ù…Ø§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ ÙØ§ÛŒÙ„ ZIP: {e}")
            if st.session_state.show_individual_files:
                st.markdown("---")
                st.markdown("#### Ù„ÛŒØ³Øª ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ Ø§Ú©Ø³Ù„ Ø¨Ø±Ø§ÛŒ Ø¯Ø§Ù†Ù„ÙˆØ¯:")
                try:
                    excel_files_individual = convert_to_excel(results)
                    if not excel_files_individual:
                        st.warning("Ù‡ÛŒÚ† ÙØ§ÛŒÙ„ Ø§Ú©Ø³Ù„ÛŒ Ø¨Ø±Ø§ÛŒ Ø¯Ø§Ù†Ù„ÙˆØ¯ ØªÙˆÙ„ÛŒØ¯ Ù†Ø´Ø¯.")
                    else:
                        for index, excel_file in enumerate(excel_files_individual):
                            filename = os.path.basename(excel_file)
                            with open(excel_file, 'rb') as f:
                                file_data = f.read()
                            st.download_button(label=f"ğŸ“„ {filename}", data=file_data, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"download_file_{index}", use_container_width=True)
                except Exception as e:
                    st.error(f"Ø®Ø·Ø§ Ø¯Ø± Ø¢Ù…Ø§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ ØªÚ©ÛŒ: {e}")

    def create_stats_section(results):
        if not results:
            st.info("Ø¯Ø§Ø¯Ù‡â€ŒØ§ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ Ø¢Ù…Ø§Ø± ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯.")
            return
        st.markdown('<div class="modern-card">', unsafe_allow_html=True)
        st.subheader("ğŸ“ˆ Ø¢Ù…Ø§Ø± ØªØ­Ù„ÛŒÙ„")
        st.divider()
        successful = sum(1 for _, r in results if 'error' not in r)
        failed = len(results) - successful
        risk_counts = {'Ù¾Ø§ÛŒÛŒÙ†': 0, 'Ù…ØªÙˆØ³Ø·': 0, 'Ø¨Ø§Ù„Ø§': 0, 'Ø¨Ø­Ø±Ø§Ù†ÛŒ': 0}
        opinion_types = {'Ù…Ù‚Ø¨ÙˆÙ„': 0, 'Ù…Ø´Ø±ÙˆØ·': 0, 'Ù…Ø±Ø¯ÙˆØ¯': 0, 'Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±': 0}
        company_data = []
        for filename, result in results:
            if 'error' not in result:
                try:
                    analysis = result['ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ']['Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ']
                    risk = analysis['Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_Ú©Ù„ÛŒ_Ø¨Ù†Ø§_Ø¨Ù‡_Ú¯Ø²Ø§Ø±Ø´']
                    if risk in risk_counts: 
                        risk_counts[risk] += 1
                    opinion = analysis['Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±']
                    if opinion in opinion_types:
                        opinion_types[opinion] += 1
                    company_data.append({'Ù†Ø§Ù…_Ø´Ø±Ú©Øª': analysis['Ù†Ø§Ù…_Ø´Ø±Ú©Øª'], 'Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ': analysis['Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ'], 'Ø­Ø³Ø§Ø¨Ø±Ø³': analysis.get('Ù†Ø§Ù…_Ø­Ø³Ø§Ø¨Ø±Ø³', 'Ù†Ø§Ù…Ø´Ø®Øµ'), 'Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±': opinion, 'Ø³Ø·Ø­_Ø±ÛŒØ³Ú©': risk})
                except KeyError: 
                    pass
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
                <div class="metric-modern metric-success">
                    <p>ØªØ­Ù„ÛŒÙ„ Ù…ÙˆÙÙ‚</p>
                    <div class="metric-value-box">
                        <div class="metric-value">{successful}</div>
                    </div>
                </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
                <div class="metric-modern metric-failed">
                    <p>Ù†Ø§Ù…ÙˆÙÙ‚</p>
                    <div class="metric-value-box">
                        <div class="metric-value">{failed}</div>
                    </div>
                </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
                <div class="metric-modern metric-highrisk">
                    <p>Ø±ÛŒØ³Ú© Ø¨Ø§Ù„Ø§ / Ø¨Ø­Ø±Ø§Ù†ÛŒ</p>
                    <div class="metric-value-box">
                        <div class="metric-value">{risk_counts["Ø¨Ø§Ù„Ø§"] + risk_counts["Ø¨Ø­Ø±Ø§Ù†ÛŒ"]}</div>
                    </div>
                </div>
            """, unsafe_allow_html=True)
        with col4:
            st.markdown(f"""
                <div class="metric-modern metric-lowrisk">
                    <p>Ø±ÛŒØ³Ú© Ù¾Ø§ÛŒÛŒÙ†</p>
                    <div class="metric-value-box">
                        <div class="metric-value">{risk_counts["Ù¾Ø§ÛŒÛŒÙ†"]}</div>
                    </div>
                </div>
            """, unsafe_allow_html=True)

        if company_data:
            # <<< Ø¨Ø®Ø´ ØªÙˆØ²ÛŒØ¹ Ø³Ø·Ø­ Ø±ÛŒØ³Ú© Ø¨Ø§ Ø³Ø§Ø®ØªØ§Ø± HTML Ø¬Ø¯ÛŒØ¯ >>>
            st.markdown("<br><h3>ğŸ“Š ØªÙˆØ²ÛŒØ¹ Ø³Ø·Ø­ Ø±ÛŒØ³Ú©</h3>", unsafe_allow_html=True)
            risk_icons = {'Ù¾Ø§ÛŒÛŒÙ†': 'ğŸŸ¢', 'Ù…ØªÙˆØ³Ø·': 'ğŸŸ¡', 'Ø¨Ø§Ù„Ø§': 'ğŸŸ ', 'Ø¨Ø­Ø±Ø§Ù†ÛŒ': 'ğŸ”´'}
            
            col1, col2, col3, col4 = st.columns(4)
            # ØªØ±ØªÛŒØ¨ Ø±Ø§ Ù…Ø·Ø§Ø¨Ù‚ ØªØµÙˆÛŒØ± Ù‡Ø¯Ù ØªØºÛŒÛŒØ± Ù…ÛŒâ€ŒØ¯Ù‡ÛŒÙ… (Ø§Ø² Ø³Ø¨Ø² Ø¨Ù‡ Ù‚Ø±Ù…Ø²)
            risk_types = ['Ù¾Ø§ÛŒÛŒÙ†', 'Ù…ØªÙˆØ³Ø·', 'Ø¨Ø§Ù„Ø§', 'Ø¨Ø­Ø±Ø§Ù†ÛŒ']
            columns = [col1, col2, col3, col4]
            
            for col, risk_type in zip(columns, risk_types):
                with col:
                    st.markdown(f'''
                        <div class="stat-card-v2 {risk_type.lower()}">
                            <div class="card-icon-v2">{risk_icons[risk_type]}</div>
                            <div class="card-number-v2">{risk_counts[risk_type]}</div>
                            <div class="card-title-v2">Ø±ÛŒØ³Ú© {risk_type}</div>
                        </div>
                    ''', unsafe_allow_html=True)

            # <<< Ø¨Ø®Ø´ ØªÙˆØ²ÛŒØ¹ Ù†ÙˆØ¹ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± Ø¨Ø§ Ø³Ø§Ø®ØªØ§Ø± HTML Ø¬Ø¯ÛŒØ¯ >>>
            st.markdown("<br><h3>ğŸ“Š ØªÙˆØ²ÛŒØ¹ Ù†ÙˆØ¹ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±</h3>", unsafe_allow_html=True)
            opinion_icons = {'Ù…Ù‚Ø¨ÙˆÙ„': 'âœ…', 'Ù…Ø´Ø±ÙˆØ·': 'âš ï¸', 'Ù…Ø±Ø¯ÙˆØ¯': 'âŒ', 'Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±': 'â­•'}
            
            col1, col2, col3, col4 = st.columns(4)
            opinion_keys = ['Ù…Ù‚Ø¨ÙˆÙ„', 'Ù…Ø´Ø±ÙˆØ·', 'Ù…Ø±Ø¯ÙˆØ¯', 'Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±']
            columns = [col1, col2, col3, col4]
            
            for col, op_type in zip(columns, opinion_keys):
                with col:
                    st.markdown(f'''
                        <div class="stat-card-v2 {op_type.lower().replace(' ', '-')}">
                            <div class="card-icon-v2">{opinion_icons[op_type]}</div>
                            <div class="card-number-v2">{opinion_types[op_type]}</div>
                            <div class="card-title-v2">{op_type}</div>
                        </div>
                    ''', unsafe_allow_html=True)

                
            st.markdown("<br><h3>ğŸ¢ Ø¬Ø¯ÙˆÙ„ Ø®Ù„Ø§ØµÙ‡ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§</h3>", unsafe_allow_html=True)
            html_table = '<table class="summary-table"><thead><tr><th>Ø´Ø±Ú©Øª</th><th>Ø¯ÙˆØ±Ù‡ Ù…Ø§Ù„ÛŒ</th><th>Ø­Ø³Ø§Ø¨Ø±Ø³</th><th>Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±</th><th>Ø³Ø·Ø­ Ø±ÛŒØ³Ú©</th></tr></thead><tbody>'
            for i, row in enumerate(company_data):
                risk_class = ""
                if row['Ø³Ø·Ø­_Ø±ÛŒØ³Ú©'] == 'Ø¨Ø­Ø±Ø§Ù†ÛŒ': risk_class = "risk-critical"
                elif row['Ø³Ø·Ø­_Ø±ÛŒØ³Ú©'] == 'Ø¨Ø§Ù„Ø§': risk_class = "risk-high"
                elif row['Ø³Ø·Ø­_Ø±ÛŒØ³Ú©'] == 'Ù…ØªÙˆØ³Ø·': risk_class = "risk-medium"
                elif row['Ø³Ø·Ø­_Ø±ÛŒØ³Ú©'] == 'Ù¾Ø§ÛŒÛŒÙ†': risk_class = "risk-low"
                opinion_class = ""
                if row['Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±'] == 'Ù…Ù‚Ø¨ÙˆÙ„': opinion_class = "opinion-accepted"
                elif row['Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±'] == 'Ù…Ø´Ø±ÙˆØ·': opinion_class = "opinion-conditional"
                elif row['Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±'] == 'Ù…Ø±Ø¯ÙˆØ¯': opinion_class = "opinion-rejected"
                elif row['Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±'] == 'Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±': opinion_class = "opinion-no"
                html_table += f'<tr><td>{row["Ù†Ø§Ù…_Ø´Ø±Ú©Øª"]}</td><td>{row["Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ"]}</td><td>{row["Ø­Ø³Ø§Ø¨Ø±Ø³"]}</td><td class="opinion-cell {opinion_class}">{row["Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±"]}</td><td class="risk-cell {risk_class}">{row["Ø³Ø·Ø­_Ø±ÛŒØ³Ú©"]}</td></tr>'
            html_table += '</tbody></table>'
            st.markdown(html_table, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)


    # ========================================================================
# Ø¨Ø®Ø´ 9: ØªÙˆØ§Ø¨Ø¹ UI (Ø§ØµÙ„Ø§Ø­ Ù†Ù‡Ø§ÛŒÛŒ Ø¨Ø±Ø§ÛŒ Ø¢Ù…Ø§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ùˆ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±)
# ========================================================================

    def normalize_company_name(name: str) -> str:
        """
        Ù†Ø±Ù…Ø§Ù„â€ŒØ³Ø§Ø²ÛŒ Ù†Ø§Ù… Ø´Ø±Ú©Øª Ø¨Ø±Ø§ÛŒ ØªØ´Ø®ÛŒØµ ÛŒÚ©Ø³Ø§Ù† Ø¨ÙˆØ¯Ù† Ø¨ÛŒÙ† Ø³Ø§Ù„â€ŒÙ‡Ø§
        - Ø­Ø°Ù ØªÙØ§ÙˆØªâ€ŒÙ‡Ø§ÛŒ Ø¬Ø²Ø¦ÛŒ Ø¯Ø± Ù†Ø§Ù…â€ŒÙ†ÙˆÛŒØ³ÛŒ
        - ÛŒÚ©Ø³Ø§Ù†â€ŒØ³Ø§Ø²ÛŒ Ø­Ø±ÙˆÙ ÙØ§Ø±Ø³ÛŒ Ùˆ Ø¹Ø±Ø¨ÛŒ
        - Ø­Ø°Ù Ú©Ù„Ù…Ø§Øª Ùˆ Ø¹Ù„Ø§Ø¦Ù… Ø§Ø¶Ø§ÙÛŒ
        """
        if not name:
            return ""
        
        # Ø­Ø°Ù ÙØ¶Ø§Ù‡Ø§ÛŒ Ø§Ø¨ØªØ¯Ø§ Ùˆ Ø§Ù†ØªÙ‡Ø§
        name = name.strip()
        
        # ÛŒÚ©Ø³Ø§Ù†â€ŒØ³Ø§Ø²ÛŒ Ø­Ø±ÙˆÙ ÙØ§Ø±Ø³ÛŒ Ùˆ Ø¹Ø±Ø¨ÛŒ
        name = name.replace("ÙŠ", "ÛŒ").replace("Ùƒ", "Ú©")
        name = name.replace("Ø©", "Ù‡").replace("Ø¤", "Ùˆ").replace("Ø¥", "Ø§").replace("Ø£", "Ø§")
        
        # ØªØ¨Ø¯ÛŒÙ„ Ø¨Ù‡ Ø­Ø±ÙˆÙ Ú©ÙˆÚ†Ú© (Ø¨Ø±Ø§ÛŒ Ø²Ø¨Ø§Ù† Ø§Ù†Ú¯Ù„ÛŒØ³ÛŒ)
        name = name.lower()
        
        # Ø­Ø°Ù ØªÙ…Ø§Ù… Ø¹Ù„Ø§Ø¦Ù… Ù†Ú¯Ø§Ø±Ø´ÛŒ Ùˆ Ù¾Ø±Ø§Ù†ØªØ²Ù‡Ø§
        name = re.sub(r'[()\[\]{}ØŒ,.:;""\'`~!@#$%^&*_+=|\\/<>?]', '', name)
        
        # Ø­Ø°Ù Ú©Ù„Ù…Ø§Øª Ùˆ Ø¹Ø¨Ø§Ø±Ø§Øª Ø±Ø§ÛŒØ¬ (case-insensitive)
        patterns_to_remove = [
            r'(?i)Ø´Ø±Ú©Øª\s*',
            r'(?i)Ø³Ù‡Ø§Ù…ÛŒ\s*Ø¹Ø§Ù…\s*',
            r'(?i)Ø³Ù‡Ø§Ù…ÛŒ\s*Ø®Ø§Øµ\s*',
            r'(?i)Ø¨Ø§\s*Ù…Ø³Ø¦ÙˆÙ„ÛŒØª\s*Ù…Ø­Ø¯ÙˆØ¯\s*',
            r'(?i)Ø¹Ø§Ù…\s*',
            r'(?i)Ø®Ø§Øµ\s*',
        ]
        
        for pattern in patterns_to_remove:
            name = re.sub(pattern, '', name)
        
        # Ø­Ø°Ù ÙØ§ØµÙ„Ù‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¶Ø§ÙÛŒ Ùˆ ØªØ¨Ø¯ÛŒÙ„ Ú†Ù†Ø¯ ÙØ§ØµÙ„Ù‡ Ø¨Ù‡ ÛŒÚ© ÙØ§ØµÙ„Ù‡
        name = re.sub(r'\s+', ' ', name)
        
        # Ø­Ø°Ù ÙØ¶Ø§Ù‡Ø§ÛŒ Ø§Ø¨ØªØ¯Ø§ Ùˆ Ø§Ù†ØªÙ‡Ø§ Ù…Ø¬Ø¯Ø¯
        name = name.strip()
        
        return name


    def process_and_prepare_dataframes(results: List) -> (Dict, bool):
        """
        Ø§ÛŒÙ† ØªØ§Ø¨Ø¹ Ù†ØªØ§ÛŒØ¬ JSON Ø±Ø§ Ù…Ø³ØªÙ‚ÛŒÙ…Ø§Ù‹ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ú©Ø±Ø¯Ù‡ØŒ Ø³ØªÙˆÙ† 'year' Ø±Ø§ Ø¨Ù‡ Ù‡Ø± Ø´ÛŒØª Ø§Ø¶Ø§ÙÙ‡ Ù…ÛŒâ€ŒÚ©Ù†Ø¯
        Ùˆ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ Ø±Ø§ Ø¨Ø±Ø§ÛŒ Ø§Ø¯ØºØ§Ù… Ø¢Ù…Ø§Ø¯Ù‡ Ù…ÛŒâ€ŒÚ©Ù†Ø¯.
        """
        all_dataframes = {}
        company_names = {}  # ØªØºÛŒÛŒØ± Ø§Ø² set Ø¨Ù‡ dict Ø¨Ø±Ø§ÛŒ Ù†Ú¯Ù‡Ø¯Ø§Ø±ÛŒ Ù†Ø§Ù… Ø§ØµÙ„ÛŒ Ùˆ Ù†Ø±Ù…Ø§Ù„ Ø´Ø¯Ù‡
        
        # 1. Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù‡Ø± ÙØ§ÛŒÙ„ Ø¨Ù‡ ØµÙˆØ±Øª Ù…Ø¬Ø²Ø§
        for filename, data in results:
            if "error" in data:
                continue
                
            try:
                report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                summary_data = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                
                # Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø³Ø§Ù„ Ùˆ Ù†Ø§Ù… Ø´Ø±Ú©Øª
                company_name = summary_data.get("Ù†Ø§Ù…_Ø´Ø±Ú©Øª", f"Unknown_{filename}")
                normalized_name = normalize_company_name(company_name)
                
                financial_year = summary_data.get("Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ", "")
                year_match = re.search(r'(\d{4})', financial_year)
                year = int(year_match.group(1)) if year_match else None
                
                if year is None:
                    logger.warning(f"Ø³Ø§Ù„ Ù…Ø§Ù„ÛŒ Ø¨Ø±Ø§ÛŒ ÙØ§ÛŒÙ„ {filename} ÛŒØ§ÙØª Ù†Ø´Ø¯. Ø§Ø² Ø§ÛŒÙ† ÙØ§ÛŒÙ„ Ø¯Ø± Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ ØµØ±Ùâ€ŒÙ†Ø¸Ø± Ù…ÛŒâ€ŒØ´ÙˆØ¯.")
                    continue

                # Ø°Ø®ÛŒØ±Ù‡ Ù†Ø§Ù… Ù†Ø±Ù…Ø§Ù„ Ø´Ø¯Ù‡ Ø¨Ø§ Ù†Ø§Ù… Ø§ØµÙ„ÛŒ
                if normalized_name not in company_names:
                    company_names[normalized_name] = company_name

                # 2. Ø³Ø§Ø®Øª Ø¯ÛŒØªØ§ÙØ±ÛŒÙ… Ø¨Ø±Ø§ÛŒ Ù‡Ø± Ø´ÛŒØª Ùˆ Ø§Ø¶Ø§ÙÙ‡ Ú©Ø±Ø¯Ù† Ø³ØªÙˆÙ† 'year'
                # Ø´ÛŒØª Ø®Ù„Ø§ØµÙ‡
                if 'df_summary' not in all_dataframes: 
                    all_dataframes['df_summary'] = []
                df1 = pd.DataFrame([summary_data])
                df1['year'] = year
                all_dataframes['df_summary'].append(df1)
                
                # Ø´ÛŒØª ØªØ§Ú©ÛŒØ¯ Ø¨Ø± Ù…Ø·Ø§Ù„Ø¨ Ø®Ø§Øµ
                part2 = report.get("Ø¨Ø®Ø´Û²_ØªØ¬Ø²ÛŒÙ‡_ØªØ­Ù„ÛŒÙ„_Ú¯Ø²Ø§Ø±Ø´", {})
                if "Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ" in part2:
                    if 'df_emphasis' not in all_dataframes: 
                        all_dataframes['df_emphasis'] = []
                    emphasis_data = part2["Ø¨Ù†Ø¯_ØªØ§Ú©ÛŒØ¯_Ø¨Ø±_Ù…Ø·Ø§Ù„Ø¨_Ø®Ø§Øµ"]
                    if emphasis_data.get("Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯", False) and "Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡" in emphasis_data:
                        df_emphasis = pd.DataFrame(emphasis_data["Ù…ÙˆØ§Ø±Ø¯_Ù…Ø·Ø±Ø­_Ø´Ø¯Ù‡"])
                        df_emphasis['year'] = year
                        df_emphasis['Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯'] = True
                        all_dataframes['df_emphasis'].append(df_emphasis)
                    else:
                        all_dataframes['df_emphasis'].append(pd.DataFrame([{'Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯': False, 'year': year}]))

                # Ø´ÛŒØª ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ
                if "Ú¯Ø²Ø§Ø±Ø´_Ø±Ø¹Ø§ÛŒØª_Ø§Ù„Ø²Ø§Ù…Ø§Øª_Ù‚Ø§Ù†ÙˆÙ†ÛŒ" in part2:
                    if 'df_violations' not in all_dataframes: 
                        all_dataframes['df_violations'] = []
                    legal_data = part2["Ú¯Ø²Ø§Ø±Ø´_Ø±Ø¹Ø§ÛŒØª_Ø§Ù„Ø²Ø§Ù…Ø§Øª_Ù‚Ø§Ù†ÙˆÙ†ÛŒ"]
                    if legal_data.get("Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯", False) and "ØªØ®Ù„ÙØ§Øª" in legal_data:
                        df_legal = pd.DataFrame(legal_data["ØªØ®Ù„ÙØ§Øª"])
                        df_legal['year'] = year
                        df_legal['Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯'] = True
                        all_dataframes['df_violations'].append(df_legal)
                    else:
                        all_dataframes['df_violations'].append(pd.DataFrame([{'Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯': False, 'year': year}]))
                
                # Ø´ÛŒØª Ú†Ú©â€ŒÙ„ÛŒØ³Øª
                if "Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ" in report:
                    if 'df_checklist' not in all_dataframes: 
                        all_dataframes['df_checklist'] = []
                    df3 = pd.DataFrame(report["Ø¨Ø®Ø´Û³_Ú†Ú©_Ù„ÛŒØ³Øª_Ù…ÙˆØ¶ÙˆØ¹ÛŒ"])
                    df3['year'] = year
                    all_dataframes['df_checklist'].append(df3)

            except Exception as e:
                logger.error(f"Ø®Ø·Ø§ Ø¯Ø± Ø¢Ù…Ø§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ø¯ÛŒØªØ§ÙØ±ÛŒÙ… Ø¨Ø±Ø§ÛŒ ÙØ§ÛŒÙ„ {filename}: {e}")
                continue

        # 3. Ø§Ø¹ØªØ¨Ø§Ø±Ø³Ù†Ø¬ÛŒ Ù†Ø§Ù… Ø´Ø±Ú©Øª (Ø§ØµÙ„Ø§Ø­ Ø´Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ø§Ø¬Ø§Ø²Ù‡ Ø¯Ø§Ø¯Ù† Ø¨Ù‡ Ù†Ø§Ù…â€ŒÙ‡Ø§ÛŒ Ù…ØªÙØ§ÙˆØª)
        if len(company_names) > 1:
            original_names = list(company_names.values())
            # Ø¨Ù‡ Ø¬Ø§ÛŒ Ø¨Ø±Ú¯Ø±Ø¯Ø§Ù†Ø¯Ù† FalseØŒ ÙÙ‚Ø· Ø§Ø·Ù„Ø§Ø¹â€ŒØ±Ø³Ø§Ù†ÛŒ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
            st.info(f"â„¹ï¸ Ø¯Ø± Ø­Ø§Ù„ ØªØ¬Ù…ÛŒØ¹ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§ÛŒ {len(company_names)} Ø´Ø±Ú©Øª Ù…Ø®ØªÙ„Ù Ø¨Ø±Ø§ÛŒ ØªØ­Ù„ÛŒÙ„ Ú¯Ø±ÙˆÙ‡ÛŒ: {', '.join(original_names)}")
            # return {}, False  <-- Ø§ÛŒÙ† Ø®Ø· Ø­Ø°Ù ÛŒØ§ Ú©Ø§Ù…Ù†Øª Ø´Ø¯

        # 4. Ø§Ø¯ØºØ§Ù… Ù†Ù‡Ø§ÛŒÛŒ Ø¯ÛŒØªØ§ÙØ±ÛŒÙ…â€ŒÙ‡Ø§
        merged_data = {}
        for key, df_list in all_dataframes.items():
            if df_list:
                merged_data[key] = pd.concat(df_list, ignore_index=True)

        return merged_data, True


    def load_font_as_base64(font_path):
        """
        ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª Ø¨Ø§ÛŒÙ†Ø±ÛŒ Ø±Ø§ Ù…ÛŒâ€ŒØ®ÙˆØ§Ù†Ø¯ Ùˆ Ø¢Ù† Ø±Ø§ Ø¨Ù‡ Ø±Ø´ØªÙ‡ Base64 Ø¨Ø§ Ù¾ÛŒØ´ÙˆÙ†Ø¯ data URI ØªØ¨Ø¯ÛŒÙ„ Ù…ÛŒâ€ŒÚ©Ù†Ø¯.
        """
        try:
            with open(font_path, "rb") as f:  # 'rb' Ø¨Ø±Ø§ÛŒ Ø®ÙˆØ§Ù†Ø¯Ù† ÙØ§ÛŒÙ„ Ø¨Ù‡ ØµÙˆØ±Øª Ø¨Ø§ÛŒÙ†Ø±ÛŒ
                font_data = f.read()
            
            # Ø¯Ø§Ø¯Ù‡ Ø¨Ø§ÛŒÙ†Ø±ÛŒ Ø±Ø§ Ø¨Ù‡ Base64 ØªØ¨Ø¯ÛŒÙ„ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
            base64_encoded_data = base64.b64encode(font_data).decode('utf-8')
            
            # Ø¨Ø± Ø§Ø³Ø§Ø³ Ù¾Ø³ÙˆÙ†Ø¯ ÙØ§ÛŒÙ„ØŒ Ù¾ÛŒØ´ÙˆÙ†Ø¯ ØµØ­ÛŒØ­ Ø±Ø§ ØªØ¹ÛŒÛŒÙ† Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
            if font_path.endswith(".woff2"):
                mime_type = "font/woff2"
            elif font_path.endswith(".woff"):
                mime_type = "font/woff"
            elif font_path.endswith(".ttf"):
                mime_type = "font/truetype"
            else:
                # Ø§Ú¯Ø± ÙØ±Ù…Øª Ø¯ÛŒÚ¯Ø±ÛŒ Ø¨ÙˆØ¯ØŒ ÛŒÚ© Ù…Ù‚Ø¯Ø§Ø± Ù¾ÛŒØ´â€ŒÙØ±Ø¶ Ø¯Ø± Ù†Ø¸Ø± Ù…ÛŒâ€ŒÚ¯ÛŒØ±ÛŒÙ…
                mime_type = "application/font-octet-stream"

            # Ø±Ø´ØªÙ‡ Ù†Ù‡Ø§ÛŒÛŒ Ø¨Ø§ ÙØ±Ù…Øª data URI Ø±Ø§ Ø¨Ø±Ù…ÛŒâ€ŒÚ¯Ø±Ø¯Ø§Ù†ÛŒÙ…
            return f"data:{mime_type};base64,{base64_encoded_data}"

        except FileNotFoundError:
            st.error(f"Ø®Ø·Ø§ÛŒ Ø­ÛŒØ§ØªÛŒ: ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª Ø¯Ø± Ù…Ø³ÛŒØ± '{font_path}' Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯.")
            return None
        except Exception as e:
            st.error(f"Ø®Ø·Ø§ Ø¯Ø± Ù¾Ø±Ø¯Ø§Ø²Ø´ ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª: {e}")
            return None

# Ø¨Ø®Ø´ Ø§ØµÙ„Ø§Ø­â€ŒØ´Ø¯Ù‡ Ø¨Ø±Ø§ÛŒ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ø¨Ø§ Ø¹Ù†Ø§ÙˆÛŒÙ† markdown Ùˆ expander ØªÙˆØ¶ÛŒØ­Ø§Øª

    def create_charts_section(results):
        # 1. Ø¨Ø±Ø±Ø³ÛŒ Ø§ÙˆÙ„ÛŒÙ‡ Ù†ØªØ§ÛŒØ¬ (Ø§ÛŒÙ† Ø¨Ø®Ø´ Ø¨Ø¯ÙˆÙ† ØªØºÛŒÛŒØ± Ø§Ø³Øª)
        if not results or not any('error' not in r for _, r in results):
            st.info("Ø¯Ø§Ø¯Ù‡ Ù…Ø¹ØªØ¨Ø±ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯.")
            return

        st.markdown('<div class="modern-card">', unsafe_allow_html=True)
        st.subheader("ğŸ“ˆ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ùˆ ØªØ­Ù„ÛŒÙ„ Ø±ÙˆÙ†Ø¯")

        # ==================== Ø¨Ø®Ø´ Ú©Ù„ÛŒØ¯ÛŒ: Ø®ÙˆØ§Ù†Ø¯Ù† ÙÙˆÙ†Øª Ø§Ø² ÙØ§ÛŒÙ„ ====================
        # âœ… ØªÙ…Ø§Ù… Ø§ÛŒÙ† Ù…Ù†Ø·Ù‚ Ø¨Ø§ÛŒØ¯ Ø¯Ø§Ø®Ù„ ØªØ§Ø¨Ø¹ Ø¨Ø§Ø´Ø¯
        
        # 2. ÙÙˆÙ†Øª Ø±Ø§ Ø§Ø² ÙØ§ÛŒÙ„ Ù…ØªÙ†ÛŒ Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡ Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
        font_base64_string = load_font_as_base64("fonts/BMITRA.woff2")

        # 3. Ø¨Ø±Ø±Ø³ÛŒ Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ… Ú©Ù‡ Ø¢ÛŒØ§ ÙÙˆÙ†Øª Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø´Ø¯Ù‡ Ø§Ø³Øª ÛŒØ§ Ø®ÛŒØ±
        # Ø§Ú¯Ø± ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª Ù¾ÛŒØ¯Ø§ Ù†Ø´Ø¯ØŒ ÛŒÚ© Ù¾ÛŒØ§Ù… Ù‡Ø´Ø¯Ø§Ø± Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ùˆ Ø§Ø¬Ø±Ø§ÛŒ Ø§ÛŒÙ† ØªØ§Ø¨Ø¹ Ø±Ø§ Ù…ØªÙˆÙ‚Ù Ù…ÛŒâ€ŒÚ©Ù†ÛŒÙ…
        if not font_base64_string:
            st.warning("ÙØ§ÛŒÙ„ ÙÙˆÙ†Øª Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ù†Ø´Ø¯ØŒ Ø¯Ø± Ù†ØªÛŒØ¬Ù‡ ÙÙ‡Ø±Ø³Øª Ú¯Ø±Ø§ÙÛŒÚ©ÛŒ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ù†Ù…ÛŒâ€ŒØ´ÙˆØ¯.")
            # Ø§ÛŒÙ† 'return' Ù…ØªØ¹Ù„Ù‚ Ø¨Ù‡ ØªØ§Ø¨Ø¹ create_charts_section Ø§Ø³Øª Ùˆ Ø¨Ø§Ø¹Ø« Ø®Ø±ÙˆØ¬ Ø§Ø² Ø¢Ù† Ù…ÛŒâ€ŒØ´ÙˆØ¯
            st.markdown("</div>", unsafe_allow_html=True)
            return

        # 4. Ø§Ú¯Ø± ÙÙˆÙ†Øª Ø¨Ø§ Ù…ÙˆÙÙ‚ÛŒØª Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ø´Ø¯ØŒ Ø§Ø¯Ø§Ù…Ù‡ Ù…ÛŒâ€ŒØ¯Ù‡ÛŒÙ… Ùˆ Ú©Ø§Ù…Ù¾ÙˆÙ†Ù†Øª HTML Ø±Ø§ Ù…ÛŒâ€ŒØ³Ø§Ø²ÛŒÙ…
        #ØªÙ†Ø¸ÛŒÙ… ÙÙˆÙ†Øª Ø¨Ø±Ø§ÛŒ ÙÙ‡Ø±Ø³Øª Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§
        html_with_embedded_font = f"""
        <!DOCTYPE html>
        <html lang="fa" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                @font-face {{
                    font-family: 'B Mitra';
                    src: url('{font_base64_string}') format('woff2');
                    font-weight: normal;
                    font-style: normal;
                }}
                * {{
                    font-family: 'B Mitra','Noto Naskh Arabic', Tahoma, Arial, sans-serif !important;
                }}
            </style>
        </head>
        <body>
            <div style="background: #f0f4f8; padding: 1.5rem; border-radius: 12px; margin-bottom: 2rem; direction: rtl;">
                <h2 style="text-align: center; color: #2c3e50; margin-bottom: 1.5rem; margin-top: 0; font-size: 1.7rem;">ğŸ“Š ÙÙ‡Ø±Ø³Øª Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§</h2>
                <div style="background: white; padding: 1.2rem; border-radius: 10px; margin-bottom: 1rem; border-right: 5px solid #667eea; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    <h3 style="color: #4c51bf; margin: 0 0 0.5rem 0; font-size: 1.4rem;">ğŸ“Š Ø¨Ø®Ø´ Û±: ØªØ­Ù„ÛŒÙ„ Ø±ÙˆÙ†Ø¯Ù‡Ø§ÛŒ Ú©Ù„Ø§Ù† Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ</h3>
                </div>
                <div style="background: white; padding: 1.2rem; border-radius: 10px; margin-bottom: 1rem; border-right: 5px solid #f97316; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    <h3 style="color: #4c51bf; margin: 0 0 0.5rem 0; font-size: 1.4rem;">âš ï¸ Ø¨Ø®Ø´ Û²: ØªØ­Ù„ÛŒÙ„ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡ Ø´Ø¯Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´</h3>
                </div>
                <div style="background: white; padding: 1.2rem; border-radius: 10px; margin-bottom: 1rem; border-right: 5px solid #f59e0b; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    <h3 style="color: #4c51bf; margin: 0 0 0.5rem 0; font-size: 1.4rem;">âš–ï¸ Ø¨Ø®Ø´ Û³: ØªØ­Ù„ÛŒÙ„ ØªØ®Ù„ÙØ§Øª Ùˆ Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ</h3>
                </div>
                <div style="background: white; padding: 1.2rem; border-radius: 10px; border-right: 5px solid #ef4444; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    <h3 style="color: #4c51bf; margin: 0 0 0.5rem 0; font-size: 1.4rem;">ğŸ”¥ Ø¨Ø®Ø´ Û´: Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ</h3>
                </div>
                <div style="margin-top: 1.5rem; padding: 1rem; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); border-radius: 10px; text-align: center;">
                    <p style="color: white; margin: 0; font-size: 1.4rem; font-weight: bold;">ğŸ“ˆ Ù…Ø¬Ù…ÙˆØ¹: 7 Ù†Ù…ÙˆØ¯Ø§Ø± ØªØ­Ù„ÛŒÙ„ÛŒ Ø¯Ø± 4 Ø¨Ø®Ø´ </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        components.html(html_with_embedded_font, height=600)
                
        # st.divider()

        try:
            font = setup_persian_font()
            
            # Ø¢Ù…Ø§Ø¯Ù‡â€ŒØ³Ø§Ø²ÛŒ Ùˆ Ø§Ø¯ØºØ§Ù… Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§
            merged_data, is_consistent = process_and_prepare_dataframes(results)
            
            if not is_consistent or not merged_data:
                st.markdown("</div>", unsafe_allow_html=True)
                return

            df_summary = merged_data.get("df_summary")
            df_emphasis = merged_data.get("df_emphasis")
            df_checklist = merged_data.get("df_checklist")
            df_violations = merged_data.get("df_violations")

            # ====================================================================
            # Ø¨Ø®Ø´ Û±: Ø±ÙˆÙ†Ø¯Ù‡Ø§ÛŒ Ú©Ù„Ø§Ù†
            # ====================================================================
            st.markdown('''
                <h2 style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 2rem; 
                        border-radius: 20px; 
                        margin: 2.5rem 0; 
                        box-shadow: 0 10px 30px rgba(102, 126, 234, 0.3);
                        color: white; 
                        text-align: center; 
                        font-size: 1.8rem; 
                        font-weight: 700; 
                        letter-spacing: 1px;">
                    ğŸ“Š Ø¨Ø®Ø´ Û±: ØªØ­Ù„ÛŒÙ„ Ø±ÙˆÙ†Ø¯Ù‡Ø§ÛŒ Ú©Ù„Ø§Ù† Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ
                </h2>
                ''', unsafe_allow_html=True)
       
            
            col1, col2 = st.columns(2, gap="large")
            
            # Ù†Ù…ÙˆØ¯Ø§Ø± Û±: Ø±ÙˆÙ†Ø¯ Ø³Ø·Ø­ Ø±ÛŒØ³Ú©
            with col1:
                if df_summary is not None and not df_summary.empty:
                    st.markdown('''
                        <h3 style="background: rgba(240, 147, 251, 0.3); 
                                padding: 0.8rem 1.5rem; 
                                border-radius: 12px; 
                                margin: 0.5rem 0 0 0;
                                color: #2c3e50; 
                                text-align: center; 
                                font-size: 1.2rem; 
                                font-weight: 600;">
                            âš ï¸ Ø±ÙˆÙ†Ø¯ Ø³Ø·Ø­ Ø±ÛŒØ³Ú© Ú©Ù„ÛŒ
                        </h3>
                    ''', unsafe_allow_html=True)
                    
                    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                    fig1 = plot_risk_trend(df_summary, font)
                    st.pyplot(fig1)
                    plt.close(fig1)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù…ÙˆØ¯Ø§Ø± Ø±ÙˆÙ†Ø¯ Ø±ÛŒØ³Ú©"):
                        st.markdown("""
                        <div style="text-align: right; direction: rtl; padding: 1rem;">
                        
                        **Ø§ÛŒÙ† Ù†Ù…ÙˆØ¯Ø§Ø± Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                        
                        - **Ù…Ø­ÙˆØ± Ø¹Ù…ÙˆØ¯ÛŒ**: Ø³Ø·Ø­ Ø±ÛŒØ³Ú© (Ù¾Ø§ÛŒÛŒÙ†ØŒ Ù…ØªÙˆØ³Ø·ØŒ Ø¨Ø§Ù„Ø§ØŒ Ø¨Ø­Ø±Ø§Ù†ÛŒ)
                        - **Ù…Ø­ÙˆØ± Ø§ÙÙ‚ÛŒ**: Ø³Ø§Ù„â€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ Ù…ÙˆØ±Ø¯ Ø¨Ø±Ø±Ø³ÛŒ
                        - **Ø®Ø· Ø±ÙˆÙ†Ø¯**: ØªØºÛŒÛŒØ±Ø§Øª Ø³Ø·Ø­ Ø±ÛŒØ³Ú© Ú©Ù„ÛŒ Ø¯Ø± Ø·ÙˆÙ„ Ø²Ù…Ø§Ù†
                        
                        **Ù†Ú©Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ:**
                        - Ø§ÙØ²Ø§ÛŒØ´ Ø±ÙˆÙ†Ø¯ Ø¨Ù‡ Ø³Ù…Øª Ø¨Ø§Ù„Ø§ Ù†Ø´Ø§Ù†â€ŒØ¯Ù‡Ù†Ø¯Ù‡ Ø§ÙØ²Ø§ÛŒØ´ Ø±ÛŒØ³Ú© Ø§Ø³Øª
                        - Ú©Ø§Ù‡Ø´ Ø±ÙˆÙ†Ø¯ Ù†Ø´Ø§Ù†â€ŒØ¯Ù‡Ù†Ø¯Ù‡ Ø¨Ù‡Ø¨ÙˆØ¯ ÙˆØ¶Ø¹ÛŒØª Ú©Ù†ØªØ±Ù„â€ŒÙ‡Ø§ÛŒ Ø¯Ø§Ø®Ù„ÛŒ Ùˆ Ú©Ø§Ù‡Ø´ Ø±ÛŒØ³Ú©
                        - Ø«Ø¨Ø§Øª Ø¯Ø± ÛŒÚ© Ø³Ø·Ø­ Ù†Ø´Ø§Ù†â€ŒØ¯Ù‡Ù†Ø¯Ù‡ ÙˆØ¶Ø¹ÛŒØª Ù¾Ø§ÛŒØ¯Ø§Ø± Ø³Ø§Ø²Ù…Ø§Ù† Ø§Ø³Øª
                        
                        </div>
                        """, unsafe_allow_html=True)
            
            # Ù†Ù…ÙˆØ¯Ø§Ø± Û²: Ø±ÙˆÙ†Ø¯ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±
            with col2:
                if df_summary is not None and not df_summary.empty:
                    st.markdown('''
                        <h3 style="background: rgba(250, 112, 154, 0.3); 
                                padding: 0.8rem 1.5rem; 
                                border-radius: 12px; 
                                margin: 0.5rem 0 0 0;
                                color: #2c3e50; 
                                text-align: center; 
                                font-size: 1.2rem; 
                                font-weight: 600;">
                            âœ… Ø±ÙˆÙ†Ø¯ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± Ø­Ø³Ø§Ø¨Ø±Ø³
                        </h3>
                    ''', unsafe_allow_html=True)
                    
                    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                    fig2 = plot_opinion_trend(df_summary, font)
                    st.pyplot(fig2)
                    plt.close(fig2)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù…ÙˆØ¯Ø§Ø± Ø±ÙˆÙ†Ø¯ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±"):
                        st.markdown("""
                        <div style="text-align: right; direction: rtl; padding: 1rem;">
                        
                        **Ø§ÛŒÙ† Ù†Ù…ÙˆØ¯Ø§Ø± Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                        
                        - **Ù…Ø­ÙˆØ± Ø¹Ù…ÙˆØ¯ÛŒ**: Ù†ÙˆØ¹ Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± (Ù…Ù‚Ø¨ÙˆÙ„ØŒ Ù…Ø´Ø±ÙˆØ·ØŒ Ù…Ø±Ø¯ÙˆØ¯ØŒ Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±)
                        - **Ù…Ø­ÙˆØ± Ø§ÙÙ‚ÛŒ**: Ø³Ø§Ù„â€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ
                        - **Ø®Ø· Ø±ÙˆÙ†Ø¯**: ØªØºÛŒÛŒØ±Ø§Øª Ù†ÙˆØ¹ Ù†Ø¸Ø± Ø­Ø³Ø§Ø¨Ø±Ø³ Ø¯Ø± Ø·ÙˆÙ„ Ø²Ù…Ø§Ù†
                        
                        **ØªÙØ³ÛŒØ±:**
                        - **Ù…Ù‚Ø¨ÙˆÙ„**: ØµÙˆØ±Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ Ø¹Ø§Ø±ÛŒ Ø§Ø² ØªØ­Ø±ÛŒÙ Ø¨Ø§ Ø§Ù‡Ù…ÛŒØª
                        - **Ù…Ø´Ø±ÙˆØ·**: ÙˆØ¬ÙˆØ¯ Ù…Ø­Ø¯ÙˆØ¯ÛŒØª ÛŒØ§ Ø§Ù†Ø­Ø±Ø§Ù Ù‚Ø§Ø¨Ù„ ØªÙˆØ¬Ù‡
                        - **Ù…Ø±Ø¯ÙˆØ¯**: Ø¹Ø¯Ù… Ø§Ù†Ø·Ø¨Ø§Ù‚ Ø¨Ø§ Ø§Ø³ØªØ§Ù†Ø¯Ø§Ø±Ø¯Ù‡Ø§
                        - **Ø¹Ø¯Ù… Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±**: Ø¹Ø¯Ù… Ø§Ù…Ú©Ø§Ù† Ø±Ø³ÛŒØ¯Ú¯ÛŒ Ú©Ø§ÙÛŒ
                        
                        </div>
                        """, unsafe_allow_html=True)

 

            # ====================================================================
            # Ø¨Ø®Ø´ Û³: Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡
            # ====================================================================
            st.markdown('''
                <h2 style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); 
                        padding: 2rem; 
                        border-radius: 20px; 
                        margin: 2.5rem 0; 
                        box-shadow: 0 10px 30px rgba(79, 172, 254, 0.3);
                        color: white; 
                        text-align: center; 
                        font-size: 1.8rem; 
                        font-weight: 700; 
                        letter-spacing: 1px;">
                    âš ï¸ Ø¨Ø®Ø´ Û²: ØªØ­Ù„ÛŒÙ„ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡ Ø´Ø¯Ù‡ Ø¯Ø± Ú¯Ø²Ø§Ø±Ø´
                </h2>
            ''', unsafe_allow_html=True)
            
            if df_emphasis is not None and not df_emphasis.empty and df_emphasis['Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯'].any():
                col3, col4 = st.columns([1, 1], gap="large")
                
                # Ù†Ù…ÙˆØ¯Ø§Ø± Û´: Ø³ØªÙˆÙ†ÛŒ Ø§Ù†Ø¨Ø§Ø´ØªÙ‡ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§
                with col3:
                    st.markdown('''
                        <h3 style="background: rgba(255, 183, 197, 0.3); 
                                padding: 0.8rem 1.5rem; 
                                border-radius: 12px; 
                                margin: 0.5rem 0 0 0;
                                color: #2c3e50; 
                                text-align: center; 
                                font-size: 1.2rem; 
                                font-weight: 600;">
                            ğŸ“Š Ø±ÙˆÙ†Ø¯ ØªØ¹Ø¯Ø§Ø¯ Ùˆ ØªØ±Ú©ÛŒØ¨ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§
                        </h3>
                    ''', unsafe_allow_html=True)
                    
                    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                    fig4 = plot_risk_stacked_bar(df_emphasis, font)
                    st.pyplot(fig4)
                    plt.close(fig4)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù…ÙˆØ¯Ø§Ø± Ø³ØªÙˆÙ†ÛŒ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§"):
                        st.markdown("""
                        <div style="text-align: right; direction: rtl; padding: 1rem;">
                        
                        **Ø§ÛŒÙ† Ù†Ù…ÙˆØ¯Ø§Ø± Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                        
                        - **Ø³ØªÙˆÙ†â€ŒÙ‡Ø§**: ØªØ¹Ø¯Ø§Ø¯ Ú©Ù„ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡ Ø´Ø¯Ù‡ Ø¯Ø± Ù‡Ø± Ø³Ø§Ù„
                        - **Ø±Ù†Ú¯â€ŒÙ‡Ø§ÛŒ Ù…Ø®ØªÙ„Ù**: Ø¯Ø³ØªÙ‡â€ŒÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ Ø±ÛŒØ³Ú© (Ø§Ø¹ØªØ¨Ø§Ø±ÛŒØŒ Ø¨Ø§Ø²Ø§Ø±ØŒ Ø¹Ù…Ù„ÛŒØ§ØªÛŒ Ùˆ...)
                        - **Ø¹Ø¯Ø¯ Ø±ÙˆÛŒ Ù‡Ø± Ø¨Ø®Ø´**: ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ§Ø±Ø¯ Ø¢Ù† Ø¯Ø³ØªÙ‡ Ø±ÛŒØ³Ú©
                        
                        **ØªØ­Ù„ÛŒÙ„:**
                        - Ø§ÙØ²Ø§ÛŒØ´ Ø§Ø±ØªÙØ§Ø¹ Ø³ØªÙˆÙ† = Ø§ÙØ²Ø§ÛŒØ´ ØªØ¹Ø¯Ø§Ø¯ Ú©Ù„ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§
                        - ØªØºÛŒÛŒØ± ØªØ±Ú©ÛŒØ¨ Ø±Ù†Ú¯â€ŒÙ‡Ø§ = ØªØºÛŒÛŒØ± Ø¯Ø± Ù†ÙˆØ¹ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ ØºØ§Ù„Ø¨
                        - ØºÙ„Ø¨Ù‡ ÛŒÚ© Ø±Ù†Ú¯ = ØªÙ…Ø±Ú©Ø² Ø±ÛŒØ³Ú© Ø¯Ø± ÛŒÚ© Ø¯Ø³ØªÙ‡ Ø®Ø§Øµ
                        
                        **Ù‡Ø´Ø¯Ø§Ø±:** Ø§ÙØ²Ø§ÛŒØ´ Ù†Ø§Ú¯Ù‡Ø§Ù†ÛŒ Ø¯Ø± ÛŒÚ© Ø¯Ø³ØªÙ‡ Ø±ÛŒØ³Ú© Ø®Ø§Øµ Ù†ÛŒØ§Ø² Ø¨Ù‡ ØªÙˆØ¬Ù‡ ÙˆÛŒÚ˜Ù‡ Ø¯Ø§Ø±Ø¯.
                        
                        </div>
                        """, unsafe_allow_html=True)
                
                # Ù†Ù…ÙˆØ¯Ø§Ø± Ûµ: Sunburst Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§
                with col4:
                    st.markdown('''
                        <h3 style="background: rgba(210, 153, 194, 0.3); 
                                padding: 0.8rem 1.5rem; 
                                border-radius: 12px; 
                                margin: 0.5rem 0 0 0;
                                color: #2c3e50; 
                                text-align: center; 
                                font-size: 1.2rem; 
                                font-weight: 600;">
                            ğŸ¯ ØªÙˆØ²ÛŒØ¹ Ø³Ù„Ø³Ù„Ù‡â€ŒÙ…Ø±Ø§ØªØ¨ÛŒ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§
                        </h3>
                    ''', unsafe_allow_html=True)
                    
                    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                    fig5 = plot_risk_sunburst(df_emphasis)
                    st.plotly_chart(fig5, use_container_width=True)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù…ÙˆØ¯Ø§Ø± Ø³Ù„Ø³Ù„Ù‡ Ù…Ø±Ø§ØªØ¨ÛŒ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§"):
                        st.markdown("""
                        <div style="text-align: right; direction: rtl; padding: 1rem;">
                        
                        **Ø§ÛŒÙ† Ù†Ù…ÙˆØ¯Ø§Ø± Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                        
                        - **Ù…Ø±Ú©Ø² Ø¯Ø§ÛŒØ±Ù‡**: Ú©Ù„ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡ Ø´Ø¯Ù‡
                        - **Ø­Ù„Ù‚Ù‡ Ø¯Ø§Ø®Ù„ÛŒ**: Ø¯Ø³ØªÙ‡â€ŒÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ Ø±ÛŒØ³Ú©
                        - **Ø­Ù„Ù‚Ù‡ Ø®Ø§Ø±Ø¬ÛŒ**: Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§ÛŒ Ù‡Ø± Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ
                        
                        **Ù†Ø­ÙˆÙ‡ Ø§Ø³ØªÙØ§Ø¯Ù‡:**
                        - Ø§Ù†Ø¯Ø§Ø²Ù‡ Ù‡Ø± Ù‚Ø·Ø§Ø¹ Ù…ØªÙ†Ø§Ø³Ø¨ Ø¨Ø§ ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ§Ø±Ø¯ Ø¢Ù† Ø§Ø³Øª
                        - Ú©Ù„ÛŒÚ© Ø±ÙˆÛŒ Ù‡Ø± Ù‚Ø³Ù…Øª Ø¨Ø±Ø§ÛŒ Ø¨Ø²Ø±Ú¯â€ŒÙ†Ù…Ø§ÛŒÛŒ
                        - Hover Ø¨Ø±Ø§ÛŒ Ù…Ø´Ø§Ù‡Ø¯Ù‡ Ø¬Ø²Ø¦ÛŒØ§Øª (Ù†Ø§Ù…ØŒ ØªØ¹Ø¯Ø§Ø¯ØŒ Ø¯Ø±ØµØ¯)
                        
                        **Ù…Ø²Ø§ÛŒØ§:**
                        - Ø¯ÛŒØ¯ Ú©Ù„ÛŒ Ø§Ø² ØªÙˆØ²ÛŒØ¹ Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ Ø¯Ø± ÛŒÚ© Ù†Ú¯Ø§Ù‡
                        - Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø³Ø±ÛŒØ¹ Ù¾Ø±ØªÚ©Ø±Ø§Ø±ØªØ±ÛŒÙ† Ø¯Ø³ØªÙ‡â€ŒÙ‡Ø§ÛŒ Ø±ÛŒØ³Ú©
                        - Ø¯Ø±Ú© Ø±Ø§Ø¨Ø·Ù‡ Ø¨ÛŒÙ† Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ùˆ Ø²ÛŒØ±Ø´Ø§Ø®Ù‡â€ŒÙ‡Ø§
                        
                        </div>
                        """, unsafe_allow_html=True)
            else:
                st.info("ğŸ“Œ Ø±ÛŒØ³Ú© Ø¨Ø±Ø¬Ø³ØªÙ‡â€ŒØ§ÛŒ Ø¯Ø± Ø§ÛŒÙ† Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ Ø°Ú©Ø± Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")

            # ====================================================================
            # Ø¨Ø®Ø´ Û´: ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ
            # ====================================================================
            st.markdown('''
                <h2 style="background: linear-gradient(135deg, #30cfd0 0%, #330867 100%); 
                        padding: 2rem; 
                        border-radius: 20px; 
                        margin: 2.5rem 0; 
                        box-shadow: 0 10px 30px rgba(48, 207, 208, 0.3);
                        color: white; 
                        text-align: center; 
                        font-size: 1.8rem; 
                        font-weight: 700; 
                        letter-spacing: 1px;">
                    âš–ï¸ Ø¨Ø®Ø´ Û³: ØªØ­Ù„ÛŒÙ„ ØªØ®Ù„ÙØ§Øª Ùˆ Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ
                </h2>
            ''', unsafe_allow_html=True)
            
            if df_violations is not None and not df_violations.empty and df_violations['Ù…ÙˆØ¶ÙˆØ¹ÛŒØª_Ø¯Ø§Ø±Ø¯'].any():
                col5, col6 = st.columns([1, 1], gap="large")
                
                # Ù†Ù…ÙˆØ¯Ø§Ø± Û¶: Ø³ØªÙˆÙ†ÛŒ Ø§Ù†Ø¨Ø§Ø´ØªÙ‡ ØªØ®Ù„ÙØ§Øª
                with col5:
                    st.markdown('''
                        <h3 style="background: rgba(252, 74, 26, 0.3); 
                                padding: 0.8rem 1.5rem; 
                                border-radius: 12px; 
                                margin: 0.5rem 0 0 0;
                                color: #2c3e50; 
                                text-align: center; 
                                font-size: 1.2rem; 
                                font-weight: 600;">
                            ğŸ“‰ Ø±ÙˆÙ†Ø¯ ØªØ¹Ø¯Ø§Ø¯ Ùˆ ØªØ±Ú©ÛŒØ¨ ØªØ®Ù„ÙØ§Øª
                        </h3>
                    ''', unsafe_allow_html=True)
                    
                    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                    fig6 = plot_violations_stacked_bar(df_violations, font)
                    st.pyplot(fig6)
                    plt.close(fig6)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù…ÙˆØ¯Ø§Ø± ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ"):
                        st.markdown("""
                        <div style="text-align: right; direction: rtl; padding: 1rem;">
                                
                        **Ø§ÛŒÙ† Ù†Ù…ÙˆØ¯Ø§Ø± Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                        
                        - **Ø³ØªÙˆÙ†â€ŒÙ‡Ø§**: ØªØ¹Ø¯Ø§Ø¯ Ú©Ù„ Ù…ÙˆØ§Ø±Ø¯ Ø¹Ø¯Ù… Ø±Ø¹Ø§ÛŒØª Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ
                        - **Ø±Ù†Ú¯â€ŒÙ‡Ø§**: Ø¯Ø³ØªÙ‡â€ŒØ¨Ù†Ø¯ÛŒ ØªØ®Ù„ÙØ§Øª (Ù†Ù‡Ø§Ø¯ Ù†Ø§Ø¸Ø±ØŒ Ø¨Ø§Ø²Ø§Ø± Ø³Ø±Ù…Ø§ÛŒÙ‡ØŒ Ø­Ø§Ú©Ù…ÛŒØªÛŒØŒ Ù…Ø§Ù„ÛŒØ§ØªÛŒ Ùˆ...)
                        - **Ø§Ø¹Ø¯Ø§Ø¯**: ØªØ¹Ø¯Ø§Ø¯ Ù…ÙˆØ§Ø±Ø¯ ØªØ®Ù„Ù Ø¯Ø± Ù‡Ø± Ø¯Ø³ØªÙ‡
                        
                        **Ø§Ù‡Ù…ÛŒØª:**
                        - Ø§ÙØ²Ø§ÛŒØ´ ØªØ®Ù„ÙØ§Øª = Ø¶Ø¹Ù Ø¯Ø± Ø±Ø¹Ø§ÛŒØª Ø§Ù„Ø²Ø§Ù…Ø§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ
                        - Ú©Ø§Ù‡Ø´ ØªØ®Ù„ÙØ§Øª = Ø¨Ù‡Ø¨ÙˆØ¯ Ø³ÛŒØ³ØªÙ…â€ŒÙ‡Ø§ÛŒ Ú©Ù†ØªØ±Ù„ÛŒ Ùˆ Ø­Ø§Ú©Ù…ÛŒØªÛŒ
                        - ØªÚ©Ø±Ø§Ø± ØªØ®Ù„Ù Ø¯Ø± ÛŒÚ© Ø¯Ø³ØªÙ‡ = Ù†ÛŒØ§Ø² Ø¨Ù‡ Ø§Ù‚Ø¯Ø§Ù… Ø§ØµÙ„Ø§Ø­ÛŒ Ø§Ø³Ø§Ø³ÛŒ
                        
                        **ØªÙˆØ¬Ù‡:** ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ù…ÛŒâ€ŒØªÙˆØ§Ù†Ø¯ Ù…Ù†Ø¬Ø± Ø¨Ù‡ Ø¬Ø±ÛŒÙ…Ù‡â€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ Ùˆ Ø¢Ø³ÛŒØ¨ Ø¨Ù‡ Ø´Ù‡Ø±Øª Ø³Ø§Ø²Ù…Ø§Ù† Ø´ÙˆØ¯.
                                    
                        </div>
                        """,unsafe_allow_html=True)
                
                # Ù†Ù…ÙˆØ¯Ø§Ø± Û·: Sunburst ØªØ®Ù„ÙØ§Øª
                with col6:
                    st.markdown('''
                        <h3 style="background: rgba(238, 156, 167, 0.3); 
                                padding: 0.8rem 1.5rem; 
                                border-radius: 12px; 
                                margin: 0.5rem 0 0 0;
                                color: #2c3e50; 
                                text-align: center; 
                                font-size: 1.2rem; 
                                font-weight: 600;">
                            ğŸ” Ø³Ø§Ø®ØªØ§Ø± Ùˆ ØªÙˆØ²ÛŒØ¹ ØªØ®Ù„ÙØ§Øª
                        </h3>
                    ''', unsafe_allow_html=True)
                    
                    st.markdown('<div class="chart-container">', unsafe_allow_html=True)
                    fig7 = plot_violations_sunburst(df_violations)
                    st.plotly_chart(fig7, use_container_width=True)
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù…ÙˆØ¯Ø§Ø± Ø³Ù„Ø³Ù„Ù‡ Ù…Ø±Ø§ØªØ¨ÛŒ ØªØ®Ù„ÙØ§Øª"):
                        st.markdown("""
                            <div style="text-align: right; direction: rtl; padding: 1rem;">
                                    
                        **Ø§ÛŒÙ† Ù†Ù…ÙˆØ¯Ø§Ø± Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                        
                        - **Ø­Ù„Ù‚Ù‡ Ù…Ø±Ú©Ø²ÛŒ**: Ú©Ù„ ØªØ®Ù„ÙØ§Øª
                        - **Ø­Ù„Ù‚Ù‡ Ù…ÛŒØ§Ù†ÛŒ**: Ø¯Ø³ØªÙ‡â€ŒÙ‡Ø§ÛŒ Ø§ØµÙ„ÛŒ Ù‚ÙˆØ§Ù†ÛŒÙ† (Ø§Ù„Ø²Ø§Ù…Ø§Øª Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒØŒ Ø¨ÙˆØ±Ø³ØŒ Ø­Ø§Ú©Ù…ÛŒØªÛŒ Ùˆ...)
                        - **Ø­Ù„Ù‚Ù‡ Ø®Ø§Ø±Ø¬ÛŒ**: Ø¬Ø²Ø¦ÛŒØ§Øª ØªØ®Ù„Ù Ø¯Ø± Ù‡Ø± Ø¯Ø³ØªÙ‡
                        
                        **Ú©Ø§Ø±Ø¨Ø±Ø¯Ù‡Ø§:**
                        - Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ù¾Ø±ØªÚ©Ø±Ø§Ø±ØªØ±ÛŒÙ† Ù†ÙˆØ¹ ØªØ®Ù„Ù
                        - Ø¯Ø±Ú© Ø±Ø§Ø¨Ø·Ù‡ Ø¨ÛŒÙ† Ø¯Ø³ØªÙ‡ Ø§ØµÙ„ÛŒ Ù‚Ø§Ù†ÙˆÙ† Ùˆ Ù…ÙˆØ§Ø±Ø¯ ØªØ®Ù„Ù
                        - Ø§ÙˆÙ„ÙˆÛŒØªâ€ŒØ¨Ù†Ø¯ÛŒ Ø¨Ø±Ø§ÛŒ Ø§Ù‚Ø¯Ø§Ù…Ø§Øª Ø§ØµÙ„Ø§Ø­ÛŒ
                        
                        **Ù†Ú©ØªÙ‡ Ù…Ù‡Ù…:** 
                        ØªÙ…Ø±Ú©Ø² ØªØ®Ù„ÙØ§Øª Ø¯Ø± ÛŒÚ© Ø¯Ø³ØªÙ‡ Ø®Ø§Øµ (Ù…Ø«Ù„Ø§Ù‹ Ø§Ù„Ø²Ø§Ù…Ø§Øª Ø¨Ø§Ù†Ú© Ù…Ø±Ú©Ø²ÛŒ) 
                        Ù†Ø´Ø§Ù†â€ŒØ¯Ù‡Ù†Ø¯Ù‡ Ù†ÛŒØ§Ø² Ø¨Ù‡ ØªÙˆØ¬Ù‡ ÙˆÛŒÚ˜Ù‡ Ø¨Ù‡ Ø¢Ù† Ø­ÙˆØ²Ù‡ Ø§Ø³Øª.
                                    
                        </div>
                        """,unsafe_allow_html=True)
            else:
                st.info("âœ… Ø¯Ø± Ø§ÛŒÙ† Ú¯Ø²Ø§Ø±Ø´â€ŒÙ‡Ø§ØŒ ØªØ®Ù„Ù Ù‚Ø§Ù†ÙˆÙ†ÛŒ Ú¯Ø²Ø§Ø±Ø´ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")


           # ====================================================================
            # Ø¨Ø®Ø´ 4: Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ
            # ====================================================================
            st.markdown('''
                <h2 style="background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%); 
                        padding: 2rem; 
                        border-radius: 20px; 
                        margin: 2.5rem 0; 
                        box-shadow: 0 10px 30px rgba(168, 237, 234, 0.3);
                        color: #2c3e50; 
                        text-align: center; 
                        font-size: 1.8rem; 
                        font-weight: 700; 
                        letter-spacing: 1px;">
                    ğŸ”¥ Ø¨Ø®Ø´ Û´: Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ
                </h2>
            ''', unsafe_allow_html=True)
            
            if df_checklist is not None and not df_checklist.empty:
                st.markdown('''
                    <h3 style="background: rgba(255, 154, 158, 0.3); 
                            padding: 0.8rem 1.5rem; 
                            border-radius: 12px; 
                            margin: 0.5rem 0 0 0;
                            color: #2c3e50; 
                            text-align: center; 
                            font-size: 1.2rem; 
                            font-weight: 600;">
                        ğŸ¯ Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ ÙˆØ¶Ø¹ÛŒØª Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ
                    </h3>
                ''', unsafe_allow_html=True)
                
                st.markdown('<div class="chart-container-full">', unsafe_allow_html=True)
                fig3 = plot_checklist_heatmap(df_checklist, font)
                st.pyplot(fig3)
                plt.close(fig3)
                st.markdown('</div>', unsafe_allow_html=True)
                
                with st.expander("ğŸ“– ØªÙˆØ¶ÛŒØ­Ø§Øª Ù†Ù‚Ø´Ù‡ Ø­Ø±Ø§Ø±ØªÛŒ"):
                    st.markdown("""
                    <div style="text-align: right; direction: rtl; padding: 1rem;">
                    
                    **Ø§ÛŒÙ† Ù†Ù‚Ø´Ù‡ Ú†Ù‡ Ú†ÛŒØ²ÛŒ Ø±Ø§ Ù†Ø´Ø§Ù† Ù…ÛŒâ€ŒØ¯Ù‡Ø¯ØŸ**
                    
                    - **Ø±Ø¯ÛŒÙâ€ŒÙ‡Ø§**: Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ Ù…ÙˆØ±Ø¯ Ø¨Ø±Ø±Ø³ÛŒ (Ú©ÙØ§ÛŒØª Ø³Ø±Ù…Ø§ÛŒÙ‡ØŒ ØªØ³Ø¹ÛŒØ± Ø§Ø±Ø²ØŒ Ù…Ø§Ù„ÛŒØ§Øª Ùˆ...)
                    - **Ø³ØªÙˆÙ†â€ŒÙ‡Ø§**: Ø³Ø§Ù„â€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ
                    - **Ø±Ù†Ú¯ Ø³Ù„ÙˆÙ„â€ŒÙ‡Ø§**: Ø³Ø·Ø­ Ø±ÛŒØ³Ú© ÛŒØ§ ÙˆØ¶Ø¹ÛŒØª Ù‡Ø± Ù…ÙˆØ¶ÙˆØ¹
        
                    **Ú©Ø§Ø±Ø¨Ø±Ø¯:**
                    - Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ù…ÙˆØ¶ÙˆØ¹Ø§Øª ØªÚ©Ø±Ø§Ø±ÛŒ Ø¯Ø± Ø·ÙˆÙ„ Ø²Ù…Ø§Ù†
                    - ØªØ´Ø®ÛŒØµ Ø±ÙˆÙ†Ø¯ Ø¨Ù‡Ø¨ÙˆØ¯ ÛŒØ§ Ø¨Ø¯ØªØ± Ø´Ø¯Ù† Ù‡Ø± Ù…ÙˆØ¶ÙˆØ¹
                    - Ø§ÙˆÙ„ÙˆÛŒØªâ€ŒØ¨Ù†Ø¯ÛŒ Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ù¾Ø±Ø±ÛŒØ³Ú©
                    
                    </div>
                    """, unsafe_allow_html=True)
                
        except Exception as e:
            st.error(f"âŒ Ø®Ø·Ø§ Ø¯Ø± Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§: {str(e)}")
            logger.error(f"Chart error: {traceback.format_exc()}")
            
        st.markdown("</div>", unsafe_allow_html=True)

    def normalize_company_name(name: str) -> str:
        """
        Ù†Ø±Ù…Ø§Ù„â€ŒØ³Ø§Ø²ÛŒ Ø®ÛŒÙ„ÛŒ Ø¶Ø¹ÛŒÙ: ÙÙ‚Ø· Ø¨Ø±Ø§ÛŒ Ø¬Ù„ÙˆÚ¯ÛŒØ±ÛŒ Ø§Ø² Ø®Ø·Ø§Ù‡Ø§ÛŒ ØªØ§ÛŒÙ¾ÛŒ
        Ø¨ÛŒØ´ØªØ± Ù†Ø§Ù…â€ŒÙ‡Ø§ Ù…ØªÙØ§ÙˆØª Ø¨Ø§Ù‚ÛŒ Ù…ÛŒâ€ŒÙ…Ø§Ù†Ù†Ø¯ ØªØ§ Ú©Ø§Ø±Ø¨Ø± Ø®ÙˆØ¯Ø´ ØªØµÙ…ÛŒÙ… Ø¨Ú¯ÛŒØ±Ø¯
        """
        if not name:
            return ""
        
        # ÙÙ‚Ø· Ø­Ø°Ù ÙØ¶Ø§Ù‡Ø§ÛŒ Ø§Ø¶Ø§ÙÛŒ Ùˆ ÛŒÚ©Ø³Ø§Ù†â€ŒØ³Ø§Ø²ÛŒ Ø­Ø±ÙˆÙ Ø¹Ø±Ø¨ÛŒ/ÙØ§Ø±Ø³ÛŒ
        name = name.strip()
        name = name.replace("ÙŠ", "ÛŒ").replace("Ùƒ", "Ú©")
        name = name.replace("Ø©", "Ù‡").replace("Ø¤", "Ùˆ").replace("Ø¥", "Ø§").replace("Ø£", "Ø§")
        name = re.sub(r'\s+', ' ', name)
        
        return name


    def analyze_companies(results: List) -> Dict:
        """
        ØªØ­Ù„ÛŒÙ„ Ùˆ Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ù…ÙˆØ¬ÙˆØ¯ Ø¯Ø± Ù†ØªØ§ÛŒØ¬
        Ø¨Ø¯ÙˆÙ† Ø§Ø¯ØºØ§Ù… Ø®ÙˆØ¯Ú©Ø§Ø± - Ú©Ø§Ø±Ø¨Ø± Ø¨Ø§ÛŒØ¯ Ø®ÙˆØ¯Ø´ ØªØµÙ…ÛŒÙ… Ø¨Ú¯ÛŒØ±Ø¯
        """
        companies_info = {}
        
        for filename, data in results:
            if "error" in data:
                continue
                
            try:
                report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                summary = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                
                company_name = summary.get("Ù†Ø§Ù…_Ø´Ø±Ú©Øª", f"Unknown_{filename}")
                normalized = normalize_company_name(company_name)
                
                # Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø³Ø§Ù„
                financial_year = summary.get("Ø¯ÙˆØ±Ù‡_Ù…Ø§Ù„ÛŒ", "")
                year_match = re.search(r'(\d{4})', financial_year)
                year = int(year_match.group(1)) if year_match else None
                
                # Ø°Ø®ÛŒØ±Ù‡ Ø§Ø·Ù„Ø§Ø¹Ø§Øª
                if normalized not in companies_info:
                    companies_info[normalized] = {
                        'original_name': company_name,
                        'normalized_name': normalized,
                        'files': [],
                        'years': [],
                        'risk_levels': [],
                        'opinion_types': []
                    }
                
                companies_info[normalized]['files'].append(filename)
                if year:
                    companies_info[normalized]['years'].append(year)
                companies_info[normalized]['risk_levels'].append(
                    summary.get("Ø³Ø·Ø­_Ø±ÛŒØ³Ú©_Ú©Ù„ÛŒ_Ø¨Ù†Ø§_Ø¨Ù‡_Ú¯Ø²Ø§Ø±Ø´", "Ù†Ø§Ù…Ø´Ø®Øµ")
                )
                companies_info[normalized]['opinion_types'].append(
                    summary.get("Ù†ÙˆØ¹_Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø±", "Ù†Ø§Ù…Ø´Ø®Øµ")
                )
                
            except Exception as e:
                logger.error(f"Ø®Ø·Ø§ Ø¯Ø± ØªØ­Ù„ÛŒÙ„ {filename}: {e}")
                continue
        
        # Ù…Ø±ØªØ¨â€ŒØ³Ø§Ø²ÛŒ Ø³Ø§Ù„â€ŒÙ‡Ø§
        for info in companies_info.values():
            info['years'] = sorted(info['years'])
            info['year_range'] = f"{min(info['years'])} - {max(info['years'])}" if info['years'] else "Ù†Ø§Ù…Ø´Ø®Øµ"
            info['file_count'] = len(info['files'])
        
        return {
            'companies': companies_info,
            'total_companies': len(companies_info),
            'is_single_company': len(companies_info) == 1,
            'company_names': [info['original_name'] for info in companies_info.values()]
        }


    def create_company_merger_interface(analysis: Dict) -> Dict[str, List[str]]:
        """
        Ø±Ø§Ø¨Ø· Ú©Ø§Ø±Ø¨Ø±ÛŒ Ø¨Ø±Ø§ÛŒ Ø§Ø¯ØºØ§Ù… Ø¯Ø³ØªÛŒ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§
        """
        st.subheader("Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡")
        
        # Ù…Ù‚Ø¯Ø§Ø±Ø¯Ù‡ÛŒ session state
        if 'company_mergers' not in st.session_state:
            st.session_state.company_mergers = {}
        
        companies = list(analysis['companies'].keys())
        
        if len(companies) < 2:
            st.info("ğŸ“Œ ÙÙ‚Ø· ÛŒÚ© Ø´Ø±Ú©Øª Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡ Ø§Ø³Øª.")
            return {}
        
        
        
        # Ù†Ù…Ø§ÛŒØ´ ÙØ´Ø±Ø¯Ù‡ Ù„ÛŒØ³Øª Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§
        for normalized_name, info in sorted(
            analysis['companies'].items(), 
            key=lambda x: x[1]['file_count'], 
            reverse=True
        ):
            st.markdown(f"""
                <div style="background: white; 
                            padding: 0.8rem; 
                            border-radius: 8px; 
                            border-right: 3px solid #2196f3;
                            margin-bottom: 0.5rem;
                            box-shadow: 0 1px 3px rgba(0,0,0,0.1);">
                    <span style="font-weight: bold; color: #2c3e50;">
                        {info['original_name']}
                    </span>
                    <span style="background: #e3f2fd; 
                                padding: 0.2rem 0.5rem; 
                                border-radius: 10px; 
                                font-size: 0.8rem;
                                margin-right: 0.5rem;">
                        ğŸ“… {info['year_range']}
                    </span>
                    <span style="background: #f3e5f5; 
                                padding: 0.2rem 0.5rem; 
                                border-radius: 10px; 
                                font-size: 0.8rem;">
                        ğŸ“Š {info['file_count']} Ø³Ø§Ù„
                    </span>
                </div>
            """, unsafe_allow_html=True)
        
        # ÙØ±Ù… Ø§ÛŒØ¬Ø§Ø¯ Ú¯Ø±ÙˆÙ‡ Ø¬Ø¯ÛŒØ¯
        st.markdown("### â• Ø§ÛŒØ¬Ø§Ø¯ Ú¯Ø±ÙˆÙ‡ Ø§Ø¯ØºØ§Ù… Ø¬Ø¯ÛŒØ¯ Ø¨Ø±Ø§ÛŒ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ ÛŒÚ©Ø³Ø§Ù†")
        with st.expander("â„¹ï¸ Ø±Ø§Ù‡Ù†Ù…Ø§: Ú†Ú¯ÙˆÙ†Ù‡ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ Ø±Ø§ Ø§Ø¯ØºØ§Ù… Ú©Ù†ÛŒÙ…ØŸ", expanded=False):
            st.markdown("""
            <div style="text-align: right; direction: rtl; padding: 1rem; border-radius: 8px;">
            
            **ğŸ“Œ Ù‡Ø¯Ù Ø§Ø² Ø§Ø¯ØºØ§Ù…:**
            
            Ø§Ú¯Ø± ØµÙˆØ±Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ ÛŒÚ© Ø´Ø±Ú©Øª Ø¨Ø§ Ù†Ø§Ù…â€ŒÙ‡Ø§ÛŒ Ù…Ø®ØªÙ„Ù Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡â€ŒØ§Ù†Ø¯ØŒ Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒØ¯ Ø¢Ù†â€ŒÙ‡Ø§ Ø±Ø§ Ø¯Ø± ÛŒÚ© Ú¯Ø±ÙˆÙ‡ Ù‚Ø±Ø§Ø± Ø¯Ù‡ÛŒØ¯.
            
            **Ù…Ø«Ø§Ù„â€ŒÙ‡Ø§ÛŒ Ù†ÛŒØ§Ø² Ø¨Ù‡ Ø§Ø¯ØºØ§Ù…:**
            
            âœ… **ØªØºÛŒÛŒØ± Ø¬Ø²Ø¦ÛŒ Ù†Ø§Ù…:**
            - "Ø¨Ø§Ù†Ú© Ù…Ù„Øª" Ùˆ "Ø¨Ø§Ù†Ú© Ù…Ù„Øª (Ø³Ù‡Ø§Ù…ÛŒ Ø¹Ø§Ù…)"
            - "ØµÙ†Ø§ÛŒØ¹ Ù¾ØªØ±ÙˆØ´ÛŒÙ…ÛŒ Ø®Ù„ÛŒØ¬ ÙØ§Ø±Ø³" Ùˆ "Ø´Ø±Ú©Øª ØµÙ†Ø§ÛŒØ¹ Ù¾ØªØ±ÙˆØ´ÛŒÙ…ÛŒ Ø®Ù„ÛŒØ¬ ÙØ§Ø±Ø³"
            
            âœ… **ØªØºÛŒÛŒØ± Ù†Ø§Ù… ØªØ¬Ø§Ø±ÛŒ:**
            - "Ù¾Ø§Ø±Ø³ Ø®ÙˆØ¯Ø±Ùˆ" â†’ "Ø§ÛŒØ±Ø§Ù† Ø®ÙˆØ¯Ø±Ùˆ"
            - "Ú¯Ø±ÙˆÙ‡ ØµÙ†Ø¹ØªÛŒ Ø§ÛŒØ±Ø§Ù†" â†’ "Ù‡Ù„Ø¯ÛŒÙ†Ú¯ ØµÙ†Ø¹ØªÛŒ Ø§ÛŒØ±Ø§Ù†"
            
            âœ… **Ø§Ø¯ØºØ§Ù… ÛŒØ§ ØªØºÛŒÛŒØ± Ù†Ø§Ù… Ù‚Ø§Ù†ÙˆÙ†ÛŒ:**
            - "Ø¨Ø§Ù†Ú© Ø§Ù‚ØªØµØ§Ø¯ Ù†ÙˆÛŒÙ†" + "Ù…ÙˆØ³Ø³Ù‡ Ø§Ø¹ØªØ¨Ø§Ø±ÛŒ Ù†ÙˆÛŒÙ†"
            
            ---
            
            **ğŸ”§ Ù†Ø­ÙˆÙ‡ Ø§Ø³ØªÙØ§Ø¯Ù‡:**
            
            1. Ø§Ø² Ù„ÛŒØ³Øª Ù¾Ø§ÛŒÛŒÙ†ØŒ ÛŒÚ© Ø´Ø±Ú©Øª Ø±Ø§ Ø¨Ù‡ Ø¹Ù†ÙˆØ§Ù† **"Ù†Ø§Ù… Ø§ØµÙ„ÛŒ"** Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯
            2. Ø³Ø§ÛŒØ± Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒÛŒ Ú©Ù‡ Ø¨Ø§ÛŒØ¯ Ø§Ø¯ØºØ§Ù… Ø´ÙˆÙ†Ø¯ Ø±Ø§ Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯
            3. Ø±ÙˆÛŒ Ø¯Ú©Ù…Ù‡ **"Ø«Ø¨Øª Ú¯Ø±ÙˆÙ‡"** Ú©Ù„ÛŒÚ© Ú©Ù†ÛŒØ¯
            4. Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±ØŒ ÙÙ‚Ø· Ù†Ø§Ù… Ø§ØµÙ„ÛŒ Ø±Ø§ Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯
            
            </div>
            """, unsafe_allow_html=True)


        col1, col2 = st.columns(2)
        
        with col1:
            main_company = st.selectbox(
                "Ù†Ø§Ù… Ø§ØµÙ„ÛŒ Ø´Ø±Ú©Øª (Ù…Ø±Ø¬Ø¹):",
                options=companies,
                format_func=lambda x: f"{analysis['companies'][x]['original_name']} ({analysis['companies'][x]['file_count']} Ø³Ø§Ù„)",
                key="main_company_select"
            )
        
        with col2:
            # ÙÛŒÙ„ØªØ± Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒÛŒ Ú©Ù‡ Ù‚Ø¨Ù„Ø§Ù‹ Ø¯Ø± Ú¯Ø±ÙˆÙ‡ Ù†ÛŒØ³ØªÙ†Ø¯
            available_companies = [
                c for c in companies 
                if c not in st.session_state.company_mergers.keys() and c != main_company
            ]
            
            related_companies = st.multiselect(
                "Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø±ØªØ¨Ø· (Ù‡Ù…Ø§Ù† Ø´Ø±Ú©Øª Ø¨Ø§ Ù†Ø§Ù… Ø¯ÛŒÚ¯Ø±):",
                options=available_companies,
                format_func=lambda x: f"{analysis['companies'][x]['original_name']} ({analysis['companies'][x]['file_count']} Ø³Ø§Ù„)",
                key="related_companies_select"
            )
        
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
        
        with col_btn1:
            if st.button("âœ… Ø«Ø¨Øª Ú¯Ø±ÙˆÙ‡", type="primary", use_container_width=True):
                if main_company:
                    st.session_state.company_mergers[main_company] = [main_company] + related_companies
                    
                    total_files = sum([
                        analysis['companies'][c]['file_count'] 
                        for c in [main_company] + related_companies
                    ])
                    # st.success(f"âœ… Ú¯Ø±ÙˆÙ‡ Ø«Ø¨Øª Ø´Ø¯: {analysis['companies'][main_company]['original_name']} (Ù…Ø¬Ù…ÙˆØ¹ {total_files} ÙØ§ÛŒÙ„)")
                    st.rerun()
        
        with col_btn2:
            if st.button("ğŸ”„ Ù¾Ø§Ú© Ú©Ø±Ø¯Ù† Ù‡Ù…Ù‡", use_container_width=True):
                st.session_state.company_mergers = {}
                st.rerun()
                
        # --- Ø¬Ø§Ø¨Ø¬Ø§ÛŒÛŒ Ù…Ø­Ù„ Ù†Ù…Ø§ÛŒØ´: Ù„ÛŒØ³Øª Ú¯Ø±ÙˆÙ‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¯ØºØ§Ù… Ø´Ø¯Ù‡ Ø§ÛŒÙ†Ø¬Ø§ Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ù…ÛŒâ€ŒØ´ÙˆØ¯ ---
        if st.session_state.company_mergers:
            # st.markdown("### ğŸ“Œ Ú¯Ø±ÙˆÙ‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¯ØºØ§Ù… Ø´Ø¯Ù‡:")
            
            for main_name, merged_list in st.session_state.company_mergers.items():
                main_original = analysis['companies'][main_name]['original_name']
                merged_originals = [
                    analysis['companies'][n]['original_name'] 
                    for n in merged_list if n != main_name
                ]
                
                total_files = sum([
                    analysis['companies'][n]['file_count'] 
                    for n in merged_list
                ])
                
                # col_res1, col_res2 = st.columns([5, 1])
                # with col_res1:
                st.markdown(f"""
                    <div style="background: #e8f5e9; 
                                padding: 1rem; 
                                border-radius: 10px; 
                                margin-bottom: 0.5rem;
                                border-right: 5px solid #4caf50;">
                        <strong>ğŸ¢ Ù†Ø§Ù… Ø§ØµÙ„ÛŒ:</strong> {main_original} 
                        <span style="background: #4caf50; color: white; padding: 0.2rem 0.6rem; 
                                    border-radius: 12px; font-size: 0.85rem; margin-right: 0.5rem;">
                            {total_files} ÙØ§ÛŒÙ„
                        </span>
                        <br>
                        {'<strong>ğŸ”— Ø´Ø§Ù…Ù„:</strong> ' + ', '.join(merged_originals) if merged_originals else ''}
                    </div>
                """, unsafe_allow_html=True)
                
                # with col_res2:
                #     if st.button("ğŸ—‘ï¸", key=f"delete_{main_name}", use_container_width=True):
                #         del st.session_state.company_mergers[main_name]
                #         st.rerun()
        
        return st.session_state.company_mergers


    def apply_company_mergers(results: List, mergers: Dict[str, List[str]]) -> Dict[str, List]:
        """
        Ø§Ø¹Ù…Ø§Ù„ Ø§Ø¯ØºØ§Ù…â€ŒÙ‡Ø§ÛŒ Ø´Ø±Ú©Øª Ø±ÙˆÛŒ Ù†ØªØ§ÛŒØ¬
        """
        
        if not mergers:
            return {}
        
        # Ø§ÛŒØ¬Ø§Ø¯ Ù†Ú¯Ø§Ø´Øª Ù…Ø¹Ú©ÙˆØ³
        reverse_map = {}
        for main_name, merged_list in mergers.items():
            for name in merged_list:
                reverse_map[name] = main_name
        
        # Ú¯Ø±ÙˆÙ‡â€ŒØ¨Ù†Ø¯ÛŒ Ù†ØªØ§ÛŒØ¬
        grouped_results = {main_name: [] for main_name in mergers.keys()}
        
        for filename, data in results:
            if "error" in data:
                continue
            
            try:
                report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                summary = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                company_name = summary.get("Ù†Ø§Ù…_Ø´Ø±Ú©Øª", "")
                normalized = normalize_company_name(company_name)
                
                if normalized in reverse_map:
                    main_group = reverse_map[normalized]
                    grouped_results[main_group].append((filename, data))
            
            except Exception as e:
                logger.error(f"Ø®Ø·Ø§ Ø¯Ø± Ú¯Ø±ÙˆÙ‡â€ŒØ¨Ù†Ø¯ÛŒ {filename}: {e}")
                continue
        
        return grouped_results



    def create_company_selector_with_merger(results: List) -> Tuple[str, Dict, List]:
        """
        Ø±Ø§Ø¨Ø· Ú©Ø§Ø±Ø¨Ø±ÛŒ Ø¨Ø±Ø§ÛŒ Ø§Ù†ØªØ®Ø§Ø¨ Ø´Ø±Ú©Øª Ø¨Ø§ Ù‚Ø§Ø¨Ù„ÛŒØª Ø§Ø¯ØºØ§Ù…
        """
        # ØªØ­Ù„ÛŒÙ„ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§
        analysis = analyze_companies(results)
        
        # Ø±Ø§Ø¨Ø· Ø§Ø¯ØºØ§Ù… Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§
        mergers = create_company_merger_interface(analysis)
        
        # st.markdown("---")
        
        # Ø§Ø¹Ù…Ø§Ù„ Ø§Ø¯ØºØ§Ù…â€ŒÙ‡Ø§
        if mergers:
            # Ù†ØªØ§ÛŒØ¬ Ø§Ø¯ØºØ§Ù… Ø´Ø¯Ù‡
            grouped_results = apply_company_mergers(results, mergers)
            
            # 1) Ù†Ø§Ù…â€ŒÙ‡Ø§ÛŒ Ø´Ø±Ú©Øª Ø§ØµÙ„ÛŒ Ø§Ø² Ú¯Ø±ÙˆÙ‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¯ØºØ§Ù… Ø´Ø¯Ù‡
            merged_main_names = list(mergers.keys())
            
            # 2) Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒÛŒ Ú©Ù‡ Ø¯Ø± Ù‡ÛŒÚ† Ú¯Ø±ÙˆÙ‡ Ø§Ø¯ØºØ§Ù…ÛŒ Ù†ÛŒØ³ØªÙ†Ø¯ (Ù…Ø³ØªÙ‚Ù„)
            merged_members = set()
            for main, members in mergers.items():
                for m in members:
                    merged_members.add(m)
            
            all_company_keys = list(analysis['companies'].keys())
            independent_companies = [c for c in all_company_keys if c not in merged_members]
            
            # 3) Ù„ÛŒØ³Øª Ù†Ù‡Ø§ÛŒÛŒ: Ú¯Ø±ÙˆÙ‡â€ŒÙ‡Ø§ÛŒ Ø§Ø¯ØºØ§Ù…ÛŒ + Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø³ØªÙ‚Ù„
            available_companies = []
            seen = set()
            
            for c in merged_main_names + independent_companies:
                if c not in seen:
                    available_companies.append(c)
                    seen.add(c)
            
            # ØªØ§Ø¨Ø¹ Ú©Ù…Ú©ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ ØªØ¹Ø¯Ø§Ø¯ ÙØ§ÛŒÙ„
            def file_count(x):
                if x in grouped_results:
                    return len(grouped_results[x])
                return analysis['companies'][x]['file_count']
            
            def company_label(x):
                if x in mergers:
                    return f"ğŸ”— {analysis['companies'][x]['original_name']} (Ú¯Ø±ÙˆÙ‡ Ø§Ø¯ØºØ§Ù… - {file_count(x)} ÙØ§ÛŒÙ„)"
                else:
                    return f"{analysis['companies'][x]['original_name']} ({file_count(x)} ÙØ§ÛŒÙ„)"
            
            st.subheader("ğŸ¢Ø§Ù†ØªØ®Ø§Ø¨ Ø´Ø±Ú©Øª Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±")
            
            # if len(available_companies) > 1:
            #     st.info(f"ğŸ“Š ØªØ¹Ø¯Ø§Ø¯ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ù‚Ø§Ø¨Ù„ Ø§Ù†ØªØ®Ø§Ø¨: {len(available_companies)} ({len(merged_main_names)} Ú¯Ø±ÙˆÙ‡ Ø§Ø¯ØºØ§Ù…ÛŒ + {len(independent_companies)} Ù…Ø³ØªÙ‚Ù„)")
            
            selected_company = st.selectbox(
                "Ø´Ø±Ú©Øª Ù…ÙˆØ±Ø¯ Ù†Ø¸Ø± Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±:",
                options=available_companies,
                format_func=company_label,
                key="selected_company_for_chart"
            )
            
            if selected_company:
                # Ø§Ú¯Ø± Ú¯Ø±ÙˆÙ‡ Ø§Ø¯ØºØ§Ù…ÛŒ Ø§Ø³Øª
                if selected_company in grouped_results:
                    # --- ØªØºÛŒÛŒØ±: Ù†Ù…Ø§ÛŒØ´ Ø¬Ø²Ø¦ÛŒØ§Øª Ú¯Ø±ÙˆÙ‡ Ø§Ø¯ØºØ§Ù…ÛŒ Ø¯Ø± Ø§ÛŒÙ†Ø¬Ø§ Ø­Ø°Ù Ø´Ø¯ ---
                    return selected_company, analysis, grouped_results[selected_company]
                
                # Ø´Ø±Ú©Øª Ù…Ø³ØªÙ‚Ù„ Ø§Ø³Øª
                else:
                    company_results = []
                    for filename, data in results:
                        if "error" not in data:
                            try:
                                report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                                summary = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                                comp_name = summary.get("Ù†Ø§Ù…_Ø´Ø±Ú©Øª", "")
                                
                                if normalize_company_name(comp_name) == selected_company:
                                    company_results.append((filename, data))
                            except:
                                continue
                    
                    # --- ØªØºÛŒÛŒØ±: Ù†Ù…Ø§ÛŒØ´ Ø¬Ø²Ø¦ÛŒØ§Øª Ø´Ø±Ú©Øª Ù…Ø³ØªÙ‚Ù„ Ø¯Ø± Ø§ÛŒÙ†Ø¬Ø§ Ø­Ø°Ù Ø´Ø¯ ---
                    return selected_company, analysis, company_results
        
        else:
            # Ø­Ø§Ù„Øª Ø¨Ø¯ÙˆÙ† Ø§Ø¯ØºØ§Ù… - Ù†Ù…Ø§ÛŒØ´ Ù‡Ù…Ù‡ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§
            st.subheader("ğŸ¢ Ø§Ù†ØªØ®Ø§Ø¨ Ø´Ø±Ú©Øª Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±")

            
            if analysis['is_single_company']:
                company_name = list(analysis['companies'].keys())[0]
                st.success(f"âœ… ÙÙ‚Ø· ÛŒÚ© Ø´Ø±Ú©Øª Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ø´Ø¯Ù‡: {analysis['companies'][company_name]['original_name']}")
                
                # ÙÛŒÙ„ØªØ± Ù†ØªØ§ÛŒØ¬
                company_results = []
                for filename, data in results:
                    if "error" not in data:
                        try:
                            report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                            summary = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                            comp_name = summary.get("Ù†Ø§Ù…_Ø´Ø±Ú©Øª", "")
                            
                            if normalize_company_name(comp_name) == company_name:
                                company_results.append((filename, data))
                        except:
                            continue
                
                return company_name, analysis, company_results
            
            else:
                companies = list(analysis['companies'].keys())
                
                selected_company = st.selectbox(
                    "ÛŒÚ© Ø´Ø±Ú©Øª Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø± Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯:",
                    options=companies,
                    format_func=lambda x: f"{analysis['companies'][x]['original_name']} ({analysis['companies'][x]['file_count']} ÙØ§ÛŒÙ„)",
                    key="selected_single_company"
                )
                
                if selected_company:
                    company_results = []
                    for filename, data in results:
                        if "error" not in data:
                            try:
                                report = data["ØªØ­Ù„ÛŒÙ„_Ø¬Ø§Ù…Ø¹_Ú¯Ø²Ø§Ø±Ø´_Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ"]
                                summary = report["Ø¨Ø®Ø´Û±_Ø®Ù„Ø§ØµÙ‡_Ùˆ_Ø§Ø·Ù„Ø§Ø¹Ø§Øª_Ú©Ù„ÛŒØ¯ÛŒ"]
                                comp_name = summary.get("Ù†Ø§Ù…_Ø´Ø±Ú©Øª", "")
                                
                                if normalize_company_name(comp_name) == selected_company:
                                    company_results.append((filename, data))
                            except:
                                continue
                    
                    return selected_company, analysis, company_results
                
                return None, analysis, []


    def create_filtered_charts_section(results: List):
        """
        Ø¨Ø®Ø´ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ø¨Ø§ Ù‚Ø§Ø¨Ù„ÛŒØª Ø§Ø¯ØºØ§Ù… Ùˆ Ø§Ù†ØªØ®Ø§Ø¨ Ø´Ø±Ú©Øª
        """
        
        if not results or not any('error' not in r for _, r in results):
            st.info("Ø¯Ø§Ø¯Ù‡ Ù…Ø¹ØªØ¨Ø±ÛŒ Ø¨Ø±Ø§ÛŒ Ù†Ù…Ø§ÛŒØ´ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ ÙˆØ¬ÙˆØ¯ Ù†Ø¯Ø§Ø±Ø¯.")
            return
        
        # Ú¯Ø§Ù… 1: Ø§Ù†ØªØ®Ø§Ø¨ Ø´Ø±Ú©Øª Ø¨Ø§ Ù‚Ø§Ø¨Ù„ÛŒØª Ø§Ø¯ØºØ§Ù…
        selected_company, analysis, company_results = create_company_selector_with_merger(results)
        
        if not selected_company or not company_results:
            st.warning("âš ï¸ Ù„Ø·ÙØ§Ù‹ ÛŒÚ© Ø´Ø±Ú©Øª Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯ ØªØ§ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ø±Ø³Ù… Ø´ÙˆÙ†Ø¯.")
            return
        
        # Ø¨Ø±Ø±Ø³ÛŒ ØªØ¹Ø¯Ø§Ø¯ Ø¯Ø§Ø¯Ù‡â€ŒÙ‡Ø§
        # if len(company_results) < 2:
        #     st.error(f"âŒ Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø± Ø±ÙˆÙ†Ø¯ØŒ Ø­Ø¯Ø§Ù‚Ù„ 2 Ø³Ø§Ù„ Ø¯Ø§Ø¯Ù‡ Ù†ÛŒØ§Ø² Ø§Ø³Øª. ØªØ¹Ø¯Ø§Ø¯ ÙØ¹Ù„ÛŒ: {len(company_results)}")
        #     return
        
        # Ú¯Ø§Ù… 2: Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§
        # st.markdown("---")
        st.success(f"âœ… Ø¯Ø± Ø­Ø§Ù„ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ø¨Ø±Ø§ÛŒ: **{analysis['companies'][selected_company]['original_name']}** ({len(company_results)} ÙØ§ÛŒÙ„)")
        
        try:
            # ÙØ±Ø§Ø®ÙˆØ§Ù†ÛŒ ØªØ§Ø¨Ø¹ Ø§ØµÙ„ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±
            create_charts_section(company_results)
            
        except Exception as e:
            st.error(f"âŒ Ø®Ø·Ø§ Ø¯Ø± Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§: {str(e)}")
            logger.error(f"Chart error: {traceback.format_exc()}")
    def main():
        # Ù…Ù‚Ø¯Ø§Ø±Ø¯Ù‡ÛŒ Ø§ÙˆÙ„ÛŒÙ‡ session_state Ø¯Ø± Ø§Ø¨ØªØ¯Ø§ÛŒ Ø¨Ø±Ù†Ø§Ù…Ù‡
        if 'results' not in st.session_state:
            st.session_state.results = None
        if 'processing_active' not in st.session_state:
            st.session_state.processing_active = False

        create_header()
        tab1, tab2, tab3, tab4 = st.tabs(["ğŸ“¤ Ø¢Ù¾Ù„ÙˆØ¯ Ùˆ Ù¾Ø±Ø¯Ø§Ø²Ø´", "ğŸ“ŠÙ†ØªØ§ÛŒØ¬ ØªØ­Ù„ÛŒÙ„", "ğŸ“ˆ Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ø¢Ù…Ø§Ø±ÛŒ", "ğŸ“‰ ØªØ±Ù†Ø¯ Ùˆ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§"])
        
        with tab1:
            with st.expander("ğŸ“‹ Ø±Ø§Ù‡Ù†Ù…Ø§ÛŒ Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ ÙØ§ÛŒÙ„", expanded=False):
                st.markdown("""
                <div style="text-align: right; direction: rtl; padding: 1rem; border-radius: 12px;">
                
                #### âœ¨ Ù…Ø±Ø§Ø­Ù„ Ú©Ø§Ø±:
                
                * ğŸ“„ **ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ÛŒ PDF Ú¯Ø²Ø§Ø±Ø´ Ø­Ø³Ø§Ø¨Ø±Ø³ÛŒ** Ø±Ø§ Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯
                * ğŸ“¦ ÛŒØ§ ÛŒÚ© **ÙØ§ÛŒÙ„ ZIP** Ø­Ø§ÙˆÛŒ Ú†Ù†Ø¯ÛŒÙ† PDF Ø¨Ø§Ø±Ú¯Ø°Ø§Ø±ÛŒ Ú©Ù†ÛŒØ¯
                * âœ… Ù¾Ø³ Ø§Ø² Ø§Ù†ØªØ®Ø§Ø¨ØŒ ÙØ§ÛŒÙ„â€ŒÙ‡Ø§ Ø±Ø§ Ø¨Ø±Ø±Ø³ÛŒ Ú©Ù†ÛŒØ¯
                * ğŸš€ Ø¯Ú©Ù…Ù‡ **"Ø´Ø±ÙˆØ¹ ØªØ­Ù„ÛŒÙ„"** Ø±Ø§ Ø¨Ø²Ù†ÛŒØ¯ ØªØ§ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ø¢ØºØ§Ø² Ø´ÙˆØ¯
                
                ---
                
                #### âš ï¸ Ù†Ú©Ø§Øª Ù…Ù‡Ù…:
                
                * **ÙØ±Ù…Øª Ù¾Ø´ØªÛŒØ¨Ø§Ù†ÛŒ Ø´Ø¯Ù‡:** ÙÙ‚Ø· PDF
                * **Ø­Ø¯Ø§Ú©Ø«Ø± Ø­Ø¬Ù… Ù‡Ø± ÙØ§ÛŒÙ„:** 50 Ù…Ú¯Ø§Ø¨Ø§ÛŒØª
                * **Ú©ÛŒÙÛŒØª Ù…Ù‡Ù… Ø§Ø³Øª:** Ø§Ø³Ú©Ù†â€ŒÙ‡Ø§ÛŒ ÙˆØ§Ø¶Ø­ Ù†ØªØ§ÛŒØ¬ Ø¨Ù‡ØªØ±ÛŒ Ø¯Ø§Ø±Ù†Ø¯
                
                ---
                
                #### ğŸ¯ ÙˆÛŒÚ˜Ú¯ÛŒâ€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø±Ø³ÛŒ Ø´Ø¯Ù‡:
                
                âœ”ï¸ **Ø§Ø·Ù„Ø§Ø¹Ø§Øª Ø´Ø±Ú©Øª** - Ù†Ø§Ù…ØŒ Ø¯ÙˆØ±Ù‡ Ù…Ø§Ù„ÛŒ  
                âœ”ï¸ **Ø§Ø¸Ù‡Ø§Ø±Ù†Ø¸Ø± Ø­Ø³Ø§Ø¨Ø±Ø³** - Ù†ÙˆØ¹ Ùˆ Ø³Ø·Ø­ Ø±ÛŒØ³Ú©  
                âœ”ï¸ **Ø±ÛŒØ³Ú©â€ŒÙ‡Ø§ÛŒ Ø¨Ø±Ø¬Ø³ØªÙ‡** - ØªØ´Ø®ÛŒØµ Ùˆ Ø¯Ø³ØªÙ‡â€ŒØ¨Ù†Ø¯ÛŒ  
                âœ”ï¸ **ØªØ®Ù„ÙØ§Øª Ù‚Ø§Ù†ÙˆÙ†ÛŒ** - Ø´Ù†Ø§Ø³Ø§ÛŒÛŒØŒ ØªØ­Ù„ÛŒÙ„ Ùˆ Ø¯Ø³ØªÙ‡ Ø¨Ù†Ø¯ÛŒ  
                âœ”ï¸ **Ù…ÙˆØ¶ÙˆØ¹Ø§Øª Ú©Ù„ÛŒØ¯ÛŒ** - Ø´Ù†Ø§Ø³Ø§ÛŒÛŒ Ùˆ ØªØ­Ù„ÛŒÙ„  
                            
                </div>
                """, unsafe_allow_html=True)

            uploaded_files = create_file_upload_section()
            if uploaded_files:
                create_processing_section(uploaded_files)

        with tab2:
            if st.session_state.results:
                create_results_section(st.session_state.results)
            else:
                st.info("Ù‡Ù†ÙˆØ² ÙØ§ÛŒÙ„ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")
        
        with tab3:
            if st.session_state.results:
                create_stats_section(st.session_state.results)
            else:
                st.info("Ù‡Ù†ÙˆØ² ÙØ§ÛŒÙ„ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")
        
        with tab4:
            with st.expander("ğŸ“‰ Ø±Ø§Ù‡Ù†Ù…Ø§ÛŒ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ùˆ ØªØ±Ù†Ø¯", expanded=False):
                st.markdown("""
                    <div style="text-align: right; direction: rtl; padding: 1rem; border-radius: 12px;">

                    #### **âš ï¸ ØªÙˆØ¬Ù‡:**

                    - Ø§Ú¯Ø± ØµÙˆØ±Øªâ€ŒÙ‡Ø§ÛŒ Ù…Ø§Ù„ÛŒ Ø´Ù…Ø§ Ù…ØªØ¹Ù„Ù‚ Ø¨Ù‡ **ÛŒÚ© Ø´Ø±Ú©Øª** Ø¨Ø§Ø´Ù†Ø¯ØŒ Ù†Ù…ÙˆØ¯Ø§Ø±Ù‡Ø§ Ø¨Ù‡ ØµÙˆØ±Øª Ø®ÙˆØ¯Ú©Ø§Ø± Ø±Ø³Ù… Ù…ÛŒâ€ŒØ´ÙˆÙ†Ø¯.
                    - Ø§Ú¯Ø± **Ú†Ù†Ø¯ Ø´Ø±Ú©Øª Ù…Ø®ØªÙ„Ù** Ø¯Ø§Ø±ÛŒØ¯ØŒ Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒØ¯ Ø´Ø±Ú©Øª Ù…ÙˆØ±Ø¯ Ù†Ø¸Ø± Ø±Ø§ Ø¨Ø±Ø§ÛŒ Ø±Ø³Ù… Ù†Ù…ÙˆØ¯Ø§Ø± Ø§Ù†ØªØ®Ø§Ø¨ Ú©Ù†ÛŒØ¯.
                    - Ø§Ú¯Ø± Ù†Ø§Ù… ÛŒÚ© Ø´Ø±Ú©Øª Ø¯Ø± Ø³Ø§Ù„â€ŒÙ‡Ø§ÛŒ Ù…Ø®ØªÙ„Ù **Ú©Ø§Ù…Ù„Ø§Ù‹ ØªØºÛŒÛŒØ± Ú©Ø±Ø¯Ù‡ Ø¨Ø§Ø´Ø¯**ØŒ Ù…ÛŒâ€ŒØªÙˆØ§Ù†ÛŒØ¯ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ Ø±Ø§ Ø¨Ù‡â€ŒØµÙˆØ±Øª Ø¯Ø³ØªÛŒ Ø§Ø¯ØºØ§Ù… Ú©Ù†ÛŒØ¯.
                    - Ø¯Ø± ØµÙˆØ±Øª Ø§Ø¯ØºØ§Ù…ØŒ ØªÙ…Ø§Ù… Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ø¹Ø¶Ùˆ ÛŒÚ© Ú¯Ø±ÙˆÙ‡ØŒ Ø¨Ø§ **Ù†Ø§Ù… Ø´Ø±Ú©Øª Ø§ØµÙ„ÛŒ Ú¯Ø±ÙˆÙ‡** Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ù…ÛŒâ€ŒØ´ÙˆÙ†Ø¯ Ùˆ Ø´Ø±Ú©Øªâ€ŒÙ‡Ø§ÛŒ Ø²ÛŒØ±Ú¯Ø±ÙˆÙ‡ Ø¨Ù‡â€ŒØµÙˆØ±Øª Ø¬Ø¯Ø§Ú¯Ø§Ù†Ù‡ Ù†Ù…Ø§ÛŒØ´ Ø¯Ø§Ø¯Ù‡ Ù†Ù…ÛŒâ€ŒØ´ÙˆÙ†Ø¯.

                    </div>
                """, unsafe_allow_html=True)

            
            if st.session_state.results:
                # âœ… Ø§Ø³ØªÙØ§Ø¯Ù‡ Ø§Ø² ØªØ§Ø¨Ø¹ Ø¬Ø¯ÛŒØ¯ Ø¨Ù‡ Ø¬Ø§ÛŒ ØªØ§Ø¨Ø¹ Ù‚Ø¯ÛŒÙ…ÛŒ
                create_filtered_charts_section(st.session_state.results)
            else:
                st.info("Ù‡Ù†ÙˆØ² ÙØ§ÛŒÙ„ÛŒ Ù¾Ø±Ø¯Ø§Ø²Ø´ Ù†Ø´Ø¯Ù‡ Ø§Ø³Øª.")


    if __name__ == "__main__":
        main()
